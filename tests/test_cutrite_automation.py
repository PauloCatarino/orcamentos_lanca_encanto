import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook

from Martelo_Orcamentos_V2.app.services import cutrite_automation


class CutRiteAutomationTests(unittest.TestCase):
    class _Rect:
        def __init__(self, left=0, top=0, right=1200, bottom=900):
            self.left = left
            self.top = top
            self.right = right
            self.bottom = bottom

        def width(self):
            return self.right - self.left

        def height(self):
            return self.bottom - self.top

    class _ElementInfo:
        def __init__(self, name="", control_type="Button", **extra):
            self.name = name
            self.control_type = control_type
            for key, value in extra.items():
                setattr(self, key, value)

    class _Element:
        def __init__(self, name="", control_type="Button", children=None, **extra):
            self.element_info = CutRiteAutomationTests._ElementInfo(name, control_type, **extra)
            self._children = list(children or [])

        def descendants(self, control_type=None):
            items = []
            for child in self._children:
                if control_type is None or child.element_info.control_type == control_type:
                    items.append(child)
                items.extend(child.descendants(control_type=control_type))
            return items

        def children(self):
            return list(self._children)

    def test_normalize_cutrite_ui_text_removes_accents(self):
        self.assertEqual(
            cutrite_automation._normalize_cutrite_ui_text("Importação - Peças"),
            "importacao - pecas",
        )

    def test_sanitize_cutrite_plan_name_replaces_invalid_chars(self):
        value = cutrite_automation.sanitize_cutrite_plan_name(' 0417:01/01*JF?VIVA ')
        self.assertEqual(value, "0417_01_01_JF_VIVA")

    def test_find_lista_material_workbook_prefers_exact_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            other = folder / "Lista_Material_OUTRO.xlsm"
            exact = folder / "Lista_Material_0417_01_26_JF_VIVA.xlsm"
            other.write_text("x", encoding="utf-8")
            exact.write_text("x", encoding="utf-8")

            result = cutrite_automation.find_lista_material_workbook(
                folder,
                nome_enc_imos="0417_01_26_JF_VIVA",
            )

        self.assertEqual(result, exact)

    def test_find_lista_material_workbook_uses_single_prefix_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "Lista_Material_TESTE.xlsx"
            source.write_text("x", encoding="utf-8")

            result = cutrite_automation.find_lista_material_workbook(folder)

        self.assertEqual(result, source)

    def test_build_cutrite_import_row_preserves_listagem_cut_rite_layout(self):
        row = [
            "Costa",
            "AGL_MLM_LINHO_CANCUN_19MM",
            2380,
            589,
            1,
            "N",
            None,
            "JF_VIVA",
            "0417",
            "0417_01_01_JF_VIVA",
            "RP_01(a)",
            "CNC",
            19.2,
            "111:111:000:000",
            "PVC_1.0_LINHO",
            None,
            "PVC_0.4_LINHO",
            "PVC_0.4_LINHO",
            15,
            10125,
            None,
            None,
            None,
            19,
            19,
        ]

        mapped = cutrite_automation.build_cutrite_import_row(row)

        self.assertEqual(mapped[0], "Costa")
        self.assertEqual(mapped[1], "AGL_MLM_LINHO_CANCUN_19MM")
        self.assertEqual(mapped[2], 2380)
        self.assertEqual(mapped[3], 589)
        self.assertEqual(mapped[4], 1)
        self.assertEqual(mapped[5], "N")
        self.assertEqual(mapped[9], "0417_01_01_JF_VIVA")
        self.assertEqual(mapped[10], "RP_01(a)")
        self.assertEqual(mapped[14], "PVC_1.0_LINHO")
        self.assertEqual(mapped[18], 15)
        self.assertEqual(mapped[19], 10125)
        self.assertEqual(mapped[23], 19)
        self.assertEqual(mapped[24], 19)

    def test_build_cutrite_import_headers_match_excel_layout(self):
        self.assertEqual(
            cutrite_automation.build_cutrite_import_headers(),
            [
                "Descricao",
                "Material",
                "Comp",
                "Larg",
                "Qt",
                "Veio",
                "Orla",
                "Cliente",
                "Ref_Cliente",
                "Processo",
                "Artigo",
                "Notas",
                "Esp",
                "Grafico Orlas",
                "Orla ESQ",
                "Orla DIR",
                "Orla CIMA",
                "Orla BAIXO",
                "ID",
                "CNC_1",
                "CNC_2",
                "+comp",
                "+Larg",
            ],
        )

    def test_load_cutrite_source_table_reads_xlsm_directly_and_forces_grafico_orlas(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "Lista_Material_0395_01_26_JF_VIVA.xlsm"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = cutrite_automation.LISTAGEM_CUT_RITE_SHEET
            headers = cutrite_automation.build_cutrite_import_headers()
            for index, header in enumerate(headers, start=1):
                worksheet.cell(row=2, column=index).value = header
            values = [
                "Costa",
                "AGL_MLM_LINHO_CANCUN_19MM",
                2380,
                589,
                1,
                "N",
                "",
                "JF_VIVA",
                "2603018",
                "0395_01_26",
                "RP_01(a)",
                "Nota teste",
                19.2,
                "000:000:000:000",
                "PVC_1.0_LINHO",
                "PVC_0.4_LINHO",
                "PVC_0.4_LINHO",
                "PVC_0.4_LINHO",
                15,
                10125,
                "",
                "",
                "",
            ]
            for index, value in enumerate(values, start=1):
                worksheet.cell(row=3, column=index).value = value
            workbook.save(source)

            with patch.object(
                cutrite_automation,
                "_load_cutrite_source_table_from_excel_macro",
                side_effect=AssertionError("macro nao devia ser chamada"),
            ):
                loaded_headers, rows, macro_text = cutrite_automation._load_cutrite_source_table(source)

        self.assertEqual(loaded_headers[:5], headers[:5])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][13], cutrite_automation.CUTRITE_GRAFICO_ORLAS_VALUE)
        self.assertIsNone(macro_text)

    def test_load_cutrite_source_table_falls_back_to_macro_when_file_locked(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "Lista_Material_0395_01_26_JF_VIVA.xlsm"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = cutrite_automation.LISTAGEM_CUT_RITE_SHEET
            workbook.save(source)

            macro_text = "A\tB\nC\tD"
            macro_rows = [["A", "B"], ["C", "D"]]

            with patch.object(cutrite_automation, "load_workbook", side_effect=PermissionError("locked")):
                with patch.object(
                    cutrite_automation,
                    "_load_cutrite_source_table_from_excel_macro",
                    return_value=(macro_text, macro_rows),
                ):
                    headers, rows, returned_macro_text = cutrite_automation._load_cutrite_source_table(source)

            self.assertEqual(returned_macro_text, macro_text)
            self.assertEqual(rows, macro_rows)
            self.assertEqual(headers, cutrite_automation.build_cutrite_import_headers())

    def test_build_cutrite_import_workbook_no_longer_creates_intermediate_xls(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "Lista_Material_0395_01_26_JF_VIVA.xlsm"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = cutrite_automation.LISTAGEM_CUT_RITE_SHEET
            for index, header in enumerate(cutrite_automation.build_cutrite_import_headers(), start=1):
                worksheet.cell(row=2, column=index).value = header
            worksheet.cell(row=3, column=1).value = "Costa"
            worksheet.cell(row=3, column=2).value = "AGL"
            worksheet.cell(row=3, column=3).value = 2379.4
            worksheet.cell(row=3, column=4).value = 950
            worksheet.cell(row=3, column=5).value = 1
            workbook.save(source)

            context = cutrite_automation.CutRiteImportContext(
                processo=SimpleNamespace(id=1),
                folder_path=root,
                source_workbook_path=source,
                import_workbook_path=root / "0395_01_01_26_JF_VIVA.xls",
                cutrite_input_path=root / "IMPORT" / "0395_01_01_26_JF_VIVA.xls",
                plan_name="0395_01_01_26_JF_VIVA",
                cutrite_exe_path=Path(r"C:\V12\V12.exe"),
                import_exe_path=Path(r"C:\V12\Import.exe"),
                cutrite_root=Path(r"C:\V12"),
                cutrite_profile_dir=Path(r"\\SERVER_LE\Homag_iX\Cutrite\V12-Data\Paulo_Catarino"),
                cutrite_workdir=Path(r"C:\V12\_WORK\USER1"),
                cutrite_data_dir=root / "DATA",
                cutrite_target_data_dir=root / "DATA",
            )

            import_path, rows, paste_text = cutrite_automation.build_cutrite_import_workbook(context)

        self.assertEqual(import_path, context.import_workbook_path)
        self.assertFalse(import_path.exists())
        self.assertEqual(len(rows), 1)
        self.assertTrue(paste_text)

    def test_select_cutrite_export_columns_excludes_excel_only_columns(self):
        row = [
            "Descricao",
            "Material",
            2379.4,
            950,
            1,
            "N",
            "",
            "JF_VIVA",
            "2603018",
            "0417_01_26",
            "RP_01(a)",
            "Nota",
            19.2,
            "111:111:000:000",
            "PVC_1.0_LINHO",
            "PVC_0.4_LINHO",
            "PVC_0.4_LINHO",
            "PVC_0.4_LINHO",
            1,
            10143,
            10243,
            0,
            0,
            19,
            19,
            "IGNORAR",
        ]
        self.assertEqual(
            cutrite_automation._select_cutrite_export_columns(row),
            row[:23],
        )

    def test_parse_cutrite_clipboard_rows_preserves_cut_rite_import_layout(self):
        clipboard_text = (
            "Costa\tAGL_MLM_LINHO_CANCUN_10MM\t2379.4\t950\t1\tN\t\tJF_VIVA\t2603018\t0417_01_26\tRP_01(a)\t"
            "\t10.4\t111:111:000:000\tPVC_1.0_LINHO\tPVC_0.4_LINHO\tPVC_0.4_LINHO\tPVC_0.4_LINHO\t13\t"
            "10143\t\t\t\r\n"
        )

        rows = cutrite_automation._parse_cutrite_clipboard_rows(clipboard_text)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "Costa")
        self.assertEqual(rows[0][2], 2379.4)
        self.assertEqual(rows[0][4], 1)
        self.assertEqual(rows[0][11], "")
        self.assertEqual(rows[0][14], "PVC_1.0_LINHO")
        self.assertEqual(rows[0][15], "PVC_0.4_LINHO")
        self.assertEqual(rows[0][18], 13)
        self.assertEqual(rows[0][19], 10143)
        self.assertEqual(len(rows[0]), 23)

    def test_build_cutrite_paste_text_formats_rows_for_clipboard(self):
        text = cutrite_automation._build_cutrite_paste_text(
            [
                ["Costa", "AGL", 2379.4, 950.0, 1.0, "N", "", "JF_VIVA"],
                ["Tampo", "AGL", 911.6, 586.0, 1.0, "N", "", "JF_VIVA"],
            ]
        )

        lines = text.split("\r\n")
        self.assertEqual(lines[0], "Costa\tAGL\t2379.4\t950\t1\tN\t\tJF_VIVA")
        self.assertEqual(lines[1], "Tampo\tAGL\t911.6\t586\t1\tN\t\tJF_VIVA")

    def test_normalize_cutrite_macro_paste_text_preserves_ready_clipboard_rows(self):
        text = cutrite_automation._normalize_cutrite_macro_paste_text("A\tB\tC\r\n\r\n1\t2\t3\n")

        self.assertEqual(text, "A\tB\tC\r\n1\t2\t3")

    def test_set_cutrite_clipboard_text_falls_back_to_qt_when_win32clipboard_is_missing(self):
        with patch.object(
            cutrite_automation,
            "_set_cutrite_clipboard_text_win32",
            side_effect=ModuleNotFoundError("win32clipboard"),
        ) as win32_mock, patch.object(
            cutrite_automation,
            "_set_cutrite_clipboard_text_qt",
            return_value=None,
        ) as qt_mock:
            cutrite_automation._set_cutrite_clipboard_text("A\tB")

        win32_mock.assert_called_once()
        qt_mock.assert_called_once_with("A\tB")

    def test_set_cutrite_clipboard_text_raises_when_all_clipboard_backends_fail(self):
        with patch.object(
            cutrite_automation,
            "_set_cutrite_clipboard_text_win32",
            side_effect=ModuleNotFoundError("win32clipboard"),
        ), patch.object(
            cutrite_automation,
            "_set_cutrite_clipboard_text_qt",
            side_effect=RuntimeError("qt clipboard indisponivel"),
        ):
            with self.assertRaisesRegex(RuntimeError, "Nao foi possivel preparar o clipboard"):
                cutrite_automation._set_cutrite_clipboard_text("A\tB")

    def test_close_cutrite_desenho_window_closes_detected_window(self):
        class WindowStub:
            def __init__(self):
                self.closed = 0

            def close(self):
                self.closed += 1

        window = WindowStub()
        finder_results = [window, None]

        with patch.object(
            cutrite_automation,
            "_find_cutrite_descendant_window",
            side_effect=lambda *args, **kwargs: finder_results.pop(0) if finder_results else None,
        ), patch.object(cutrite_automation.time, "sleep", return_value=None):
            result = cutrite_automation._close_cutrite_desenho_window(object(), object(), timeout_seconds=1)

        self.assertTrue(result)
        self.assertEqual(window.closed, 1)

    def test_click_cutrite_main_import_button_prefers_named_button(self):
        target = self._Element(
            name="Importacao - Pecas",
            control_type="Button",
            help_text="Importacao - Pecas",
        )
        main_window = self._Element(control_type="Window", children=[target])

        with patch.object(cutrite_automation, "_click_cutrite_element") as click_mock:
            result = cutrite_automation._click_cutrite_main_import_button(main_window, mouse=object())

        self.assertTrue(result)
        click_mock.assert_called_once()

    def test_find_cutrite_parts_buttons_prefers_lista_de_pecas_tooltip(self):
        target = self._Element(
            name="Pecas",
            control_type="Button",
            help_text="Lista de pecas",
        )
        distractor = self._Element(
            name="Pecas",
            control_type="Button",
            help_text="Importacao - Pecas",
        )
        main_window = self._Element(control_type="Window", children=[distractor, target])

        buttons = cutrite_automation._find_cutrite_parts_buttons(main_window)

        self.assertEqual(buttons[0], target)

    def test_find_cutrite_window_falls_back_to_desktop_window(self):
        desktop_window = self._Element(name="Lista de pecas", control_type="Window")

        class DesktopStub:
            def windows(self):
                return [desktop_window]

        result = cutrite_automation._find_cutrite_window(
            DesktopStub(),
            self._Element(name="Principal", control_type="Window"),
            cutrite_automation.CUTRITE_PARTS_WINDOW_NAME,
        )

        self.assertEqual(result, desktop_window)

    def test_submit_cutrite_import_window_uses_toolbar_when_ok_absent(self):
        window = self._Element(control_type="Window")
        with patch.object(
            cutrite_automation,
            "_click_cutrite_named_button_if_present",
            return_value=False,
        ), patch.object(cutrite_automation, "_click_cutrite_toolbar_button") as toolbar_mock:
            cutrite_automation._submit_cutrite_import_window(window, mouse=object())

        toolbar_mock.assert_called_once_with(
            window,
            cutrite_automation.CUTRITE_IMPORT_WINDOW_TOOLBAR_IMPORT_INDEX,
            unittest.mock.ANY,
        )

    def test_save_cutrite_cutlist_uses_toolbar_before_coordinate_fallback(self):
        cutlist_window = self._Element(name="Lista de pecas", control_type="Window")
        main_window = self._Element(name="CutRite", control_type="Window")
        save_window = self._Element(name="Guardar em baixo", control_type="Window")
        desktop = object()
        keyboard = unittest.mock.Mock()

        with patch.object(cutrite_automation, "_activate_cutrite_window"), patch.object(
            cutrite_automation, "_sleep_cutrite"
        ), patch.object(
            cutrite_automation, "_find_cutrite_save_button_candidates", return_value=[]
        ), patch.object(
            cutrite_automation, "_click_cutrite_save_fallback"
        ) as fallback_mock, patch.object(
            cutrite_automation, "_click_cutrite_save_via_toolbar"
        ) as toolbar_mock, patch.object(
            cutrite_automation, "_wait_for_cutrite_window", return_value=save_window
        ) as wait_mock:
            result = cutrite_automation._save_cutrite_cutlist(
                cutlist_window,
                desktop,
                main_window,
                mouse=object(),
                keyboard=keyboard,
                timeout_seconds=10,
            )

        self.assertEqual(result, save_window)
        toolbar_mock.assert_called_once_with(cutlist_window, unittest.mock.ANY)
        fallback_mock.assert_not_called()
        wait_mock.assert_called_once()

    def test_focus_cutrite_first_data_cell_targets_description_on_row_one(self):
        description_header = unittest.mock.Mock()
        description_header.element_info = self._ElementInfo(name="Descricao", control_type="Text")
        description_header.rectangle.return_value = self._Rect(left=52, top=99, right=248, bottom=122)

        row_one = unittest.mock.Mock()
        row_one.element_info = self._ElementInfo(name="1.", control_type="Text")
        row_one.rectangle.return_value = self._Rect(left=3, top=141, right=45, bottom=161)

        parts_window = unittest.mock.Mock()
        parts_window.descendants.return_value = [description_header, row_one]
        mouse = unittest.mock.Mock()
        keyboard = unittest.mock.Mock()

        with patch.object(cutrite_automation, "_activate_cutrite_window"), patch.object(
            cutrite_automation, "_sleep_cutrite"
        ):
            cutrite_automation._focus_cutrite_first_data_cell(parts_window, mouse, keyboard)

        first_click = mouse.click.call_args_list[0].kwargs["coords"]
        second_click = mouse.click.call_args_list[1].kwargs["coords"]
        self.assertEqual(first_click, second_click)
        self.assertEqual(first_click, (80, 151))
        keyboard.send_keys.assert_not_called()

    def test_focus_cutrite_first_data_cell_moves_down_after_global_fallback(self):
        description_header = unittest.mock.Mock()
        description_header.element_info = self._ElementInfo(name="Descricao", control_type="Text")
        description_header.rectangle.return_value = self._Rect(left=52, top=99, right=248, bottom=122)

        global_row = unittest.mock.Mock()
        global_row.element_info = self._ElementInfo(name="Global", control_type="Text")
        global_row.rectangle.return_value = self._Rect(left=2, top=120, right=44, bottom=140)

        parts_window = unittest.mock.Mock()
        parts_window.descendants.return_value = [description_header, global_row]
        mouse = unittest.mock.Mock()
        keyboard = unittest.mock.Mock()

        with patch.object(cutrite_automation, "_activate_cutrite_window"), patch.object(
            cutrite_automation, "_sleep_cutrite"
        ):
            cutrite_automation._focus_cutrite_first_data_cell(parts_window, mouse, keyboard)

        keyboard.send_keys.assert_called_once_with("{DOWN}", pause=0.02)
    def test_save_cutrite_parts_list_confirms_without_forcing_maximize(self):
        parts_window = self._Element(name="Lista de pecas", control_type="Window")
        main_window = self._Element(name="CutRite", control_type="Window")
        save_window = self._Element(name="Guardar em baixo", control_type="Window")
        desktop = object()
        keyboard = unittest.mock.Mock()

        with patch.object(cutrite_automation, "_activate_cutrite_window"), patch.object(
            cutrite_automation, "_sleep_cutrite"
        ), patch.object(
            cutrite_automation, "_save_cutrite_cutlist", return_value=save_window
        ) as save_mock, patch.object(
            cutrite_automation, "_click_cutrite_named_button"
        ) as click_ok_mock:
            cutrite_automation._save_cutrite_parts_list(
                parts_window,
                desktop,
                main_window,
                mouse=object(),
                keyboard=keyboard,
                timeout_seconds=12,
            )

        save_mock.assert_called_once_with(
            parts_window,
            desktop,
            main_window,
            unittest.mock.ANY,
            keyboard,
            timeout_seconds=12,
        )
        click_ok_mock.assert_called_once_with(save_window, "ok", unittest.mock.ANY)

    def test_save_cutrite_cutlist_accepts_save_window_found_by_title(self):
        cutlist_window = self._Element(name="Lista de pecas", control_type="Window")
        main_window = self._Element(name="CutRite", control_type="Window")
        save_window = self._Element(name="Guardar em baixo", control_type="Window")
        desktop = object()
        keyboard = unittest.mock.Mock()

        with patch.object(cutrite_automation, "_activate_cutrite_window"), patch.object(
            cutrite_automation, "_sleep_cutrite"
        ), patch.object(
            cutrite_automation, "_find_cutrite_save_button_candidates", return_value=[]
        ), patch.object(
            cutrite_automation, "_click_cutrite_save_fallback"
        ) as fallback_mock, patch.object(
            cutrite_automation,
            "_wait_for_cutrite_window",
            return_value=save_window,
        ) as wait_mock:
            result = cutrite_automation._save_cutrite_cutlist(
                cutlist_window,
                desktop,
                main_window,
                mouse=object(),
                keyboard=keyboard,
                timeout_seconds=10,
            )

        self.assertEqual(result, save_window)
        fallback_mock.assert_called()
        wait_mock.assert_called_once()

    def test_maximize_cutrite_window_returns_when_window_is_already_maximized(self):
        window = self._Element(name="CutRite", control_type="Window")
        with patch.object(
            cutrite_automation, "_is_cutrite_window_maximized", return_value=True
        ), patch.object(
            cutrite_automation, "_looks_cutrite_window_maximized", return_value=False
        ), patch.object(
            cutrite_automation, "_activate_cutrite_window"
        ) as activate_mock:
            cutrite_automation._maximize_cutrite_window(window, window_label="principal do CUT-RITE")

        activate_mock.assert_not_called()

    def test_maximize_cutrite_window_accepts_monitor_geometry_fallback(self):
        window = self._Element(name="CutRite", control_type="Window")
        with patch.object(
            cutrite_automation, "_is_cutrite_window_maximized", return_value=False
        ), patch.object(
            cutrite_automation, "_looks_cutrite_window_maximized", return_value=True
        ), patch.object(
            cutrite_automation, "_activate_cutrite_window"
        ) as activate_mock:
            cutrite_automation._maximize_cutrite_window(window, window_label="principal do CUT-RITE")

        activate_mock.assert_not_called()

    def test_maximize_cutrite_window_returns_false_when_confirmation_fails(self):
        window = self._Element(name="CutRite", control_type="Window")
        with patch.object(
            cutrite_automation, "_is_cutrite_window_maximized", return_value=False
        ), patch.object(
            cutrite_automation, "_looks_cutrite_window_maximized", return_value=False
        ), patch.object(
            cutrite_automation, "_activate_cutrite_window"
        ), patch.object(
            cutrite_automation, "_get_cutrite_window_handle", return_value=0
        ), patch.object(
            cutrite_automation, "_sleep_cutrite"
        ):
            result = cutrite_automation._maximize_cutrite_window(window, window_label="principal do CUT-RITE")

        self.assertFalse(result)

    def test_prepare_cutrite_import_builds_expected_paths(self):
        processo = SimpleNamespace(id=7, codigo_processo="26.0417_01_01")
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "Lista_Material_0417_01_26_JF_VIVA.xlsm"
            source.write_text("x", encoding="utf-8")

            class SessionStub:
                def get(self, model, item_id):
                    self.model = model
                    self.item_id = item_id
                    return processo

            session = SessionStub()

            with patch.object(
                cutrite_automation,
                "resolve_cutrite_exe_path",
                return_value=r"C:\V12\V12.exe",
            ), patch.object(
                cutrite_automation,
                "resolve_cutrite_workdir",
                return_value=Path(r"C:\V12\_WORK\USER1"),
            ), patch.object(
                cutrite_automation,
                "resolve_cutrite_input_dir",
                return_value=Path(r"C:\V12\_WORK\IMPORT"),
            ):
                context = cutrite_automation.prepare_cutrite_import(
                    session,
                    current_id=7,
                    pasta_servidor=str(folder),
                    nome_plano_cut_rite="0417_01_01_JF_VIVA",
                    nome_enc_imos="0417_01_26_JF_VIVA",
                )

        self.assertEqual(context.processo, processo)
        self.assertEqual(context.source_workbook_path, source)
        self.assertEqual(context.import_workbook_path, folder / "0417_01_01_JF_VIVA.xls")
        self.assertEqual(context.import_exe_path, Path(r"C:\V12\Import.exe"))

    def test_prepare_cutrite_import_requires_pasta_servidor(self):
        processo = SimpleNamespace(id=7, codigo_processo="26.0417_01_01")

        class SessionStub:
            def get(self, model, item_id):
                return processo

        with self.assertRaisesRegex(ValueError, "Pasta Servidor em falta"):
            cutrite_automation.prepare_cutrite_import(
                SessionStub(),
                current_id=7,
                pasta_servidor="   ",
                nome_plano_cut_rite="0417_01_01_JF_VIVA",
                nome_enc_imos="0417_01_26_JF_VIVA",
            )

    def test_resolve_cutrite_workdir_requires_explicit_choice_when_multiple_users_exist(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            work = root / "_WORK"
            user1 = work / "USER1"
            user2 = work / "USER2"
            user1.mkdir(parents=True)
            user2.mkdir(parents=True)
            (user1 / "systemv12.ctl").write_text("SYSIMPPATH,C:\\TEMP\\IMPORT\\\n", encoding="utf-8")
            (user2 / "systemv12.ctl").write_text("SYSIMPPATH,C:\\TEMP\\IMPORT\\\n", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "Pasta Trabalho CUT-RITE"):
                cutrite_automation.resolve_cutrite_workdir(root)

    def test_resolve_cutrite_workdir_uses_configured_workdir(self):
        with tempfile.TemporaryDirectory() as tmp:
            configured = Path(tmp) / "USER2"
            configured.mkdir(parents=True)
            (configured / "systemv12.ctl").write_text("SYSIMPPATH,C:\\TEMP\\IMPORT\\\n", encoding="utf-8")

            class DbStub:
                pass

            with patch.object(
                cutrite_automation,
                "resolve_configured_cutrite_workdir",
                return_value=configured,
            ):
                result = cutrite_automation.resolve_cutrite_workdir(Path(tmp), db=DbStub())

        self.assertEqual(result, configured)

    def test_resolve_cutrite_data_dir_reads_sysdatapath(self):
        with tempfile.TemporaryDirectory() as tmp:
            workdir = Path(tmp)
            (workdir / "systemv12.ctl").write_text(
                "SYSDATAPATH,\\\\SERVER_LE\\Homag_iX\\Cutrite\\V12-Data\\Data\\\n",
                encoding="utf-8",
            )

            result = cutrite_automation.resolve_cutrite_data_dir(workdir)

        self.assertEqual(result, Path(r"\\SERVER_LE\Homag_iX\Cutrite\V12-Data\Data"))

    def test_normalize_cutrite_path_inputs_swaps_profile_and_data_when_inverted(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "Data"
            profile_dir = root / "Paulo_Catarino"
            data_dir.mkdir()
            profile_dir.mkdir()
            (profile_dir / "systemv12.ctl").write_text("SYSDATAPATH,C:\\TEMP\\DATA\\\n", encoding="utf-8")

            workdir, data, warning = cutrite_automation.normalize_cutrite_path_inputs(
                str(data_dir),
                str(profile_dir),
            )

        self.assertEqual(workdir, profile_dir)
        self.assertEqual(data, data_dir)
        self.assertIn("invertidas", warning)

    def test_sync_cutrite_generated_data_copies_to_target_dir(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_data = root / "source"
            target_data = root / "target"
            source_data.mkdir()
            target_data.mkdir()
            for suffix in (".mpd", ".ctt", ".btc", ".prl"):
                (source_data / f"0417_01_01_26_JF_VIVA{suffix}").write_text("x", encoding="utf-8")

            context = cutrite_automation.CutRiteImportContext(
                processo=SimpleNamespace(id=1),
                folder_path=root,
                source_workbook_path=root / "Lista_Material.xlsm",
                import_workbook_path=root / "0417_01_01_26_JF_VIVA.xls",
                cutrite_input_path=root / "IMPORT" / "0417_01_01_26_JF_VIVA.xls",
                plan_name="0417_01_01_26_JF_VIVA",
                cutrite_exe_path=Path(r"C:\V12\V12.exe"),
                import_exe_path=Path(r"C:\V12\Import.exe"),
                cutrite_root=Path(r"C:\V12"),
                cutrite_profile_dir=Path(r"\\SERVER_LE\Homag_iX\Cutrite\V12-Data\Paulo_Catarino"),
                cutrite_workdir=Path(r"C:\V12\_WORK\USER1"),
                cutrite_data_dir=source_data,
                cutrite_target_data_dir=target_data,
            )

            synced = cutrite_automation._sync_cutrite_generated_data(context)

        self.assertEqual(len(synced), 4)
        self.assertTrue(all(path.parent == target_data for path in synced))

    def test_sync_cutrite_generated_data_skips_copy_when_dirs_are_same_location(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_data = root / "source"
            source_data.mkdir()
            expected = source_data / "0417_01_01_26_JF_VIVA.btc"
            expected.write_text("x", encoding="utf-8")

            context = cutrite_automation.CutRiteImportContext(
                processo=SimpleNamespace(id=1),
                folder_path=root,
                source_workbook_path=root / "Lista_Material.xlsm",
                import_workbook_path=root / "0417_01_01_26_JF_VIVA.xls",
                cutrite_input_path=root / "IMPORT" / "0417_01_01_26_JF_VIVA.xls",
                plan_name="0417_01_01_26_JF_VIVA",
                cutrite_exe_path=Path(r"C:\V12\V12.exe"),
                import_exe_path=Path(r"C:\V12\Import.exe"),
                cutrite_root=Path(r"C:\V12"),
                cutrite_profile_dir=Path(r"\\SERVER_LE\Homag_iX\Cutrite\V12-Data\Paulo_Catarino"),
                cutrite_workdir=Path(r"C:\V12\_WORK\USER1"),
                cutrite_data_dir=source_data,
                cutrite_target_data_dir=Path(r"\\SERVER_LE\Homag_iX\Cutrite\V12-Data\Data"),
            )

            with patch.object(
                cutrite_automation,
                "_paths_refer_to_same_location",
                side_effect=lambda a, b: str(a).endswith("source") or a == expected,
            ):
                synced = cutrite_automation._sync_cutrite_generated_data(context)

        self.assertEqual(synced, [expected])

    def test_copy_file_with_retry_retries_winerror_32(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "a.txt"
            target = root / "b.txt"
            source.write_text("ok", encoding="utf-8")
            calls = {"count": 0}

            def fake_copy2(src, dst):
                calls["count"] += 1
                if calls["count"] == 1:
                    err = PermissionError("locked")
                    err.winerror = 32
                    raise err
                Path(dst).write_text(Path(src).read_text(encoding="utf-8"), encoding="utf-8")
                return str(dst)

            with patch.object(cutrite_automation.shutil, "copy2", side_effect=fake_copy2), patch.object(
                cutrite_automation.time,
                "sleep",
                return_value=None,
            ):
                cutrite_automation._copy_file_with_retry(source, target, retries=3, delay_seconds=0)

            copied_text = target.read_text(encoding="utf-8")

        self.assertEqual(calls["count"], 2)
        self.assertEqual(copied_text, "ok")

    def test_prepare_cutrite_runtime_workdir_creates_local_shadow_for_remote_profile(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cutrite_root = root / "V12"
            template_dir = cutrite_root / "_WORK" / "USER1"
            remote_profile = root / "Paulo_Catarino"
            template_dir.mkdir(parents=True)
            remote_profile.mkdir(parents=True)
            (template_dir / "Excel-XLS.upx").write_text("xls", encoding="utf-8")
            (template_dir / "Excel-CSV.upx").write_text("csv", encoding="utf-8")
            (remote_profile / "systemv12.ctl").write_text("SYSDATAPATH,\\\\SERVER\\Data\\\n", encoding="utf-8")
            (remote_profile / "v12defs.ctl").write_text(
                "ImportPartFormat=0\nImportPartParams=\nShowImportFileDialog=0\n",
                encoding="utf-8",
            )

            with patch.object(cutrite_automation, "_is_local_path", return_value=False):
                runtime_dir = cutrite_automation.prepare_cutrite_runtime_workdir(cutrite_root, remote_profile)

            text = (runtime_dir / "v12defs.ctl").read_text(encoding="utf-8")

        self.assertTrue(runtime_dir.is_dir())
        self.assertNotEqual(runtime_dir, remote_profile)
        self.assertTrue((runtime_dir / "systemv12.ctl").is_file())
        self.assertTrue((runtime_dir / "Excel-XLS.upx").is_file())
        self.assertIn("ImportPartFormat=10", text)
        self.assertIn("ImportPartParams=Excel-XLS", text)

    def test_prepare_cutrite_runtime_workdir_uses_unique_directory_per_run(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cutrite_root = root / "V12"
            template_dir = cutrite_root / "_WORK" / "USER1"
            remote_profile = root / "Paulo_Catarino"
            template_dir.mkdir(parents=True)
            remote_profile.mkdir(parents=True)
            (template_dir / "Excel-XLS.upx").write_text("xls", encoding="utf-8")
            (template_dir / "Excel-CSV.upx").write_text("csv", encoding="utf-8")
            (remote_profile / "systemv12.ctl").write_text("SYSDATAPATH,\\\\SERVER\\Data\\\n", encoding="utf-8")
            (remote_profile / "v12defs.ctl").write_text(
                "ImportPartFormat=0\nImportPartParams=\nShowImportFileDialog=0\n",
                encoding="utf-8",
            )

            with patch.object(cutrite_automation, "_is_local_path", return_value=False):
                runtime_dir_1 = cutrite_automation.prepare_cutrite_runtime_workdir(cutrite_root, remote_profile)
                runtime_dir_2 = cutrite_automation.prepare_cutrite_runtime_workdir(cutrite_root, remote_profile)

        self.assertNotEqual(runtime_dir_1, runtime_dir_2)
        self.assertTrue(runtime_dir_1.is_dir())
        self.assertTrue(runtime_dir_2.is_dir())

    def test_wait_for_cutrite_output_files_accepts_cutlist_core_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data"
            data_dir.mkdir()
            for suffix in (".ctt", ".mpd", ".brd"):
                (data_dir / f"0417_01_01_26_JF_VIVA{suffix}").write_text("x", encoding="utf-8")

            context = cutrite_automation.CutRiteImportContext(
                processo=SimpleNamespace(id=1),
                folder_path=root,
                source_workbook_path=root / "Lista_Material.xlsm",
                import_workbook_path=root / "0417_01_01_26_JF_VIVA.xls",
                cutrite_input_path=root / "IMPORT" / "0417_01_01_26_JF_VIVA.xls",
                plan_name="0417_01_01_26_JF_VIVA",
                cutrite_exe_path=Path(r"C:\V12\V12.exe"),
                import_exe_path=Path(r"C:\V12\Import.exe"),
                cutrite_root=Path(r"C:\V12"),
                cutrite_profile_dir=Path(r"\\SERVER_LE\Homag_iX\Cutrite\V12-Data\Paulo_Catarino"),
                cutrite_workdir=Path(r"C:\V12\_WORK\USER1"),
                cutrite_data_dir=data_dir,
                cutrite_target_data_dir=data_dir,
            )

            paths = cutrite_automation._wait_for_cutrite_output_files(context, timeout_seconds=1)

        self.assertEqual(
            {path.suffix.lower() for path in paths},
            {".brd", ".ctt", ".mpd"},
        )

    def test_wait_for_cutrite_output_files_accepts_minimum_ctt_and_mpd(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data"
            data_dir.mkdir()
            for suffix in (".ctt", ".mpd"):
                (data_dir / f"0417_01_01_26_JF_VIVA{suffix}").write_text("x", encoding="utf-8")

            context = cutrite_automation.CutRiteImportContext(
                processo=SimpleNamespace(id=1),
                folder_path=root,
                source_workbook_path=root / "Lista_Material.xlsm",
                import_workbook_path=root / "0417_01_01_26_JF_VIVA.xls",
                cutrite_input_path=root / "IMPORT" / "0417_01_01_26_JF_VIVA.xls",
                plan_name="0417_01_01_26_JF_VIVA",
                cutrite_exe_path=Path(r"C:\V12\V12.exe"),
                import_exe_path=Path(r"C:\V12\Import.exe"),
                cutrite_root=Path(r"C:\V12"),
                cutrite_profile_dir=Path(r"\\SERVER_LE\Homag_iX\Cutrite\V12-Data\Paulo_Catarino"),
                cutrite_workdir=Path(r"C:\V12\_WORK\USER1"),
                cutrite_data_dir=data_dir,
                cutrite_target_data_dir=data_dir,
            )

            paths = cutrite_automation._wait_for_cutrite_output_files(context, timeout_seconds=1)

        self.assertEqual(
            {path.suffix.lower() for path in paths},
            {".ctt", ".mpd"},
        )

    def test_navigate_to_cutrite_parts_table_sends_shift_enter_three_times(self):
        parts_window = unittest.mock.Mock()
        keyboard = unittest.mock.Mock()

        with patch.object(cutrite_automation, "_sleep_cutrite"):
            cutrite_automation._navigate_to_cutrite_parts_table(parts_window, keyboard)

        self.assertEqual(keyboard.send_keys.call_count, 3)
        keyboard.send_keys.assert_has_calls([
            unittest.mock.call("+{ENTER}", pause=0.1),
            unittest.mock.call("+{ENTER}", pause=0.1),
            unittest.mock.call("+{ENTER}", pause=0.1),
        ])


if __name__ == "__main__":
    unittest.main()

import json
import unittest
import uuid
from contextlib import contextmanager
from pathlib import Path
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from Martelo_Orcamentos_V2.app.services import producao_lista_material_audit as svc_audit


BASE_HEADERS = [
    "Descricao",
    "Material",
    "Comp",
    "Larg",
    "Qt",
    "Veio",
    "Orla",
    "Cliente",
    "Ref_Cliente",
    "Processo",
    "Artigo",
    "Notas",
    "Esp",
    "Grafico Orlas",
    "Orla ESQ",
    "Orla DIR",
    "Orla CIMA",
    "Orla BAIXO",
    "ID",
    "CNC_1",
    "CNC_2",
    "+comp",
    "+Larg",
    "Esp.Mat",
    "Esp.Final",
    "Tipo_Lacagem",
]


def _build_row(
    *,
    descricao="Porta",
    material="MDF_MR_MLM_BRANCO_19MM",
    comp=1000,
    larg=500,
    qt=1,
    veio="N",
    artigo="ARM_01",
    notas="MONTADO",
    esp=19,
    orla_esq="PVC_0.4_LINHO",
    orla_dir="PVC_0.4_LINHO",
    orla_cima="PVC_0.4_LINHO",
    orla_baixo="PVC_0.4_LINHO",
    cnc_1="",
    cnc_2="",
) -> dict:
    return {
        "Descricao": descricao,
        "Material": material,
        "Comp": comp,
        "Larg": larg,
        "Qt": qt,
        "Veio": veio,
        "Orla": "",
        "Cliente": "JF_VIVA",
        "Ref_Cliente": "2500000",
        "Processo": "0395_01_26",
        "Artigo": artigo,
        "Notas": notas,
        "Esp": esp,
        "Grafico Orlas": "111:111:000:000",
        "Orla ESQ": orla_esq,
        "Orla DIR": orla_dir,
        "Orla CIMA": orla_cima,
        "Orla BAIXO": orla_baixo,
        "ID": 1,
        "CNC_1": cnc_1,
        "CNC_2": cnc_2,
        "+comp": "",
        "+Larg": "",
        "Esp.Mat": esp,
        "Esp.Final": esp,
        "Tipo_Lacagem": "",
    }


def _write_workbook(
    path: Path,
    rows: list[dict],
    *,
    with_sheet: bool = True,
    with_table: bool = True,
    missing_headers: tuple[str, ...] = (),
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = svc_audit.TARGET_SHEET_NAME if with_sheet else "OUTRA_FOLHA"

    for column_index, header in enumerate(BASE_HEADERS, start=1):
        worksheet.cell(row=2, column=column_index).value = "" if header in missing_headers else header

    for row_index, row in enumerate(rows, start=3):
        for column_index, header in enumerate(BASE_HEADERS, start=1):
            worksheet.cell(row=row_index, column=column_index).value = row.get(header, "")

    if with_sheet and with_table:
        table_ref = f"A2:Z{max(2, len(rows) + 2)}"
        table = Table(displayName=svc_audit.TARGET_TABLE_NAME, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    workbook.save(path)
    workbook.close()


class ListaMaterialAuditTests(unittest.TestCase):
    @contextmanager
    def _workspace_tmp(self):
        root = Path("tests") / ("tmp_lista_material_audit_" + uuid.uuid4().hex)
        root.mkdir(parents=True, exist_ok=False)
        try:
            yield root
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def _config_root(self, temp_dir: str) -> Path:
        config_root = Path(temp_dir) / "cfg"
        svc_audit.ensure_external_config_files(config_root)
        return config_root

    def _audit(self, temp_dir: str, workbook_path: Path):
        return svc_audit.audit_lista_material_workbook(
            workbook_path,
            config_root=self._config_root(temp_dir),
        )

    def test_missing_sheet_returns_structural_error(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row()], with_sheet=False)

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "estrutura_ficheiro" for issue in result.issues))

    def test_missing_table_uses_header_fallback(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row()], with_table=False)

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(result.used_header_fallback)
        self.assertTrue(any(issue.rule_id == "fallback_cabecalho" for issue in result.issues))

    def test_missing_required_column_is_error(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row()], missing_headers=("Notas",))

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "colunas_obrigatorias" for issue in result.issues))

    def test_invalid_qt_is_error(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row(qt="abc")])

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "qt_invalido" for issue in result.issues))

    def test_qt_formula_is_warning(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row(qt="=1+1")])

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "qt_formula" for issue in result.issues))
        self.assertFalse(any(issue.rule_id == "qt_invalido" for issue in result.issues))

    def test_empty_material_with_exception_is_not_flagged(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row(material="", notas="COLAR 19+19 + MONTAR")])

            result = self._audit(str(tmp), workbook_path)

        self.assertFalse(any(issue.rule_id == "material_vazio" for issue in result.issues))

    def test_empty_material_without_exception_is_warning(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row(material="", notas="MONTADO")])

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "material_vazio" for issue in result.issues))

    def test_note_typo_group_suggests_canonical(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            rows = [_build_row(notas="PUC"), _build_row(notas="PUX")]
            _write_workbook(workbook_path, rows)

            result = self._audit(str(tmp), workbook_path)

        typo_issues = [issue for issue in result.issues if issue.rule_id == "nota_uniformizar"]
        self.assertTrue(typo_issues)
        self.assertTrue(any(issue.suggestion == "PUX" for issue in typo_issues))

    def test_note_order_difference_is_grouped(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            rows = [
                _build_row(notas="MONTADO + PUXADOR APLICAR(JF_VIVA)"),
                _build_row(notas="PUX APLICAR (JF_VIVA) + MONTADO"),
            ]
            _write_workbook(workbook_path, rows)

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "nota_uniformizar" for issue in result.issues))

    def test_note_spacing_difference_is_grouped(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            rows = [_build_row(notas="LACAR  1 FACE"), _build_row(notas="LACAR 1 FACE")]
            _write_workbook(workbook_path, rows)

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "nota_uniformizar" for issue in result.issues))

    def test_protected_plus_tokenization_keeps_numeric_sequence(self):
        with self._workspace_tmp() as tmp:
            config = svc_audit.load_audit_config(tmp / "cfg")
            tokens = svc_audit._safe_split_note_components(
                "COLAR 19+19 + MONTAR",
                config.note_protected_plus_patterns,
            )
            self.assertEqual(tokens, ("COLAR 19+19", "MONTAR"))

    def test_floor_rule_accepts_coherent_floor(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row(artigo="P1", notas="PISO 1 + CNC")])

            result = self._audit(str(tmp), workbook_path)

        self.assertFalse(any(issue.rule_id == "piso_incoerente" for issue in result.issues))

    def test_floor_rule_does_not_run_when_article_has_no_floor(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row(artigo="ARM_01", notas="PISO 2")])

            result = self._audit(str(tmp), workbook_path)

        self.assertFalse(any(issue.rule_id == "piso_incoerente" for issue in result.issues))

    def test_uniformizacao_suggests_review_when_same_signature_has_blank_note(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            rows = [_build_row(notas=""), _build_row(notas="MONTADO")]
            _write_workbook(workbook_path, rows)

            result = self._audit(str(tmp), workbook_path)

        self.assertTrue(any(issue.rule_id == "uniformizacao_nota_em_falta" for issue in result.issues))

    def test_orla_value_cnc_fresar_is_machining(self):
        with self._workspace_tmp() as tmp:
            config = svc_audit.load_audit_config(tmp / "cfg")
            self.assertEqual(svc_audit.classify_orla_value("CNC_FRESAR", config), "maquinacao")

    def test_orla_value_pvc_is_real_edge(self):
        with self._workspace_tmp() as tmp:
            config = svc_audit.load_audit_config(tmp / "cfg")
            self.assertEqual(svc_audit.classify_orla_value("PVC_0.4_LINHO", config), "orla_real")

    def test_export_report_creates_xlsx(self):
        with self._workspace_tmp() as tmp:
            workbook_path = tmp / "Lista_Material_TESTE.xlsm"
            _write_workbook(workbook_path, [_build_row(notas="PUC"), _build_row(notas="PUX")])

            result = self._audit(str(tmp), workbook_path)
            output_path = tmp / "relatorio.xlsx"
            svc_audit.export_audit_report(result, output_path)

            exported = load_workbook(output_path, read_only=True, data_only=True)
            try:
                self.assertIn("Resumo", exported.sheetnames)
                self.assertIn("Ocorrencias", exported.sheetnames)
            finally:
                exported.close()

    def test_config_files_are_created_outside_app(self):
        with self._workspace_tmp() as tmp:
            config_root = tmp / "externa"
            svc_audit.ensure_external_config_files(config_root)

            self.assertTrue((config_root / "notes.json").is_file())
            self.assertTrue((config_root / "materials.json").is_file())
            payload = json.loads((config_root / "notes.json").read_text(encoding="utf-8"))
            self.assertIn("whole_note_aliases", payload)


if __name__ == "__main__":
    unittest.main()

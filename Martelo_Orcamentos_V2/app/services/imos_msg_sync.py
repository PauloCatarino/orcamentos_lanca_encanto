from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Dict, List, Tuple
from zipfile import BadZipFile

from openpyxl import load_workbook


KEY_IMOS_MSG_PATH = "imos_msg_sync.target_path"
KEY_IMOS_MSG_WORKBOOK_PATH = "imos_msg_sync.workbook_path"

DEFAULT_IMOS_MSG_PATH = r"C:\Program Files\imos AG\iX CAD 2025\BIN\MSG\imos.msg"
DEFAULT_IMOS_MSG_WORKBOOK_PATH = r"I:\imos_msg.xlsx"
LEGACY_IMOS_MSG_WORKBOOK_PATH = r"I:\imos_msg.xlx"

_UTF8_BOM = b"\xef\xbb\xbf"
_MSG_LINE_RE = re.compile(r"^(?P<lang>[A-Z0-9]{3});(?P<ident>\d+)(?P<spacing>\s*);(?P<text>.*)$")
_LANG_RE = re.compile(r"^[A-Z0-9]{3}$")


@dataclass(frozen=True)
class ImosMsgSyncResult:
    workbook_path: Path
    msg_path: Path
    total_overrides: int
    matched: int
    updated: int
    unchanged: int
    backup_path: Path | None
    duplicate_keys: List[str]
    missing_keys: List[str]


def resolve_msg_path(path_value: str | Path | None) -> Path:
    text = str(path_value or DEFAULT_IMOS_MSG_PATH).strip() or DEFAULT_IMOS_MSG_PATH
    return Path(text)


def resolve_workbook_path(path_value: str | Path | None) -> Path:
    text = str(path_value or DEFAULT_IMOS_MSG_WORKBOOK_PATH).strip() or DEFAULT_IMOS_MSG_WORKBOOK_PATH
    candidate = Path(text)
    if candidate.exists():
        return candidate

    fallbacks: List[Path] = []
    suffix = candidate.suffix.lower()
    if suffix == ".xlsx":
        fallbacks.append(candidate.with_suffix(".xlx"))
    elif suffix == ".xlx":
        fallbacks.append(candidate.with_suffix(".xlsx"))

    if not path_value:
        fallbacks.append(Path(LEGACY_IMOS_MSG_WORKBOOK_PATH))

    for fallback in fallbacks:
        if fallback.exists():
            return fallback
    return candidate


def load_translation_overrides(workbook_path: str | Path | None) -> Tuple[Path, Dict[str, str], List[str]]:
    resolved = resolve_workbook_path(workbook_path)
    workbook = _open_workbook_any_extension(resolved)
    overrides: Dict[str, str] = {}
    duplicate_keys: List[str] = []
    try:
        worksheet = workbook.active
        for ref_value, text_value in worksheet.iter_rows(
            min_row=1,
            min_col=2,
            max_col=3,
            values_only=True,
        ):
            key = _normalize_reference(ref_value)
            translated_text = _normalize_translation_text(text_value)
            if not key or not translated_text:
                continue
            if key in overrides and overrides[key] != translated_text and key not in duplicate_keys:
                duplicate_keys.append(key)
            overrides[key] = translated_text
    finally:
        workbook.close()

    if not overrides:
        raise ValueError(
            "O Excel nao contem traducoes validas nas colunas B e C.\n\n"
            "Esperado: coluna B = referencia (ex.: PTG;10280), coluna C = texto em portugues."
        )

    return resolved, overrides, duplicate_keys


def sync_imos_msg_translations(
    *,
    msg_path: str | Path | None,
    workbook_path: str | Path | None,
) -> ImosMsgSyncResult:
    resolved_msg_path = resolve_msg_path(msg_path)
    resolved_workbook_path, overrides, duplicate_keys = load_translation_overrides(workbook_path)
    raw = _read_bytes(resolved_msg_path, label="ficheiro imos.msg")
    try:
        original_text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError(
            f"O ficheiro imos.msg nao esta em UTF-8 valido:\n{resolved_msg_path}"
        ) from exc

    newline = "\r\n" if b"\r\n" in raw else "\n"
    has_bom = raw.startswith(_UTF8_BOM)
    had_trailing_newline = original_text.endswith("\n")

    new_lines: List[str] = []
    matched_keys: set[str] = set()
    updated = 0
    unchanged = 0

    for line in original_text.splitlines():
        match = _MSG_LINE_RE.match(line)
        if not match:
            new_lines.append(line)
            continue

        key = f"{match.group('lang').upper()};{int(match.group('ident'))}"
        replacement = overrides.get(key)
        if replacement is None:
            new_lines.append(line)
            continue

        matched_keys.add(key)
        if match.group("text") == replacement:
            unchanged += 1
            new_lines.append(line)
            continue

        updated += 1
        new_lines.append(
            f"{match.group('lang')};{match.group('ident')}{match.group('spacing')};{replacement}"
        )

    missing_keys = [key for key in overrides if key not in matched_keys]
    backup_path: Path | None = None

    if updated:
        backup_path = _build_backup_path(resolved_msg_path)
        _copy_file(resolved_msg_path, backup_path)
        new_text = newline.join(new_lines)
        if had_trailing_newline:
            new_text += newline
        payload = new_text.encode("utf-8-sig" if has_bom else "utf-8")
        _atomic_write(resolved_msg_path, payload)

    return ImosMsgSyncResult(
        workbook_path=resolved_workbook_path,
        msg_path=resolved_msg_path,
        total_overrides=len(overrides),
        matched=len(matched_keys),
        updated=updated,
        unchanged=unchanged,
        backup_path=backup_path,
        duplicate_keys=duplicate_keys,
        missing_keys=missing_keys,
    )


def format_sync_result(result: ImosMsgSyncResult) -> str:
    lines = [
        f"Excel lido: {result.workbook_path}",
        f"Ficheiro imos.msg: {result.msg_path}",
        "",
        f"Referencias consideradas no Excel: {result.total_overrides}",
        f"Referencias encontradas no imos.msg: {result.matched}",
        f"Linhas alteradas: {result.updated}",
        f"Linhas ja corretas: {result.unchanged}",
    ]
    if result.backup_path is not None:
        lines.append(f"Backup criado: {result.backup_path}")
    if result.duplicate_keys:
        lines.append(
            "Referencias repetidas no Excel (foi usada a ultima traducao): "
            + _preview_keys(result.duplicate_keys)
        )
    if result.missing_keys:
        lines.append(
            "Referencias do Excel nao encontradas no imos.msg: "
            + _preview_keys(result.missing_keys)
        )
    return "\n".join(lines)


def _normalize_reference(value: object) -> str | None:
    text = _stringify_cell(value).strip()
    if not text:
        return None

    if ";" in text:
        lang_text, ident_text = text.split(";", 1)
    else:
        lang_text, ident_text = "PTG", text

    lang = re.sub(r"\s+", "", lang_text).upper()
    ident = ident_text.strip()
    if ident.endswith(".0"):
        ident = ident[:-2]
    if not _LANG_RE.fullmatch(lang):
        return None
    if not ident.isdigit():
        return None
    return f"{lang};{int(ident)}"


def _normalize_translation_text(value: object) -> str:
    text = _stringify_cell(value)
    if not text or not text.strip():
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\n", r"\n")
    return text.strip()


def _open_workbook_any_extension(workbook_path: Path):
    raw = _read_bytes(workbook_path, label="Excel de traducoes")
    try:
        return load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except BadZipFile as exc:
        raise ValueError(
            f"O ficheiro Excel configurado nao tem um formato valido (.xlsx/.xlsm):\n{workbook_path}"
        ) from exc
    except Exception as exc:
        raise ValueError(
            f"Nao foi possivel abrir o Excel de traducoes:\n{workbook_path}\n\n{exc}"
        ) from exc


def _read_bytes(path: Path, *, label: str) -> bytes:
    try:
        return path.read_bytes()
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Nao foi possivel localizar o {label}:\n{path}") from exc
    except PermissionError as exc:
        raise PermissionError(
            f"Sem permissao para aceder ao {label}:\n{path}\n\n"
            "Confirme que o ficheiro nao esta bloqueado e que tem permissao de leitura/escrita."
        ) from exc


def _copy_file(source: Path, destination: Path) -> None:
    try:
        shutil.copy2(source, destination)
    except PermissionError as exc:
        raise PermissionError(
            f"Sem permissao para criar o backup do imos.msg em:\n{destination}\n\n"
            "Feche o IMOS e confirme que a pasta permite escrita."
        ) from exc


def _atomic_write(target: Path, payload: bytes) -> None:
    tmp_path: Path | None = None
    try:
        fd, tmp_name = tempfile.mkstemp(
            prefix=f"{target.name}.",
            suffix=".tmp",
            dir=str(target.parent),
        )
        tmp_path = Path(tmp_name)
        with os.fdopen(fd, "wb") as handle:
            handle.write(payload)
        os.replace(tmp_path, target)
    except PermissionError as exc:
        raise PermissionError(
            f"Sem permissao para atualizar o ficheiro imos.msg:\n{target}\n\n"
            "Feche o IMOS antes de executar esta operacao e confirme permissoes de escrita."
        ) from exc
    finally:
        if tmp_path is not None and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def _build_backup_path(msg_path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return msg_path.with_name(f"{msg_path.name}.bak_{stamp}")


def _preview_keys(keys: List[str], limit: int = 10) -> str:
    preview = ", ".join(keys[:limit])
    if len(keys) > limit:
        preview += f" ... (+{len(keys) - limit})"
    return preview


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)

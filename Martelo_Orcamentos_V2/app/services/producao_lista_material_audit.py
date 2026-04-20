from __future__ import annotations

import json
import logging
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from Martelo_Orcamentos_V2.app.services import cutrite_automation as svc_cutrite
from Martelo_Orcamentos_V2.app.services.settings import get_setting

logger = logging.getLogger(__name__)

KEY_LISTA_MATERIAL_AUDIT_CONFIG_ROOT = "lista_material_audit_config_root"
DEFAULT_LISTA_MATERIAL_AUDIT_FOLDER_NAME = "_Lista_Material_Audit"
TARGET_SHEET_NAME = svc_cutrite.LISTAGEM_CUT_RITE_SHEET
TARGET_TABLE_NAME = "Tabela_Cut_Rite"
SEVERITY_ORDER = {"info": 0, "sugestao": 1, "aviso": 2, "erro": 3}
CONFIDENCE_ORDER = {"low": 0, "medium": 1, "high": 2}
CATEGORY_ORDER = ("estrutura", "notas", "materiais", "uniformizacao", "orlas")
REQUIRED_COLUMNS = (
    "Descricao",
    "Material",
    "Comp",
    "Larg",
    "Qt",
    "Veio",
    "Artigo",
    "Notas",
    "Esp",
    "Orla ESQ",
    "Orla DIR",
    "Orla CIMA",
    "Orla BAIXO",
    "CNC_1",
    "CNC_2",
)
SIGNATURE_COLUMNS = (
    "Descricao",
    "Material",
    "Comp",
    "Larg",
    "Esp",
    "Veio",
    "Orla ESQ",
    "Orla DIR",
    "Orla CIMA",
    "Orla BAIXO",
    "CNC_1",
    "CNC_2",
)
ORLA_COLUMNS = ("Orla ESQ", "Orla DIR", "Orla CIMA", "Orla BAIXO")

DEFAULT_NOTES_CONFIG = {
    "whole_note_aliases": {
        "PUC": "PUX",
        "DOBRADGEM": "DOBRAGEM",
        "ECNHIMENTO": "ENCHIMENTO",
        "MONTADO + PUXADOR APLICAR(JF_VIVA)": "PUX APLICAR (JF_VIVA) + MONTADO",
        "MONTADO + PUX APLICAR": "PUX APLICAR + MONTADO",
        "LACAR  1 FACE": "LACAR 1 FACE",
        "SUTAR 1 TOPO <35.1+ MONTAR": "SUTAR 1 TOPO <35.1 + MONTAR",
        "LACAR 1 FACE + PUX TIC-TAC+SUTAR": "LACAR 1 FACE + PUX TIC-TAC + SUTAR",
    },
    "token_aliases": {
        "PUC": "PUX",
        "DOBRADGEM": "DOBRAGEM",
        "ECNHIMENTO": "ENCHIMENTO",
        "PUXADOR APLICAR(JF_VIVA)": "PUX APLICAR (JF_VIVA)",
        "PUX APLICAR(JF_VIVA)": "PUX APLICAR (JF_VIVA)",
    },
    "protected_plus_patterns": [
        r"\d+(?:[.,]\d+)?\s*(?:\+\s*\d+(?:[.,]\d+)?)+(?:\s*MM)?",
        r"\(\s*[+\-]{1,2}\s*\d+(?:[.,]\d+)?\s*\)",
    ],
    "uniform_context_token_patterns": [
        r"^CNC$",
        r"^COM CNC$",
        r"^RC$",
        r"^RES DO CHAO$",
        r"^PISO\s*-?1(?:\s*\(MENOS\))?$",
        r"^PISO\s*1$",
        r"^PISO\s*2$",
        r"^PISO\s*3$",
    ],
}
DEFAULT_MATERIALS_CONFIG = {
    "canonical_materials": [],
    "aliases": {},
    "empty_material_exception_patterns": [
        r"\bCOLAR\b",
        r"\bCOLADO\b",
        r"\bCOLADA\b",
        r"\bCOMPOSTA\b",
        r"\bCOLAGEM\b",
    ],
}
DEFAULT_FLOORS_CONFIG = {
    "article_clean_suffix_patterns": [
        r"\([A-Z]\)$",
        r"_(?:DIR|ESQ)$",
        r"-(?:DIR|ESQ)$",
    ],
    "article_patterns": [
        {"regex": r"^RC$", "floor": "RC"},
        {"regex": r"^P[-_ ]?1$", "floor": "P1"},
        {"regex": r"^P[-_ ]?2$", "floor": "P2"},
        {"regex": r"^P[-_ ]?3$", "floor": "P3"},
        {"regex": r"^P[-_ ]?-1$", "floor": "P-1"},
    ],
    "note_patterns": [
        {"regex": r"\bRES DO CHAO\b", "floor": "RC"},
        {"regex": r"\bPISO\s*1\b", "floor": "P1"},
        {"regex": r"\bPISO\s*2\b", "floor": "P2"},
        {"regex": r"\bPISO\s*3\b", "floor": "P3"},
        {"regex": r"\bPISO\s*-1(?:\s*\(MENOS\))?\b", "floor": "P-1"},
    ],
}
DEFAULT_ORLAS_CONFIG = {
    "edge_aliases": {},
    "edge_real_patterns": [
        r"^PVC(?:[_ ].+)?$",
        r"^ABS(?:[_ ].+)?$",
        r"^PP(?:[_ ].+)?$",
        r"^FITA(?:[_ ].+)?$",
    ],
    "machining_patterns": [
        r"^CNC(?:[_ ].+)?$",
        r"FRESAR",
    ],
    "sutar_patterns": [
        r"^SUTAR\b.*$",
    ],
}
DEFAULT_CONFIG_FILES = {
    "notes.json": DEFAULT_NOTES_CONFIG,
    "materials.json": DEFAULT_MATERIALS_CONFIG,
    "floors.json": DEFAULT_FLOORS_CONFIG,
    "orlas.json": DEFAULT_ORLAS_CONFIG,
}


@dataclass(frozen=True)
class AuditIssue:
    severity: str
    confidence: str
    rule_id: str
    category: str
    sheet_name: str
    excel_row: int
    column_name: str
    article: str
    raw_value: str
    normalized_value: str
    message: str
    suggestion: str = ""
    group_key: str = ""


@dataclass(frozen=True)
class IssueGroup:
    category: str
    rule_id: str
    group_key: str
    severity: str
    confidence: str
    message: str
    suggestion: str
    occurrences: int
    rows: tuple[int, ...]
    columns: tuple[str, ...]
    articles: tuple[str, ...]
    examples: tuple[str, ...]


@dataclass(frozen=True)
class AuditRow:
    excel_row: int
    values: dict[str, object]
    formula_columns: frozenset[str]

    def value(self, column_name: str) -> object:
        return self.values.get(column_name)

    def text(self, column_name: str) -> str:
        return _stringify_value(self.values.get(column_name))


@dataclass(frozen=True)
class ExcelAuditSource:
    source_path: Path
    sheet_name: str
    table_name: str
    used_header_fallback: bool
    headers: tuple[str, ...]
    rows: tuple[AuditRow, ...]


@dataclass(frozen=True)
class NoteAnalysis:
    raw_text: str
    display_text: str
    compare_text: str
    canonical_text: str
    orderless_key: str
    surface_key: str
    tokens: tuple[str, ...]
    uniform_tokens: tuple[str, ...]
    floor_code: str = ""


@dataclass(frozen=True)
class AuditConfig:
    config_root: Path
    note_aliases: dict[str, str]
    note_token_aliases: dict[str, str]
    note_protected_plus_patterns: tuple[str, ...]
    note_uniform_context_patterns: tuple[re.Pattern[str], ...]
    material_aliases: dict[str, str]
    canonical_materials: frozenset[str]
    empty_material_patterns: tuple[re.Pattern[str], ...]
    floor_article_cleaners: tuple[re.Pattern[str], ...]
    floor_article_patterns: tuple[tuple[re.Pattern[str], str], ...]
    floor_note_patterns: tuple[tuple[re.Pattern[str], str], ...]
    edge_aliases: dict[str, str]
    edge_real_patterns: tuple[re.Pattern[str], ...]
    edge_machining_patterns: tuple[re.Pattern[str], ...]
    edge_sutar_patterns: tuple[re.Pattern[str], ...]


@dataclass(frozen=True)
class AuditResult:
    source_path: Path
    config_root: Path
    sheet_name: str
    table_name: str
    used_header_fallback: bool
    total_rows: int
    issues: tuple[AuditIssue, ...]
    groups: tuple[IssueGroup, ...]
    severity_totals: dict[str, int]
    confidence_totals: dict[str, int]

    def groups_for_category(self, category: str) -> tuple[IssueGroup, ...]:
        return tuple(group for group in self.groups if group.category == category)


def default_config_root(base_producao: Optional[str]) -> str:
    base_text = str(base_producao or "").strip()
    if not base_text:
        return DEFAULT_LISTA_MATERIAL_AUDIT_FOLDER_NAME
    return str(Path(base_text) / DEFAULT_LISTA_MATERIAL_AUDIT_FOLDER_NAME)


def resolve_config_root(db) -> Path:
    base_producao = get_setting(db, "base_path_producao", None)
    configured = get_setting(
        db,
        KEY_LISTA_MATERIAL_AUDIT_CONFIG_ROOT,
        default_config_root(base_producao),
    )
    root = Path(str(configured or "").strip() or default_config_root(base_producao))
    ensure_external_config_files(root)
    return root


def ensure_external_config_files(config_root: Path) -> None:
    config_root.mkdir(parents=True, exist_ok=True)
    for filename, payload in DEFAULT_CONFIG_FILES.items():
        file_path = config_root / filename
        if file_path.exists():
            continue
        file_path.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )


def load_audit_config(config_root: Path) -> AuditConfig:
    ensure_external_config_files(config_root)
    notes = _read_json_file(config_root / "notes.json", DEFAULT_NOTES_CONFIG)
    materials = _read_json_file(config_root / "materials.json", DEFAULT_MATERIALS_CONFIG)
    floors = _read_json_file(config_root / "floors.json", DEFAULT_FLOORS_CONFIG)
    orlas = _read_json_file(config_root / "orlas.json", DEFAULT_ORLAS_CONFIG)

    note_aliases = {
        _normalize_text_compare(str(key or "")): str(value or "").strip()
        for key, value in dict(notes.get("whole_note_aliases") or {}).items()
        if str(key or "").strip() and str(value or "").strip()
    }
    note_token_aliases = {
        _normalize_text_compare(str(key or "")): str(value or "").strip()
        for key, value in dict(notes.get("token_aliases") or {}).items()
        if str(key or "").strip() and str(value or "").strip()
    }
    material_aliases = {
        _normalize_material_compare(str(key or "")): str(value or "").strip()
        for key, value in dict(materials.get("aliases") or {}).items()
        if str(key or "").strip() and str(value or "").strip()
    }
    canonical_materials = frozenset(
        _normalize_material_compare(value)
        for value in (materials.get("canonical_materials") or [])
        if str(value or "").strip()
    )
    edge_aliases = {
        _normalize_text_compare(str(key or "")): str(value or "").strip()
        for key, value in dict(orlas.get("edge_aliases") or {}).items()
        if str(key or "").strip() and str(value or "").strip()
    }

    return AuditConfig(
        config_root=config_root,
        note_aliases=note_aliases,
        note_token_aliases=note_token_aliases,
        note_protected_plus_patterns=tuple(notes.get("protected_plus_patterns") or []),
        note_uniform_context_patterns=_compile_patterns(notes.get("uniform_context_token_patterns") or []),
        material_aliases=material_aliases,
        canonical_materials=canonical_materials,
        empty_material_patterns=_compile_patterns(materials.get("empty_material_exception_patterns") or []),
        floor_article_cleaners=_compile_patterns(floors.get("article_clean_suffix_patterns") or []),
        floor_article_patterns=_compile_floor_patterns(floors.get("article_patterns") or []),
        floor_note_patterns=_compile_floor_patterns(floors.get("note_patterns") or []),
        edge_aliases=edge_aliases,
        edge_real_patterns=_compile_patterns(orlas.get("edge_real_patterns") or []),
        edge_machining_patterns=_compile_patterns(orlas.get("machining_patterns") or []),
        edge_sutar_patterns=_compile_patterns(orlas.get("sutar_patterns") or []),
    )


def resolve_workbook_for_audit(folder_path: str | Path, *, nome_enc_imos: str = "") -> Path:
    return svc_cutrite.find_lista_material_workbook(Path(folder_path), nome_enc_imos=nome_enc_imos)


def classify_orla_value(value: str, config: Optional[AuditConfig] = None) -> str:
    text = _stringify_value(value)
    if not text:
        return "vazio"

    compare_text = _normalize_text_compare(text)
    if config is not None:
        alias = config.edge_aliases.get(compare_text)
        if alias:
            compare_text = _normalize_text_compare(alias)

    if config is not None and _matches_any_pattern(compare_text, config.edge_machining_patterns):
        return "maquinacao"
    if config is not None and _matches_any_pattern(compare_text, config.edge_sutar_patterns):
        return "sutar"
    if config is not None and _matches_any_pattern(compare_text, config.edge_real_patterns):
        return "orla_real"
    if compare_text.startswith("CNC"):
        return "maquinacao"
    if compare_text.startswith("SUTAR"):
        return "sutar"
    if compare_text.startswith(("PVC", "ABS", "PP", "FITA")):
        return "orla_real"
    return "desconhecido"


def audit_lista_material_workbook(
    source_workbook_path: Path,
    *,
    config_root: Path,
) -> AuditResult:
    config = load_audit_config(config_root)
    issues: list[AuditIssue] = []

    try:
        source = _read_excel_source(source_workbook_path)
    except ValueError as exc:
        issues.append(
            AuditIssue(
                severity="erro",
                confidence="high",
                rule_id="estrutura_ficheiro",
                category="estrutura",
                sheet_name=TARGET_SHEET_NAME,
                excel_row=0,
                column_name="",
                article="",
                raw_value="",
                normalized_value="",
                message=str(exc),
                group_key="estrutura_ficheiro",
            )
        )
        return _build_audit_result(
            source_path=source_workbook_path,
            config_root=config_root,
            sheet_name=TARGET_SHEET_NAME,
            table_name="",
            used_header_fallback=False,
            total_rows=0,
            issues=issues,
        )

    if source.used_header_fallback:
        issues.append(
            AuditIssue(
                severity="aviso",
                confidence="high",
                rule_id="fallback_cabecalho",
                category="estrutura",
                sheet_name=source.sheet_name,
                excel_row=0,
                column_name="",
                article="",
                raw_value="",
                normalized_value="",
                message="A tabela Tabela_Cut_Rite nao foi encontrada. Foi usado fallback por cabecalho.",
                suggestion="Validar se o ficheiro devia ter a tabela Tabela_Cut_Rite.",
                group_key="fallback_cabecalho",
            )
        )

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in source.headers]
    if missing_columns:
        issues.append(
            AuditIssue(
                severity="erro",
                confidence="high",
                rule_id="colunas_obrigatorias",
                category="estrutura",
                sheet_name=source.sheet_name,
                excel_row=0,
                column_name="",
                article="",
                raw_value=", ".join(missing_columns),
                normalized_value="",
                message="Colunas obrigatorias em falta.",
                suggestion=", ".join(missing_columns),
                group_key="colunas_obrigatorias",
            )
        )
        return _build_audit_result(
            source_path=source.source_path,
            config_root=config_root,
            sheet_name=source.sheet_name,
            table_name=source.table_name,
            used_header_fallback=source.used_header_fallback,
            total_rows=len(source.rows),
            issues=issues,
        )

    note_rows: list[tuple[AuditRow, NoteAnalysis]] = []
    material_variants: defaultdict[str, list[tuple[AuditRow, str, str]]] = defaultdict(list)
    technical_groups: defaultdict[str, list[tuple[AuditRow, NoteAnalysis]]] = defaultdict(list)
    orla_classifications: Counter[tuple[str, str]] = Counter()

    for row in source.rows:
        article = row.text("Artigo")
        desc = row.text("Descricao")
        qt_value = row.value("Qt")
        qt_text = _stringify_value(qt_value)
        material_text = row.text("Material")
        note_text = row.text("Notas")

        if not desc:
            issues.append(_row_issue(row, category="estrutura", rule_id="descricao_vazia", column_name="Descricao", severity="erro", message="Descricao vazia.", group_key="descricao_vazia"))
        if not article:
            issues.append(_row_issue(row, category="estrutura", rule_id="artigo_vazio", column_name="Artigo", severity="erro", message="Artigo vazio.", group_key="artigo_vazio"))
        if "Qt" in row.formula_columns:
            issues.append(
                _row_issue(
                    row,
                    category="estrutura",
                    rule_id="qt_formula",
                    column_name="Qt",
                    severity="aviso",
                    confidence="high",
                    message="Qt contem formula e deve ser revista antes de seguir para producao.",
                    group_key="qt_formula",
                )
            )
        elif not _is_valid_quantity(qt_value):
            issues.append(
                _row_issue(
                    row,
                    category="estrutura",
                    rule_id="qt_invalido",
                    column_name="Qt",
                    severity="erro",
                    confidence="high",
                    raw_value=qt_text,
                    normalized_value=qt_text,
                    message="Qt vazio ou invalido.",
                    group_key="qt_invalido",
                )
            )

        for column_name in ("Comp", "Larg", "Esp"):
            if not row.text(column_name):
                issues.append(
                    _row_issue(
                        row,
                        category="estrutura",
                        rule_id="linha_incompleta",
                        column_name=column_name,
                        severity="aviso",
                        confidence="medium",
                        message=f"Campo tecnico '{column_name}' vazio.",
                        group_key=f"linha_incompleta:{column_name}",
                    )
                )

        if not material_text:
            combined_text = _normalize_text_compare(" ".join(filter(None, (desc, article, note_text))))
            if not _matches_any_pattern(combined_text, config.empty_material_patterns):
                issues.append(
                    _row_issue(
                        row,
                        category="materiais",
                        rule_id="material_vazio",
                        column_name="Material",
                        severity="aviso",
                        confidence="medium",
                        message="Material vazio e sem excecao configurada.",
                        group_key="material_vazio",
                    )
                )
        else:
            canonical_material = _canonical_material_text(material_text, config)
            material_key = _normalize_material_compare(canonical_material)
            material_variants[material_key].append((row, material_text, canonical_material))
            if config.canonical_materials and material_key not in config.canonical_materials:
                issues.append(
                    _row_issue(
                        row,
                        category="materiais",
                        rule_id="material_desconhecido",
                        column_name="Material",
                        severity="aviso",
                        confidence="medium",
                        raw_value=material_text,
                        normalized_value=material_key,
                        message="Material nao consta na configuracao canonica.",
                        suggestion=canonical_material or material_text,
                        group_key=f"material_desconhecido:{material_key}",
                    )
                )

        note_analysis = _analyze_note(note_text, config)
        if note_analysis.raw_text:
            note_rows.append((row, note_analysis))

        article_floor = _classify_article_floor(article, config)
        if article_floor and note_analysis.floor_code and article_floor != note_analysis.floor_code:
            issues.append(
                _row_issue(
                    row,
                    category="notas",
                    rule_id="piso_incoerente",
                    column_name="Notas",
                    severity="aviso",
                    confidence="high",
                    raw_value=note_text,
                    normalized_value=note_analysis.compare_text,
                    message=f"Nota de piso incoerente com o artigo ({article_floor} vs {note_analysis.floor_code}).",
                    group_key=f"piso_incoerente:{article_floor}:{note_analysis.floor_code}",
                )
            )

        technical_groups[_build_technical_signature(row)].append((row, note_analysis))

        for column_name in ORLA_COLUMNS:
            value = row.text(column_name)
            if value:
                orla_classifications[(value, classify_orla_value(value, config))] += 1

    issues.extend(_build_note_variant_issues(note_rows))
    issues.extend(_build_material_variant_issues(material_variants))
    issues.extend(_build_uniformizacao_issues(technical_groups))
    issues.extend(_build_orla_classification_issues(orla_classifications))

    return _build_audit_result(
        source_path=source.source_path,
        config_root=config_root,
        sheet_name=source.sheet_name,
        table_name=source.table_name,
        used_header_fallback=source.used_header_fallback,
        total_rows=len(source.rows),
        issues=issues,
    )


def export_audit_report(result: AuditResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        summary = workbook.active
        summary.title = "Resumo"
        summary.append(["Ficheiro", str(result.source_path)])
        summary.append(["Folha", result.sheet_name])
        summary.append(["Tabela", result.table_name or "<fallback cabecalho>"])
        summary.append(["Fallback cabecalho", "Sim" if result.used_header_fallback else "Nao"])
        summary.append(["Linhas auditadas", result.total_rows])
        summary.append(["Issues", len(result.issues)])
        summary.append([])
        summary.append(["Severidade", "Total"])
        for severity in ("erro", "aviso", "sugestao", "info"):
            summary.append([severity, result.severity_totals.get(severity, 0)])
        summary.append([])
        summary.append(["Confianca", "Total"])
        for confidence in ("high", "medium", "low"):
            summary.append([confidence, result.confidence_totals.get(confidence, 0)])

        for category in CATEGORY_ORDER:
            sheet = workbook.create_sheet(_sheet_title(category))
            sheet.append(["Severidade", "Confianca", "Regra", "Mensagem", "Sugestao", "Ocorrencias", "Linhas", "Colunas", "Artigos", "Exemplos"])
            for group in result.groups_for_category(category):
                sheet.append([
                    group.severity,
                    group.confidence,
                    group.rule_id,
                    group.message,
                    group.suggestion,
                    group.occurrences,
                    ", ".join(str(value) for value in group.rows),
                    ", ".join(group.columns),
                    ", ".join(group.articles),
                    " | ".join(group.examples),
                ])

        details = workbook.create_sheet("Ocorrencias")
        details.append(["Categoria", "Severidade", "Confianca", "Regra", "Folha", "Linha Excel", "Coluna", "Artigo", "Valor bruto", "Valor normalizado", "Mensagem", "Sugestao", "Group Key"])
        for issue in result.issues:
            details.append([
                issue.category,
                issue.severity,
                issue.confidence,
                issue.rule_id,
                issue.sheet_name,
                issue.excel_row,
                issue.column_name,
                issue.article,
                issue.raw_value,
                issue.normalized_value,
                issue.message,
                issue.suggestion,
                issue.group_key,
            ])

        workbook.save(output_path)
    finally:
        _close_workbook_resources(workbook)
    return output_path


def _build_audit_result(
    *,
    source_path: Path,
    config_root: Path,
    sheet_name: str,
    table_name: str,
    used_header_fallback: bool,
    total_rows: int,
    issues: Iterable[AuditIssue],
) -> AuditResult:
    issue_list = list(issues)
    groups = _group_issues(issue_list)
    severity_totals = {key: 0 for key in ("erro", "aviso", "sugestao", "info")}
    confidence_totals = {key: 0 for key in ("high", "medium", "low")}
    for issue in issue_list:
        severity_totals[issue.severity] = severity_totals.get(issue.severity, 0) + 1
        confidence_totals[issue.confidence] = confidence_totals.get(issue.confidence, 0) + 1
    return AuditResult(
        source_path=source_path,
        config_root=config_root,
        sheet_name=sheet_name,
        table_name=table_name,
        used_header_fallback=used_header_fallback,
        total_rows=total_rows,
        issues=tuple(issue_list),
        groups=tuple(groups),
        severity_totals=severity_totals,
        confidence_totals=confidence_totals,
    )


def _read_excel_source(source_workbook_path: Path) -> ExcelAuditSource:
    if not source_workbook_path.exists():
        raise ValueError(f"Ficheiro Excel nao encontrado:\n{source_workbook_path}")

    workbook_values = load_workbook(source_workbook_path, read_only=False, data_only=True, keep_vba=True)
    workbook_formula = load_workbook(source_workbook_path, read_only=False, data_only=False, keep_vba=True)
    try:
        if TARGET_SHEET_NAME not in workbook_values.sheetnames:
            raise ValueError(f"Folha {TARGET_SHEET_NAME} nao encontrada.")

        values_sheet = workbook_values[TARGET_SHEET_NAME]
        formula_sheet = workbook_formula[TARGET_SHEET_NAME]
        table = formula_sheet.tables.get(TARGET_TABLE_NAME)
        used_header_fallback = table is None

        if table is not None:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = [_normalize_header_name(formula_sheet.cell(row=min_row, column=col).value) for col in range(min_col, max_col + 1)]
            data_start_row = min_row + 1
        else:
            min_col, min_row, max_col, max_row = _find_header_fallback_bounds(formula_sheet)
            headers = [_normalize_header_name(formula_sheet.cell(row=min_row, column=col).value) for col in range(min_col, max_col + 1)]
            data_start_row = min_row + 1

        rows: list[AuditRow] = []
        for excel_row in range(data_start_row, max_row + 1):
            row_values: dict[str, object] = {}
            formula_columns: set[str] = set()
            has_data = False
            for offset, header in enumerate(headers, start=min_col):
                if not header:
                    continue
                value_cell = values_sheet.cell(row=excel_row, column=offset)
                formula_cell = formula_sheet.cell(row=excel_row, column=offset)
                row_values[header] = value_cell.value
                if formula_cell.data_type == "f":
                    formula_columns.add(header)
                if value_cell.value not in (None, "") or formula_cell.value not in (None, ""):
                    has_data = True
            if not has_data:
                continue
            rows.append(AuditRow(excel_row=excel_row, values=row_values, formula_columns=frozenset(formula_columns)))

        return ExcelAuditSource(
            source_path=source_workbook_path,
            sheet_name=TARGET_SHEET_NAME,
            table_name=TARGET_TABLE_NAME if table is not None else "",
            used_header_fallback=used_header_fallback,
            headers=tuple(value for value in headers if value),
            rows=tuple(rows),
        )
    finally:
        _close_workbook_resources(workbook_values)
        _close_workbook_resources(workbook_formula)


def _find_header_fallback_bounds(sheet: Worksheet) -> tuple[int, int, int, int]:
    max_scan_row = min(sheet.max_row, 10)
    for row_index in range(1, max_scan_row + 1):
        headers = [
            _normalize_header_name(sheet.cell(row=row_index, column=col).value)
            for col in range(1, min(sheet.max_column, 40) + 1)
        ]
        found_headers = {header for header in headers if header}
        if {"Descricao", "Material", "Qt", "Artigo"}.issubset(found_headers):
            last_col = max(
                col
                for col in range(1, min(sheet.max_column, 40) + 1)
                if _normalize_header_name(sheet.cell(row=row_index, column=col).value)
            )
            return 1, row_index, last_col, sheet.max_row
    raise ValueError(
        f"A tabela {TARGET_TABLE_NAME} nao foi encontrada e nao foi possivel localizar um cabecalho fallback."
    )


def _close_workbook_resources(workbook: Workbook) -> None:
    try:
        workbook.close()
    except Exception:
        pass

    archive = getattr(workbook, "vba_archive", None)
    if archive is not None:
        try:
            archive.close()
        except Exception:
            pass


def _row_issue(
    row: AuditRow,
    *,
    category: str,
    rule_id: str,
    column_name: str,
    severity: str,
    message: str,
    confidence: str = "high",
    raw_value: str = "",
    normalized_value: str = "",
    suggestion: str = "",
    group_key: str = "",
) -> AuditIssue:
    return AuditIssue(
        severity=severity,
        confidence=confidence,
        rule_id=rule_id,
        category=category,
        sheet_name=TARGET_SHEET_NAME,
        excel_row=row.excel_row,
        column_name=column_name,
        article=row.text("Artigo"),
        raw_value=raw_value or row.text(column_name),
        normalized_value=normalized_value or row.text(column_name),
        message=message,
        suggestion=suggestion,
        group_key=group_key or rule_id,
    )


def _build_note_variant_issues(note_rows: list[tuple[AuditRow, NoteAnalysis]]) -> list[AuditIssue]:
    grouped: defaultdict[str, list[tuple[AuditRow, NoteAnalysis]]] = defaultdict(list)
    for row, note in note_rows:
        grouped[note.orderless_key].append((row, note))

    issues: list[AuditIssue] = []
    for orderless_key, items in grouped.items():
        raw_variants = Counter(note.raw_text for _, note in items if note.raw_text)
        if len(raw_variants) <= 1:
            continue

        canonical_counts = Counter(note.canonical_text for _, note in items if note.canonical_text)
        canonical = canonical_counts.most_common(1)[0][0] if canonical_counts else ""
        cleaned_variants = {_normalize_text_compare(note.display_text) for _, note in items if note.display_text}
        confidence = "high" if len(cleaned_variants) <= 1 else "medium"

        for row, note in items:
            issues.append(
                AuditIssue(
                    severity="sugestao",
                    confidence=confidence,
                    rule_id="nota_uniformizar",
                    category="notas",
                    sheet_name=TARGET_SHEET_NAME,
                    excel_row=row.excel_row,
                    column_name="Notas",
                    article=row.text("Artigo"),
                    raw_value=note.raw_text,
                    normalized_value=note.compare_text,
                    message="Nota com variantes equivalentes noutras linhas; rever uniformizacao.",
                    suggestion=canonical,
                    group_key=f"nota_uniformizar:{orderless_key}",
                )
            )
    return issues


def _build_material_variant_issues(
    material_variants: defaultdict[str, list[tuple[AuditRow, str, str]]]
) -> list[AuditIssue]:
    issues: list[AuditIssue] = []
    for material_key, items in material_variants.items():
        raw_variants = Counter(raw for _, raw, _canonical in items if raw)
        if len(raw_variants) <= 1:
            continue
        canonical_counts = Counter(canonical for _row, _raw, canonical in items if canonical)
        canonical = canonical_counts.most_common(1)[0][0] if canonical_counts else ""
        for row, raw, _canonical in items:
            issues.append(
                AuditIssue(
                    severity="sugestao",
                    confidence="high",
                    rule_id="material_uniformizar",
                    category="materiais",
                    sheet_name=TARGET_SHEET_NAME,
                    excel_row=row.excel_row,
                    column_name="Material",
                    article=row.text("Artigo"),
                    raw_value=raw,
                    normalized_value=material_key,
                    message="Material equivalente com variantes de escrita.",
                    suggestion=canonical,
                    group_key=f"material_uniformizar:{material_key}",
                )
            )
    return issues


def _build_uniformizacao_issues(
    technical_groups: defaultdict[str, list[tuple[AuditRow, NoteAnalysis]]]
) -> list[AuditIssue]:
    issues: list[AuditIssue] = []
    for technical_key, items in technical_groups.items():
        if len(items) < 2:
            continue

        non_empty = [(row, note) for row, note in items if note.raw_text]
        blank = [(row, note) for row, note in items if not note.raw_text]
        base_variants = Counter("|".join(note.uniform_tokens) for _row, note in non_empty if note.uniform_tokens)

        if blank and base_variants:
            canonical_map = Counter(note.canonical_text for _row, note in non_empty if note.canonical_text)
            canonical = canonical_map.most_common(1)[0][0] if canonical_map else ""
            for row, _note in blank:
                issues.append(
                    AuditIssue(
                        severity="sugestao",
                        confidence="medium",
                        rule_id="uniformizacao_nota_em_falta",
                        category="uniformizacao",
                        sheet_name=TARGET_SHEET_NAME,
                        excel_row=row.excel_row,
                        column_name="Notas",
                        article=row.text("Artigo"),
                        raw_value="",
                        normalized_value="",
                        message="Linha tecnicamente equivalente a outras com nota preenchida; rever possivel omissao.",
                        suggestion=canonical,
                        group_key=f"uniformizacao_nota_em_falta:{technical_key}",
                    )
                )

        if len(base_variants) > 1:
            canonical_map = Counter(note.canonical_text for _row, note in non_empty if note.canonical_text)
            canonical = canonical_map.most_common(1)[0][0] if canonical_map else ""
            for row, note in non_empty:
                issues.append(
                    AuditIssue(
                        severity="sugestao",
                        confidence="medium",
                        rule_id="uniformizacao_notas_variantes",
                        category="uniformizacao",
                        sheet_name=TARGET_SHEET_NAME,
                        excel_row=row.excel_row,
                        column_name="Notas",
                        article=row.text("Artigo"),
                        raw_value=note.raw_text,
                        normalized_value="|".join(note.uniform_tokens),
                        message="Linhas tecnicamente equivalentes apresentam notas diferentes.",
                        suggestion=canonical,
                        group_key=f"uniformizacao_notas_variantes:{technical_key}",
                    )
                )
    return issues


def _build_orla_classification_issues(orla_classifications: Counter[tuple[str, str]]) -> list[AuditIssue]:
    issues: list[AuditIssue] = []
    for (raw_value, classification), count in sorted(orla_classifications.items()):
        severity = "info" if classification != "desconhecido" else "aviso"
        issues.append(
            AuditIssue(
                severity=severity,
                confidence="high",
                rule_id="orla_classificacao",
                category="orlas",
                sheet_name=TARGET_SHEET_NAME,
                excel_row=0,
                column_name="Orlas",
                article="",
                raw_value=raw_value,
                normalized_value=classification,
                message=f"Valor de orla classificado como '{classification}'.",
                suggestion=str(count),
                group_key=f"orla_classificacao:{classification}:{raw_value}",
            )
        )
    return issues


def _group_issues(issues: list[AuditIssue]) -> list[IssueGroup]:
    grouped: defaultdict[tuple[str, str, str], list[AuditIssue]] = defaultdict(list)
    for issue in issues:
        group_key = issue.group_key or f"{issue.rule_id}:{issue.normalized_value}:{issue.message}"
        grouped[(issue.category, issue.rule_id, group_key)].append(issue)

    result: list[IssueGroup] = []
    for (category, rule_id, group_key), items in grouped.items():
        rows = tuple(sorted({issue.excel_row for issue in items if issue.excel_row > 0}))
        columns = tuple(sorted({issue.column_name for issue in items if issue.column_name}))
        articles = tuple(sorted({issue.article for issue in items if issue.article}))
        examples = tuple(value for value in list(dict.fromkeys(issue.raw_value for issue in items if issue.raw_value))[:6])
        severity = max(items, key=lambda issue: SEVERITY_ORDER.get(issue.severity, -1)).severity
        confidence = max(items, key=lambda issue: CONFIDENCE_ORDER.get(issue.confidence, -1)).confidence
        suggestion = next((issue.suggestion for issue in items if issue.suggestion), "")
        message = Counter(issue.message for issue in items if issue.message).most_common(1)[0][0]
        result.append(
            IssueGroup(
                category=category,
                rule_id=rule_id,
                group_key=group_key,
                severity=severity,
                confidence=confidence,
                message=message,
                suggestion=suggestion,
                occurrences=len(items),
                rows=rows,
                columns=columns,
                articles=articles,
                examples=examples,
            )
        )
    result.sort(
        key=lambda group: (
            CATEGORY_ORDER.index(group.category) if group.category in CATEGORY_ORDER else len(CATEGORY_ORDER),
            -SEVERITY_ORDER.get(group.severity, -1),
            -CONFIDENCE_ORDER.get(group.confidence, -1),
            group.rule_id,
            group.group_key,
        )
    )
    return result


def _analyze_note(text: str, config: AuditConfig) -> NoteAnalysis:
    raw_text = _stringify_value(text)
    if not raw_text:
        return NoteAnalysis(
            raw_text="",
            display_text="",
            compare_text="",
            canonical_text="",
            orderless_key="",
            surface_key="",
            tokens=tuple(),
            uniform_tokens=tuple(),
            floor_code="",
        )

    display_text = _normalize_note_display_text(raw_text, config.note_protected_plus_patterns)
    compare_text = _normalize_text_compare(display_text)
    canonical_text = config.note_aliases.get(compare_text, display_text)
    canonical_display = _normalize_note_display_text(canonical_text, config.note_protected_plus_patterns)
    tokens = _safe_split_note_components(canonical_display, config.note_protected_plus_patterns)
    if not tokens:
        tokens = (canonical_display,)

    canonical_tokens: list[str] = []
    for token in tokens:
        token_compare = _normalize_text_compare(token)
        canonical_tokens.append(config.note_token_aliases.get(token_compare, token))

    orderless_key = "|".join(sorted(_normalize_text_compare(token) for token in canonical_tokens if token.strip()))
    surface_key = _normalize_text_compare(" + ".join(canonical_tokens))
    uniform_tokens = tuple(
        _normalize_text_compare(token)
        for token in canonical_tokens
        if token.strip()
        and not _matches_any_pattern(_normalize_text_compare(token), config.note_uniform_context_patterns)
    )
    floor_code = _classify_floor_note(" + ".join(canonical_tokens), config)
    return NoteAnalysis(
        raw_text=raw_text,
        display_text=display_text,
        compare_text=compare_text,
        canonical_text=" + ".join(canonical_tokens),
        orderless_key=orderless_key or surface_key,
        surface_key=surface_key,
        tokens=tuple(canonical_tokens),
        uniform_tokens=uniform_tokens,
        floor_code=floor_code,
    )


def _classify_article_floor(article: str, config: AuditConfig) -> str:
    compare_text = _normalize_text_compare(article)
    if not compare_text:
        return ""
    cleaned = compare_text
    for pattern in config.floor_article_cleaners:
        cleaned = pattern.sub("", cleaned).strip()
    for pattern, floor_code in config.floor_article_patterns:
        if pattern.search(cleaned):
            return floor_code
    return ""


def _classify_floor_note(note_text: str, config: AuditConfig) -> str:
    compare_text = _normalize_text_compare(note_text)
    if not compare_text:
        return ""
    for pattern, floor_code in config.floor_note_patterns:
        if pattern.search(compare_text):
            return floor_code
    return ""


def _build_technical_signature(row: AuditRow) -> str:
    parts = []
    for column_name in SIGNATURE_COLUMNS:
        value = row.text(column_name)
        if column_name in ("Descricao", "Material", "Veio"):
            normalized = _normalize_text_compare(value)
        else:
            normalized = _normalize_numericish_text(value)
        parts.append(f"{column_name}={normalized}")
    return "|".join(parts)


def _canonical_material_text(material_text: str, config: AuditConfig) -> str:
    compare_key = _normalize_material_compare(material_text)
    return config.material_aliases.get(compare_key, _normalize_material_display(material_text))


def _normalize_material_display(material_text: str) -> str:
    text = _stringify_value(material_text)
    text = text.replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def _normalize_material_compare(material_text: str) -> str:
    text = _normalize_text_compare(_normalize_material_display(material_text))
    return text.replace(" ", "_")


def _normalize_note_display_text(text: str, protected_patterns: tuple[str, ...]) -> str:
    protected_text, mapping = _protect_sequences(text, protected_patterns)
    protected_text = protected_text.replace("’", "'").replace("“", '"').replace("”", '"')
    protected_text = re.sub(r"\s+", " ", protected_text.strip())
    protected_text = re.sub(r"\s*\+\s*", " + ", protected_text)
    protected_text = re.sub(r"\s+", " ", protected_text).strip()
    for placeholder, original in mapping.items():
        protected_text = protected_text.replace(placeholder, original)
    return protected_text


def _safe_split_note_components(text: str, protected_patterns: tuple[str, ...]) -> tuple[str, ...]:
    protected_text, mapping = _protect_sequences(text, protected_patterns)
    parts = [part.strip() for part in re.split(r"\s*\+\s*", protected_text) if part.strip()]
    restored: list[str] = []
    for part in parts:
        for placeholder, original in mapping.items():
            part = part.replace(placeholder, original)
        restored.append(part.strip())
    return tuple(restored)


def _protect_sequences(text: str, patterns: tuple[str, ...]) -> tuple[str, dict[str, str]]:
    protected_text = text
    mapping: dict[str, str] = {}
    counter = 0
    for regex_text in patterns:
        pattern = re.compile(regex_text, re.IGNORECASE)
        while True:
            match = pattern.search(protected_text)
            if match is None:
                break
            placeholder = f"__PROT_{counter}__"
            mapping[placeholder] = match.group(0)
            protected_text = protected_text[: match.start()] + placeholder + protected_text[match.end() :]
            counter += 1
    return protected_text, mapping


def _normalize_header_name(value: object) -> str:
    text = _stringify_value(value)
    text = text.replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def _normalize_numericish_text(value: str) -> str:
    text = _stringify_value(value)
    text = text.replace(",", ".")
    text = re.sub(r"\s+", "", text)
    return text.upper()


def _normalize_text_compare(value: object) -> str:
    text = _stringify_value(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\n", " ").replace("\r", " ")
    text = text.replace("º", "").replace("°", "").replace("ª", "")
    text = text.replace("’", "'").replace("“", '"').replace("”", '"')
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def _stringify_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        text = f"{value:.6f}".rstrip("0").rstrip(".")
        return text or "0"
    return str(value).strip()


def _is_valid_quantity(value: object) -> bool:
    text = _stringify_value(value)
    if not text:
        return False
    try:
        return float(text.replace(",", ".")) > 0
    except Exception:
        return False


def _compile_patterns(patterns: Iterable[str]) -> tuple[re.Pattern[str], ...]:
    compiled: list[re.Pattern[str]] = []
    for regex_text in patterns:
        regex_clean = str(regex_text or "").strip()
        if not regex_clean:
            continue
        compiled.append(re.compile(regex_clean, re.IGNORECASE))
    return tuple(compiled)


def _compile_floor_patterns(entries: Iterable[dict[str, str]]) -> tuple[tuple[re.Pattern[str], str], ...]:
    compiled: list[tuple[re.Pattern[str], str]] = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        regex_text = str(entry.get("regex") or "").strip()
        floor_code = str(entry.get("floor") or "").strip()
        if not regex_text or not floor_code:
            continue
        compiled.append((re.compile(regex_text, re.IGNORECASE), floor_code))
    return tuple(compiled)


def _matches_any_pattern(text: str, patterns: Iterable[re.Pattern[str]]) -> bool:
    for pattern in patterns:
        if pattern.search(text):
            return True
    return False


def _read_json_file(path: Path, default_payload: dict) -> dict:
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if isinstance(payload, dict):
            return payload
    except Exception:
        logger.warning("Falha a ler configuracao de auditoria: %s", path)
    return dict(default_payload)


def _sheet_title(category: str) -> str:
    mapping = {
        "estrutura": "Estrutura",
        "notas": "Notas",
        "materiais": "Materiais",
        "uniformizacao": "Uniformizacao",
        "orlas": "Orlas",
    }
    return mapping.get(category, category.title())

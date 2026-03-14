from __future__ import annotations

import os
import logging
from pathlib import Path
import json
import textwrap
# carregar .env para disponibilizar OPENAI_API_KEY quando executamos localmente
from dotenv import load_dotenv
load_dotenv()  # lê .env na raiz do projecto e define variáveis de ambiente
import unicodedata
import re

from typing import Any, Dict, List, Optional, Sequence, Mapping

from openpyxl import load_workbook
from sqlalchemy import create_engine, text, bindparam
from sqlalchemy.orm import Session

from Martelo_Orcamentos_V2.app.config import settings
from Martelo_Orcamentos_V2.app.services.settings import get_setting

try:
    import faiss  # type: ignore
except Exception:  # pragma: no cover - se nao instalado
    faiss = None

try:
    from sentence_transformers import SentenceTransformer
except Exception:  # pragma: no cover - se nao instalado
    SentenceTransformer = None

try:
    from transformers import AutoModelForSeq2SeqLM, AutoTokenizer
except Exception:  # pragma: no cover - opcional
    AutoModelForSeq2SeqLM = None
    AutoTokenizer = None

try:
    from openai import OpenAI  # type: ignore
except Exception:  # pragma: no cover - opcional
    OpenAI = None

# Configurações / chaves
KEY_IA_BASE_PATH = "ia_base_pesquisa_profunda"
DEFAULT_IA_BASE_PATH = r"\\SERVER_LE\_Lanca_Encanto\LancaEncanto\Dep._Compras\Tabela e Catalogos_Fornecedores\Tabelas Preços\Pesquisa_Profunda_IA"

KEY_IA_EMB_PATH = "ia_embeddings_path"
DEFAULT_IA_SHARED_PATH = (
    r"\\SERVER_LE\_Lanca_Encanto\LancaEncanto\Dep._Orcamentos\Base_Dados_Orcamento\Pesquisa_IA_Martelo"
)
# Defaults em UNC (partilhado no servidor) para funcionar em qualquer PC.
DEFAULT_IA_EMB_PATH = DEFAULT_IA_SHARED_PATH
KEY_IA_MODEL_NAME = "ia_model_name"
DEFAULT_IA_MODEL_NAME = "google/flan-t5-base"
KEY_IA_MODEL_PATH = "ia_model_path"
DEFAULT_IA_MODEL_PATH = DEFAULT_IA_SHARED_PATH
# Fallbacks locais (dev) caso o share não esteja disponível.
FALLBACK_LOCAL_IA_EMB_PATH = str((Path(__file__).resolve().parents[3] / "data" / "ia_embeddings").resolve())
FALLBACK_LOCAL_IA_MODEL_PATH = str(
    (Path(__file__).resolve().parents[3] / "data" / "ia_models" / "flan-t5-base").resolve()
)
KEY_IA_GEN_PROVIDER = "ia_gen_provider"  # "auto" | "local" | "openai"
DEFAULT_IA_GEN_PROVIDER = "local"
KEY_IA_OPENAI_MODEL = "ia_openai_model"
DEFAULT_IA_OPENAI_MODEL = "gpt-4o-mini"

META_FILENAME = "faiss_meta.jsonl"
FAISS_FILENAME = "faiss.index"
MODEL_NAME = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
EXCEL_REF_FILENAME = "12_Placas_Referencias_COMPLETO.xlsx"

_MODEL_CACHE: Optional[SentenceTransformer] = None
_INDEX_CACHE: Dict[str, Any] = {}
_META_CACHE: Dict[str, List[Dict[str, Any]]] = {}
_GEN_MODEL_CACHE: Optional[Any] = None
_GEN_TOKENIZER_CACHE: Optional[Any] = None
_OPENAI_CLIENT: Optional[Any] = None
_EXCEL_CACHE: Dict[str, List[Dict[str, Any]]] = {}
logger = logging.getLogger(__name__)


def _normalize_txt(val: str) -> str:
    """Normaliza texto para comparação: lower, remove acentos e pontuação."""
    if not val:
        return ""
    txt = unicodedata.normalize("NFKD", val)
    txt = txt.encode("ASCII", "ignore").decode("ascii")
    txt = txt.lower()
    txt = re.sub(r"[^a-z0-9]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _ensure_shared_dirs(base: Path) -> None:
    try:
        (base / "ia_models").mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    try:
        (base / "ia_embeddings").mkdir(parents=True, exist_ok=True)
    except Exception:
        pass


def ia_base_path(db: Session) -> str:
    return get_setting(db, KEY_IA_BASE_PATH, DEFAULT_IA_BASE_PATH) or DEFAULT_IA_BASE_PATH


def ia_embeddings_path(db: Session) -> str:
    configured = (get_setting(db, KEY_IA_EMB_PATH, DEFAULT_IA_EMB_PATH) or "").strip() or DEFAULT_IA_EMB_PATH
    configured_path = Path(configured).expanduser()

    # Preferir o caminho partilhado (UNC) sempre que o valor configurado for local/relativo.
    shared_path = Path(DEFAULT_IA_SHARED_PATH).expanduser()
    shared_available = False
    try:
        shared_path.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    if shared_path.exists():
        _ensure_shared_dirs(shared_path)
        shared_available = True

    if shared_available and not str(configured).startswith("\\\\"):
        return str(shared_path)

    if shared_available and not _path_is_ascii(configured_path):
        logger.warning("Embeddings path contem caracteres especiais; a usar caminho partilhado.")
        return str(shared_path)

    if configured_path.exists():
        return str(configured_path)

    if shared_available:
        return str(shared_path)

    # fallback local (dev) se o share falhar
    alt = Path(FALLBACK_LOCAL_IA_EMB_PATH).expanduser()
    if alt.exists():
        return str(alt)
    return str(configured_path)


def _path_is_ascii(path: Path) -> bool:
    try:
        return str(path).isascii()
    except Exception:
        return False


def ia_model_name(db: Session) -> str:
    return get_setting(db, KEY_IA_MODEL_NAME, DEFAULT_IA_MODEL_NAME) or DEFAULT_IA_MODEL_NAME


def ia_model_path(db: Session) -> str:
    # Caminho local (pasta com config.json/tokenizer/model) para execução offline; opcional.
    configured = (get_setting(db, KEY_IA_MODEL_PATH, DEFAULT_IA_MODEL_PATH) or "").strip() or DEFAULT_IA_MODEL_PATH
    cfg_path = Path(configured).expanduser()

    # Preferir o caminho partilhado (UNC) sempre que o valor configurado for local/relativo.
    shared_path = Path(DEFAULT_IA_SHARED_PATH).expanduser()
    shared_available = False
    try:
        shared_path.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    if shared_path.exists():
        _ensure_shared_dirs(shared_path)
        shared_available = True

    if shared_available and not str(configured).startswith("\\\\"):
        return str(shared_path)

    if cfg_path.exists():
        return str(cfg_path)

    if shared_available:
        return str(shared_path)

    # fallback local (dev) se o share falhar
    fallback = Path(FALLBACK_LOCAL_IA_MODEL_PATH).expanduser()
    if fallback.exists():
        return str(fallback)
    return str(cfg_path)


def ia_excel_path(db: Session, filename: str = EXCEL_REF_FILENAME) -> Path:
    base = Path(ia_base_path(db)).expanduser()
    return (base / filename).resolve()


def ia_gen_provider(db: Session) -> str:
    prov = (get_setting(db, KEY_IA_GEN_PROVIDER, DEFAULT_IA_GEN_PROVIDER) or DEFAULT_IA_GEN_PROVIDER).lower()
    if prov not in ("auto", "local", "openai"):
        return DEFAULT_IA_GEN_PROVIDER
    return prov


def ia_openai_model(db: Session) -> str:
    return get_setting(db, KEY_IA_OPENAI_MODEL, DEFAULT_IA_OPENAI_MODEL) or DEFAULT_IA_OPENAI_MODEL


def _load_model() -> Optional[SentenceTransformer]:
    global _MODEL_CACHE
    if _MODEL_CACHE is not None:
        return _MODEL_CACHE
    if SentenceTransformer is None:
        return None
    try:
        _MODEL_CACHE = SentenceTransformer(MODEL_NAME)
    except Exception:
        _MODEL_CACHE = None
    return _MODEL_CACHE


def _load_gen_model(model_name: str, model_path: Optional[str] = None):
    global _GEN_MODEL_CACHE, _GEN_TOKENIZER_CACHE
    if _GEN_MODEL_CACHE is not None and _GEN_TOKENIZER_CACHE is not None:
        return _GEN_MODEL_CACHE, _GEN_TOKENIZER_CACHE
    if AutoModelForSeq2SeqLM is None or AutoTokenizer is None:
        raise RuntimeError("Pacote transformers não está instalado.")
    try:
        load_target = model_path or model_name
        tok = AutoTokenizer.from_pretrained(load_target, local_files_only=True)
        mdl = AutoModelForSeq2SeqLM.from_pretrained(load_target, local_files_only=True)
    except Exception as exc:
        hint = "Configure o caminho local do modelo nas Configurações (Pasta Modelo IA Texto)."
        raise RuntimeError(f"Falha ao carregar modelo de geração '{model_name}' (offline/local): {exc}\n{hint}")
    _GEN_MODEL_CACHE = mdl
    _GEN_TOKENIZER_CACHE = tok
    return mdl, tok


def _load_openai_client():
    global _OPENAI_CLIENT
    if _OPENAI_CLIENT is not None:
        return _OPENAI_CLIENT
    if OpenAI is None:
        raise RuntimeError("Pacote 'openai' não está instalado. Instala com: pip install openai")
    api_key = os.getenv("OPENAI_API_KEY") or ""
    if not api_key:
        raise RuntimeError(
            "OPENAI_API_KEY não definida; configure a variável de ambiente ou coloque a chave no ficheiro .env."
        )
    # Criar cliente OpenAI
    _OPENAI_CLIENT = OpenAI(api_key=api_key)
    return _OPENAI_CLIENT


def _load_index(index_path: Path) -> Optional[Any]:
    cache_key = str(index_path)
    if cache_key in _INDEX_CACHE:
        return _INDEX_CACHE[cache_key]
    if faiss is None or not index_path.exists():
        return None
    try:
        idx = faiss.read_index(str(index_path))
    except Exception:
        return None
    _INDEX_CACHE[cache_key] = idx
    return idx


def _load_meta(meta_path: Path) -> List[Dict[str, Any]]:
    cache_key = str(meta_path)
    if cache_key in _META_CACHE:
        return _META_CACHE[cache_key]
    entries: List[Dict[str, Any]] = []
    if not meta_path.exists():
        return entries
    try:
        with meta_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entries.append(json.loads(line))
                except Exception:
                    continue
    except Exception:
        entries = []
    _META_CACHE[cache_key] = entries
    return entries


def _load_excel_rows(xlsx_path: Path) -> List[Dict[str, Any]]:
    """
    Lê o Excel e retorna lista de dicts:
    - sheet, row, text, text_norm, row_data (dict coluna->valor)
    Usa uma heurística para cabeçalhos: escolhe a linha com mais valores texto
    nas primeiras linhas (até 5). Se um cabeçalho vier em branco, nomeia como
    "Coluna N".
    """
    cache_key = str(xlsx_path)
    if cache_key in _EXCEL_CACHE:
        return _EXCEL_CACHE[cache_key]
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Ficheiro Excel não encontrado: {xlsx_path}")
    rows: List[Dict[str, Any]] = []
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        # escolher linha de cabeçalho: se a linha 4 tiver texto, usar; senão heurística
        header_idx = None
        if len(all_rows) >= 4:
            candidate = all_rows[3]
            if candidate and any(isinstance(c, str) and c.strip() for c in candidate):
                header_idx = 4
        if header_idx is None:
            best_txt = -1
            for i, row in enumerate(all_rows[:5], start=1):
                if not row:
                    continue
                str_count = sum(1 for c in row if isinstance(c, str) and c.strip())
                if str_count > best_txt:
                    best_txt = str_count
                    header_idx = i
        if header_idx is None:
            continue
        raw_headers = all_rows[header_idx - 1]
        headers: List[str] = []
        for i, cell in enumerate(raw_headers):
            name = ""
            if isinstance(cell, str):
                name = cell.strip()
            elif cell is not None:
                name = str(cell).strip()
            if not name:
                name = f"Coluna {i+1}"
            headers.append(name)

        # processar linhas de dados após o cabeçalho
        for idx, row in enumerate(all_rows[header_idx:], start=header_idx + 1):
            if not row:
                continue
            values = [c for c in row if c not in (None, "")]
            if not values:
                continue
            row_data: Dict[str, Any] = {}
            for i, header in enumerate(headers):
                val = row[i] if i < len(row) else None
                if val is None:
                    continue
                row_data[header] = val
            if not row_data:
                continue
            text = " | ".join(f"{k}: {v}" for k, v in row_data.items())
            text_clean = text.replace("\n", " ").strip()
            if not text_clean:
                continue
            combined = f"{sheet} | {text_clean}"
            combined_norm = _normalize_txt(combined)
            if not combined_norm:
                continue
            rows.append(
                {
                    "sheet": sheet,
                    "row": idx,
                    "text": text_clean,
                    "text_norm": combined_norm,
                    "row_data": row_data,
                }
            )
    _EXCEL_CACHE[cache_key] = rows
    return rows


def _engine_from_settings():
    return create_engine(settings.DB_URI, pool_pre_ping=True, echo=False)


def buscar(
    db: Session,
    query: str,
    *,
    top_k: int = 5,
    embeddings_dir: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Pesquisa vetorial sobre os dados ingestados. Retorna lista de dicts com:
    score, fornecedor, ficheiro, pagina, caminho, snippet.
    """
    if not query or not query.strip():
        return []
    query = query.strip()

    base_emb = embeddings_dir or ia_embeddings_path(db)
    emb_dir = Path(base_emb).expanduser().resolve()
    index_path = emb_dir / FAISS_FILENAME
    meta_path = emb_dir / META_FILENAME

    meta = _load_meta(meta_path)
    if not meta:
        raise RuntimeError(f"Meta não encontrada ou vazia em {meta_path}")

    model = _load_model()
    index = _load_index(index_path)
    if model is None:
        raise RuntimeError("Modelo de embeddings não disponível (sentence-transformers).")
    if index is None:
        raise RuntimeError(f"Índice FAISS não encontrado em {index_path}")
    if getattr(index, "ntotal", 0) <= 0:
        raise RuntimeError(f"Índice FAISS vazio (ntotal=0) em {index_path}; reexecute ingest_profundo.py")

    vector = model.encode([query], convert_to_numpy=True, normalize_embeddings=True)
    scores, positions = index.search(vector, top_k)
    scores = scores[0]
    positions = positions[0]

    results: List[Dict[str, Any]] = []
    chunk_ids: List[int] = []
    chunk_pos: Dict[int, float] = {}
    for pos, score in zip(positions, scores):
        if pos < 0 or pos >= len(meta):
            continue
        chunk_id = meta[pos].get("chunk_id")
        if chunk_id:
            cid = int(chunk_id)
            chunk_ids.append(cid)
            chunk_pos[cid] = float(score)

    engine = _engine_from_settings()

    if not chunk_ids:
        # Fallback: procura substring simples em ia_chunks.text
        like_query = f"%{query}%"
        with engine.begin() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT c.id, c.text, c.page, d.path, d.filename, d.supplier
                    FROM ia_chunks c
                    JOIN ia_documents d ON d.id = c.document_id
                    WHERE c.text LIKE :q
                    LIMIT :lim
                    """
                ),
                {"q": like_query, "lim": top_k},
            ).fetchall()
        results: List[Dict[str, Any]] = []
        for row in rows:
            snippet = (row.text or "").strip().replace("\n", " ")
            if len(snippet) > 260:
                snippet = snippet[:260] + "..."
            results.append(
                {
                    "score": 0.0,
                    "fornecedor": row.supplier or "",
                    "ficheiro": row.filename or "",
                    "pagina": row.page,
                    "caminho": row.path or "",
                    "snippet": snippet,
                }
            )
        return results

    # fetch chunks + docs para resultado vetorial
    with engine.begin() as conn:
        stmt = text(
            """
            SELECT c.id, c.text, c.page, d.path, d.filename, d.supplier
            FROM ia_chunks c
            JOIN ia_documents d ON d.id = c.document_id
            WHERE c.id IN :ids
            """
        ).bindparams(bindparam("ids", expanding=True))
        rows = conn.execute(stmt, {"ids": chunk_ids}).fetchall()

    for row in rows:
        snippet = (row.text or "").strip().replace("\n", " ")
        if len(snippet) > 260:
            snippet = snippet[:260] + "..."
        results.append(
            {
                "score": chunk_pos.get(row.id, 0.0),
                "fornecedor": row.supplier or "",
                "ficheiro": row.filename or "",
                "pagina": row.page,
                "caminho": row.path or "",
                "snippet": snippet,
            }
        )

    results.sort(key=lambda x: x.get("score", 0.0), reverse=True)
    return results[:top_k]


def buscar_excel(
    db: Session,
    query: str,
    *,
    top_k: int = 20,
    excel_path: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Pesquisa no ficheiro Excel 12_Placas_Referencias_COMPLETO.xlsx (ou caminho indicado).
    Retorna dicts com: score, folha, linha, snippet, caminho.
    """
    if not query or not query.strip():
        return []
    norm_query = _normalize_txt(query)
    tokens = [t for t in norm_query.split() if t]
    if not tokens:
        return []

    path = Path(excel_path) if excel_path else ia_excel_path(db)
    rows = _load_excel_rows(path)

    results: List[Dict[str, Any]] = []
    for entry in rows:
        txt = entry["text_norm"]
        matched = [tok for tok in tokens if tok in txt]
        if not matched:
            continue
        score = sum(txt.count(tok) for tok in matched)
        snippet = entry["text"]
        if len(snippet) > 260:
            snippet = snippet[:260] + "..."
        results.append(
            {
                "score": float(score),
                "folha": entry["sheet"],
                "linha": entry["row"],
                "snippet": snippet,
                "caminho": str(path),
                "row_data": entry.get("row_data", {}),
            }
        )
        if len(results) >= top_k:
            break
    results.sort(key=lambda x: x.get("score", 0.0), reverse=True)
    return results[:top_k]


def gerar_resposta(
    db: Session,
    query: str,
    snippets: Sequence[Mapping[str, Any]],
    *,
    model_name: Optional[str] = None,
    model_path: Optional[str] = None,
    max_new_tokens: int = 120,
) -> str:
    """
    Gera uma resposta curta em linguagem natural usando FLAN-T5 (ou equivalente) de forma offline.
    - snippets: lista de dicts com campos 'snippet', 'ficheiro', 'pagina'.
    """
    if not query.strip():
        return ""
    if not snippets:
        return "Nenhum contexto disponível para responder."

    prov = ia_gen_provider(db)
    # Desativado uso automático de OpenAI; só usa OpenAI se selecionado explicitamente.
    if prov == "openai":
        try:
            return _gerar_resposta_openai(db, query, snippets)
        except Exception as exc:
            # fallback para local se openai falhar
            fallback_msg = f"[Aviso] OpenAI falhou: {exc}. A usar modelo local."
            return fallback_msg + "\n" + _gerar_resposta_local(db, query, snippets, model_name, model_path, max_new_tokens)
    return _gerar_resposta_local(db, query, snippets, model_name, model_path, max_new_tokens)


def _gerar_resposta_local(
    db: Session,
    query: str,
    snippets: Sequence[Mapping[str, Any]],
    model_name: Optional[str],
    model_path: Optional[str],
    max_new_tokens: int,
) -> str:
    name = model_name or ia_model_name(db)
    path = model_path or ia_model_path(db)
    model, tok = _load_gen_model(name, path if path else None)

    ctx_parts = []
    for idx, sn in enumerate(snippets, 1):
        texto = str(sn.get("snippet") or "").replace("\n", " ").strip()
        ficheiro = sn.get("ficheiro") or ""
        pagina = sn.get("pagina")
        ref = f"{ficheiro}" + (f" pág. {pagina}" if pagina not in (None, "") else "")
        chunk = f"{idx}) {ref}: {texto}"
        ctx_parts.append(chunk)
    contexto = "\n".join(ctx_parts[:5])

    prompt = textwrap.dedent(
        f"""
        Responda em português, em 3-5 tópicos claros, usando apenas a informação abaixo.
        Cada tópico deve ter uma frase curta e, se possível, a referência (ficheiro/página) entre parêntesis.
        Se não houver dados suficientes, diga que não encontrou.

        Pergunta: {query}

        Contexto:
        {contexto}
        """
    ).strip()

    inputs = tok(prompt, return_tensors="pt", truncation=True, max_length=1024)
    output_ids = model.generate(**inputs, max_new_tokens=max_new_tokens, num_beams=4)
    resposta = tok.decode(output_ids[0], skip_special_tokens=True)
    return resposta.strip()


def _gerar_resposta_openai(
    db: Session,
    query: str,
    snippets: Sequence[Mapping[str, Any]],
) -> str:
    """
    Gera resposta usando OpenAI via API.
    Estratégia: map-reduce — sumarizar batches de snippets e depois gerar resposta final.
    """
    client = _load_openai_client()
    if not snippets:
        return "Nenhum contexto disponível para responder."

    # Configuração
    model_id = ia_openai_model(db)  # ex: "gpt-4o-mini"
    BATCH_SIZE = 6                  # quantos snippets por batch de sumarização
    MAX_SUMMARY_TOKENS = 160
    MAX_FINAL_TOKENS = 320
    TEMPERATURE = 0.1

    # Formatar snippets para prompt
    def fmt_snippet(i, sn):
        texto = str(sn.get("snippet") or "").replace("\n", " ").strip()
        if len(texto) > 900:
            texto = texto[:900] + "..."
        ficheiro = sn.get("ficheiro") or ""
        pagina = sn.get("pagina")
        ref = f"{ficheiro}" + (f" pág. {pagina}" if pagina not in (None, "") else "")
        return f"[{i}] {ref}: {texto}"

    # Construir batches
    formatted = [fmt_snippet(i+1, s) for i, s in enumerate(snippets)]
    batches = [formatted[i : i + BATCH_SIZE] for i in range(0, len(formatted), BATCH_SIZE)]

    # Função auxiliar para chamar a API com retry simples
    def _call_chat(messages, max_tokens):
        retries = 0
        while True:
            try:
                resp = client.chat.completions.create(
                    model=model_id,
                    messages=messages,
                    temperature=TEMPERATURE,
                    max_tokens=max_tokens,
                )
                if resp.choices:
                    return resp.choices[0].message.content.strip()
                return ""
            except Exception as exc:
                status = getattr(exc, "status_code", None)
                err_txt = str(exc).lower()
                if status == 429 or "quota" in err_txt or "insufficient_quota" in err_txt:
                    # Propaga para o caller ativar fallback local
                    raise RuntimeError(f"OpenAI: limite/quota atingido ({exc})")
                retries += 1
                if retries > 3:
                    raise
                # espera crescente
                import time
                time.sleep(1.5 * retries)

    # 1) Sumarizar cada batch
    summaries = []
    for batch in batches:
        prompt = (
            "Resumo curto em Português (2-3 frases) do conteúdo abaixo. "
            "Indica entre parêntesis a fonte no formato (ficheiro | página) quando relevante.\n\n"
        )
        prompt += "\n\n".join(batch)
        # montar mensagens
        messages = [
            {"role": "system", "content": "És um assistente técnico que resume excertos em Português."},
            {"role": "user", "content": prompt},
        ]
        try:
            s = _call_chat(messages, MAX_SUMMARY_TOKENS)
        except RuntimeError:
            # quota/429: propaga para ativar fallback local
            raise
        except Exception:
            # se falhar (outro motivo), guardamos uma anotação
            s = "[Resumo indisponível devido a erro na API]"
        summaries.append(s)

    # 2) Criar prompt final com todos os summaries e a pergunta
    context_block = "\n\n".join(f"[{i+1}] {s}" for i, s in enumerate(summaries))
    system_prompt = (
        "És um assistente técnico que responde em Português com base apenas no contexto fornecido. "
        "Responde em 3-5 tópicos curtos. Indica fontes no formato (ficheiro | página). "
        "Se não houver dados suficientes, diz que não encontraste."
    )
    user_prompt = (
        f"Pergunta: {query}\n\n"
        f"Contexto (sumários):\n{context_block}\n\n"
        "Resposta em tópicos curtos (inclui referências entre parênteses no formato (ficheiro | página)):\n"
    )
    messages_final = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    try:
        final = _call_chat(messages_final, MAX_FINAL_TOKENS)
    except RuntimeError:
        # quota/429: propaga para ativar fallback local
        raise
    except Exception:
        # fallback: tentar resposta simples (sem sumarização)
        try:
            # como fallback tenta uma chamada simples com os primeiros 6 snippets
            fallback_ctx = "\n".join(formatted[:6])
            messages_fb = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Pergunta: {query}\n\nContexto:\n{fallback_ctx}\n\nResposta:"},
            ]
            final = _call_chat(messages_fb, MAX_FINAL_TOKENS)
        except Exception as e2:
            # deixa subir para o caller usar fallback local
            raise

    return final.strip()

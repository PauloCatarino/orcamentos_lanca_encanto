from __future__ import annotations

import ctypes
import importlib
import os
import re
import shutil
import subprocess
import tempfile
import time
import unicodedata
import warnings
from dataclasses import dataclass
from ctypes import wintypes
from pathlib import Path
from typing import Callable, Optional, Sequence

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from Martelo_Orcamentos_V2.app.models import Producao
from Martelo_Orcamentos_V2.app.services.settings import get_setting


KEY_CUTRITE_EXE_PATH = "cutrite_exe_path"
KEY_CUTRITE_WORKDIR_PATH = "cutrite_workdir_path"
KEY_CUTRITE_DATA_PATH = "cutrite_data_path"
LISTA_MATERIAL_PREFIX = "Lista_Material_"
LISTAGEM_CUT_RITE_SHEET = "LISTAGEM_CUT_RITE"
DEFAULT_CUTRITE_EXE_PATHS = (
    r"C:\V12\V12.exe",
    r"C:\Program Files\Cut Rite Modular V12\V12.exe",
    r"C:\Program Files (x86)\Cut Rite Modular V12\V12.exe",
)
CUTRITE_SOURCE_MAX_COL = 25
CUTRITE_SOURCE_HEADER_ROW = 2
CUTRITE_SOURCE_DATA_START_ROW = 3
CUTRITE_EXPORT_HEADERS = (
    "Descricao",
    "Material",
    "Comp",
    "Larg",
    "Qt",
    "Veio",
    "Orla",
    "Cliente",
    "Ref_Cliente",
    "Processo",
    "Artigo",
    "Notas",
    "Esp",
    "Grafico Orlas",
    "Orla ESQ",
    "Orla DIR",
    "Orla CIMA",
    "Orla BAIXO",
    "ID",
    "CNC_1",
    "CNC_2",
    "+comp",
    "+Larg",
)
CUTRITE_GRAFICO_ORLAS_VALUE = "111:111:000:000"
CUTRITE_EXPORT_MAX_COL = len(CUTRITE_EXPORT_HEADERS)
CUTRITE_MAIN_TITLE_FRAGMENT = "schnitt profit modular"
CUTRITE_PARTS_WINDOW_NAME = "lista de pecas"
CUTRITE_IMPORT_WINDOW_NAME = "importacao - pecas"
CUTRITE_CUTLIST_WINDOW_NAME = "lista de corte"
CUTRITE_SAVE_WINDOW_NAME = "guardar em baixo"
CUTRITE_ERROR_WINDOW_NAME = "erro"
CUTRITE_DESENHO_WINDOW_NAME = "desenho"
CUTRITE_CORE_OUTPUT_EXTENSIONS = (".ctt", ".mpd")
CUTRITE_AUX_OUTPUT_EXTENSIONS = (".brd", ".btc", ".prl", ".loc")
CUTRITE_PARTS_TOOLBAR_IMPORT_INDEX = 3
CUTRITE_PARTS_TOOLBAR_SAVE_INDEX = 5
CUTRITE_MAIN_IMPORT_TOOLBAR_INDEX = 4
CUTRITE_IMPORT_WINDOW_TOOLBAR_IMPORT_INDEX = 1
CUTRITE_MAIN_WINDOW_READY_DELAY_SECONDS = 3.0
CUTRITE_WINDOW_OPEN_DELAY_SECONDS = 3.0
CUTRITE_ACTION_DELAY_SECONDS = 0.7
CUTRITE_POST_FOCUS_DELAY_SECONDS = 0.5
CUTRITE_POST_PASTE_DELAY_SECONDS = 0.8
CUTRITE_CONFIRMATION_TIMEOUT_SECONDS = 2
CUTRITE_SAVE_WINDOW_DELAY_SECONDS = 1.2
CUTRITE_MAXIMIZE_RETRIES = 3
CUTRITE_FIRST_DATA_CELL_X_MIN = 66
CUTRITE_FIRST_DATA_CELL_X_RATIO = 0.032
CUTRITE_FIRST_DATA_CELL_TABLE_TOP_MIN = 110
CUTRITE_FIRST_DATA_CELL_TABLE_TOP_RATIO = 0.088
CUTRITE_FIRST_DATA_CELL_ROW_HEIGHT_MIN = 22
CUTRITE_FIRST_DATA_CELL_ROW_HEIGHT_RATIO = 0.021
CUTRITE_SAVE_BUTTON_X_MIN = 185
CUTRITE_SAVE_BUTTON_X_RATIO = 0.09
CUTRITE_SAVE_BUTTON_Y_MIN = 52
CUTRITE_SAVE_BUTTON_Y_RATIO = 0.048
SW_RESTORE = 9
SW_MAXIMIZE = 3
MONITOR_DEFAULTTONEAREST = 2
CUTRITE_MAXIMIZED_BOUNDS_TOLERANCE = 32


class MONITORINFO(ctypes.Structure):
    _fields_ = (
        ("cbSize", wintypes.DWORD),
        ("rcMonitor", wintypes.RECT),
        ("rcWork", wintypes.RECT),
        ("dwFlags", wintypes.DWORD),
    )


class WINDOWPLACEMENT(ctypes.Structure):
    _fields_ = (
        ("length", wintypes.UINT),
        ("flags", wintypes.UINT),
        ("showCmd", wintypes.UINT),
        ("ptMinPosition", wintypes.POINT),
        ("ptMaxPosition", wintypes.POINT),
        ("rcNormalPosition", wintypes.RECT),
    )


@dataclass(frozen=True)
class CutRiteImportContext:
    processo: Producao
    folder_path: Path
    source_workbook_path: Path
    import_workbook_path: Path
    cutrite_input_path: Path
    plan_name: str
    cutrite_exe_path: Path
    import_exe_path: Path
    cutrite_root: Path
    cutrite_profile_dir: Path
    cutrite_workdir: Path
    cutrite_data_dir: Path
    cutrite_target_data_dir: Path


@dataclass(frozen=True)
class CutRiteExecutionResult:
    source_workbook_path: Path
    import_workbook_path: Path
    cutrite_profile_dir: Path
    cutrite_workdir: Path
    cutrite_data_dir: Path
    cutrite_target_data_dir: Path
    generated_data_paths: tuple[Path, ...]
    cutrite_was_running: bool
    launched_cutrite: bool


def sanitize_cutrite_plan_name(plan_name: str) -> str:
    text = str(plan_name or "").strip()
    text = re.sub(r'[<>:"/\\\\|?*]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text


def resolve_cutrite_exe_path(db: Optional[Session]) -> Optional[str]:
    if db is not None and hasattr(db, "execute"):
        configured = (get_setting(db, KEY_CUTRITE_EXE_PATH, "") or "").strip()
        if configured and Path(configured).is_file():
            return configured

    env_value = (os.environ.get("CUTRITE_EXE_PATH") or "").strip()
    if env_value and Path(env_value).is_file():
        return env_value

    for path in DEFAULT_CUTRITE_EXE_PATHS:
        if Path(path).is_file():
            return path
    return None


def resolve_configured_cutrite_workdir(db: Optional[Session]) -> Optional[Path]:
    workdir, _data_dir = resolve_configured_cutrite_paths(db)
    return workdir


def resolve_configured_cutrite_data_dir(db: Optional[Session]) -> Optional[Path]:
    _workdir, data_dir = resolve_configured_cutrite_paths(db)
    return data_dir


def resolve_configured_cutrite_paths(db: Optional[Session]) -> tuple[Optional[Path], Optional[Path]]:
    if db is None or not hasattr(db, "execute"):
        return None, None
    raw_workdir = (get_setting(db, KEY_CUTRITE_WORKDIR_PATH, "") or "").strip()
    raw_data_dir = (get_setting(db, KEY_CUTRITE_DATA_PATH, "") or "").strip()
    workdir, data_dir, _warning = normalize_cutrite_path_inputs(raw_workdir, raw_data_dir)
    return workdir, data_dir


def normalize_cutrite_path_inputs(
    raw_workdir: str,
    raw_data_dir: str,
) -> tuple[Optional[Path], Optional[Path], Optional[str]]:
    workdir = _normalize_existing_dir_path(raw_workdir)
    data_dir = _normalize_existing_dir_path(raw_data_dir)

    workdir_is_profile = _is_cutrite_profile_dir(workdir)
    data_dir_is_profile = _is_cutrite_profile_dir(data_dir)

    if workdir_is_profile:
        return workdir, data_dir, None

    if data_dir_is_profile and data_dir is not None:
        warning = "As pastas CUT-RITE foram invertidas automaticamente para usar o perfil correto."
        return data_dir, workdir, warning

    return None, data_dir, None


def find_lista_material_workbook(folder_path: Path, *, nome_enc_imos: str = "") -> Path:
    if not folder_path.exists() or not folder_path.is_dir():
        raise ValueError(f"Pasta Servidor nao encontrada:\n{folder_path}")

    nome_enc_txt = str(nome_enc_imos or "").strip()
    exact_candidates = []
    if nome_enc_txt:
        for suffix in (".xlsm", ".xlsx", ".xls"):
            exact_candidates.append(folder_path / f"{LISTA_MATERIAL_PREFIX}{nome_enc_txt}{suffix}")
    for candidate in exact_candidates:
        if candidate.is_file():
            return candidate

    matches = sorted(
        (
            path
            for path in folder_path.iterdir()
            if path.is_file()
            and path.name.startswith(LISTA_MATERIAL_PREFIX)
            and path.suffix.lower() in {".xlsm", ".xlsx", ".xls"}
        ),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        if nome_enc_txt:
            raise ValueError(
                "Foram encontrados varios ficheiros 'Lista_Material_' na Pasta Servidor.\n\n"
                "Verifique qual e o ficheiro correto ou preencha o Nome Enc IMOS IX correspondente."
            )
        raise ValueError(
            "Existem varios ficheiros 'Lista_Material_' na Pasta Servidor.\n\n"
            "Indique o Nome Enc IMOS IX correto ou remova ambiguidades."
        )

    raise ValueError(
        "Nao foi encontrado nenhum ficheiro 'Lista_Material_*.xlsm' na Pasta Servidor.\n\n"
        "Crie primeiro o Excel em 'Lista Material_IMOS'."
    )


def build_cutrite_import_headers() -> list[str]:
    return list(CUTRITE_EXPORT_HEADERS)


def build_cutrite_import_row(values: Sequence[object]) -> list[object]:
    row = list(values[:CUTRITE_SOURCE_MAX_COL])
    if len(row) < CUTRITE_SOURCE_MAX_COL:
        row.extend([None] * (CUTRITE_SOURCE_MAX_COL - len(row)))
    return [_normalize_cutrite_cell(value) for value in row]


def prepare_cutrite_import(
    session: Session,
    *,
    current_id: Optional[int],
    pasta_servidor: str,
    nome_plano_cut_rite: str,
    nome_enc_imos: str,
) -> CutRiteImportContext:
    if not current_id:
        raise ValueError("Selecione um processo.")

    processo = session.get(Producao, int(current_id))
    if processo is None:
        raise ValueError("Processo nao encontrado.")

    pasta_txt = str(pasta_servidor or "").strip()
    if not pasta_txt:
        raise ValueError("Pasta Servidor em falta.\n\nUse 'Criar Pasta' para gerar a pasta da obra no servidor.")
    folder_path = Path(pasta_txt)

    source_workbook_path = find_lista_material_workbook(folder_path, nome_enc_imos=nome_enc_imos)

    plan_name = sanitize_cutrite_plan_name(nome_plano_cut_rite)
    if not plan_name:
        raise ValueError("Nome Plano CUT-RITE em falta.")

    cutrite_exe = resolve_cutrite_exe_path(session)
    if not cutrite_exe:
        raise ValueError(
            "Executavel CUT-RITE nao encontrado.\n\n"
            "Configure o caminho em Configuracoes -> Geral -> Executavel CUT-RITE."
        )

    cutrite_exe_path = Path(cutrite_exe)
    cutrite_root = cutrite_exe_path.parent
    import_exe_path = cutrite_root / "Import.exe"
    if not import_exe_path.is_file():
        raise ValueError(f"Import.exe nao encontrado na instalacao CUT-RITE:\n{import_exe_path}")
    cutrite_profile_dir = resolve_cutrite_workdir(cutrite_root, db=session)
    cutrite_workdir = prepare_cutrite_runtime_workdir(cutrite_root, cutrite_profile_dir)
    cutrite_input_dir = resolve_cutrite_input_dir(cutrite_profile_dir)
    cutrite_data_dir = resolve_cutrite_data_dir(cutrite_profile_dir)
    cutrite_target_data_dir = resolve_configured_cutrite_data_dir(session) or cutrite_data_dir
    cutrite_input_dir.mkdir(parents=True, exist_ok=True)

    import_workbook_path = folder_path / f"{plan_name}.xls"
    cutrite_input_path = cutrite_input_dir / import_workbook_path.name

    return CutRiteImportContext(
        processo=processo,
        folder_path=folder_path,
        source_workbook_path=source_workbook_path,
        import_workbook_path=import_workbook_path,
        cutrite_input_path=cutrite_input_path,
        plan_name=plan_name,
        cutrite_exe_path=cutrite_exe_path,
        import_exe_path=import_exe_path,
        cutrite_root=cutrite_root,
        cutrite_profile_dir=cutrite_profile_dir,
        cutrite_workdir=cutrite_workdir,
        cutrite_data_dir=cutrite_data_dir,
        cutrite_target_data_dir=cutrite_target_data_dir,
    )


def build_cutrite_import_workbook(
    context: CutRiteImportContext,
    *,
    timeout_seconds: int = 240,
) -> tuple[Path, list[list[object]], Optional[str]]:
    headers, rows, macro_paste_text = _load_cutrite_source_table(context.source_workbook_path)
    if not rows:
        raise ValueError("A folha LISTAGEM_CUT_RITE nao contem linhas validas para importar.")
    export_rows = [_select_cutrite_export_columns(row) for row in rows]
    normalized_paste_text = _normalize_cutrite_macro_paste_text(macro_paste_text)
    if not normalized_paste_text:
        normalized_paste_text = _build_cutrite_paste_text(export_rows)
    return context.import_workbook_path, export_rows, normalized_paste_text


def execute_cutrite_import(
    context: CutRiteImportContext,
    *,
    timeout_seconds: int = 120,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> CutRiteExecutionResult:
    _report_cutrite_progress(progress_callback, "A preparar a listagem CUT-RITE.")
    import_workbook_path, export_rows, macro_paste_text = build_cutrite_import_workbook(context)
    paste_text = macro_paste_text or _build_cutrite_paste_text(export_rows)
    _report_cutrite_progress(progress_callback, "A iniciar a automacao do CUT-RITE.")
    cutrite_was_running, launched_cutrite = _execute_cutrite_ui_import(
        context,
        paste_text=paste_text,
        timeout_seconds=max(timeout_seconds, 180),
        progress_callback=progress_callback,
    )
    _report_cutrite_progress(progress_callback, "A validar os ficheiros gerados pelo plano.")
    generated_data_paths = _sync_cutrite_generated_data(context)
    _report_cutrite_progress(progress_callback, "Plano CUT-RITE concluido.")

    return CutRiteExecutionResult(
        source_workbook_path=context.source_workbook_path,
        import_workbook_path=context.import_workbook_path,
        cutrite_profile_dir=context.cutrite_profile_dir,
        cutrite_workdir=context.cutrite_workdir,
        cutrite_data_dir=context.cutrite_data_dir,
        cutrite_target_data_dir=context.cutrite_target_data_dir,
        generated_data_paths=tuple(generated_data_paths),
        cutrite_was_running=cutrite_was_running,
        launched_cutrite=launched_cutrite,
    )


def _execute_cutrite_ui_import(
    context: CutRiteImportContext,
    *,
    paste_text: str,
    timeout_seconds: int,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> tuple[bool, bool]:
    Desktop, mouse, keyboard = _load_cutrite_ui_modules()
    desktop = Desktop(backend="uia")
    main_window = _find_cutrite_main_window(desktop)
    cutrite_was_running = main_window is not None
    launched_cutrite = False
    if main_window is None:
        _report_cutrite_progress(progress_callback, "A abrir o programa CUT-RITE.")
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
        subprocess.Popen(
            [str(context.cutrite_exe_path)],
            cwd=str(context.cutrite_workdir),
            creationflags=creationflags,
        )
        launched_cutrite = True
        _report_cutrite_progress(progress_callback, "A aguardar a janela principal do CUT-RITE.")
        main_window = _wait_for_cutrite_window(
            lambda: _find_cutrite_main_window(Desktop(backend="uia")),
            timeout_seconds=45,
            error_message="Nao foi possivel abrir a janela principal do CUT-RITE.",
        )
    else:
        _report_cutrite_progress(progress_callback, "CUT-RITE ja estava em execucao.")
    _activate_cutrite_window(main_window)
    _sleep_cutrite(CUTRITE_MAIN_WINDOW_READY_DELAY_SECONDS)
    _report_cutrite_progress(progress_callback, "A maximizar a janela principal do CUT-RITE.")
    _maximize_cutrite_window(main_window, window_label="principal do CUT-RITE")
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)

    _report_cutrite_progress(progress_callback, "A abrir a janela Lista de pecas.")
    parts_window = _ensure_cutrite_parts_window(desktop, main_window, mouse, timeout_seconds=25)
    _activate_cutrite_window(parts_window)
    _sleep_cutrite(CUTRITE_WINDOW_OPEN_DELAY_SECONDS)
    _wait_for_cutrite_parts_window_ready(parts_window, timeout_seconds=12)
    _report_cutrite_progress(progress_callback, "A maximizar a janela Lista de pecas.")
    _maximize_cutrite_window(parts_window, window_label="Lista de pecas do CUT-RITE")
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    _report_cutrite_progress(progress_callback, "A preparar o clipboard com a lista de pecas.")
    _set_cutrite_clipboard_text(paste_text)
    _report_cutrite_progress(progress_callback, "A definir o titulo do plano no CUT-RITE.")
    _set_cutrite_parts_title(parts_window, context.plan_name, keyboard)
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    _report_cutrite_progress(progress_callback, "A navegar para a tabela de pecas.")
    _navigate_to_cutrite_parts_table(parts_window, keyboard)
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    _report_cutrite_progress(progress_callback, "A colar a lista de pecas no CUT-RITE.")
    keyboard.send_keys("^v", pause=0.08)
    _report_cutrite_progress(progress_callback, "A confirmar a colagem multipla, se necessario.")
    _dismiss_cutrite_confirmation_dialog(
        desktop,
        main_window,
        mouse,
        button_name="sim",
        timeout_seconds=CUTRITE_CONFIRMATION_TIMEOUT_SECONDS,
    )
    _report_cutrite_progress(progress_callback, "A aguardar a colagem dos dados.")
    _sleep_cutrite(5.0)  # Esperar 5 segundos para colar os dados
    _report_cutrite_progress(progress_callback, "A maximizar a janela Lista de pecas para gravar.")
    maximized_for_save = _maximize_cutrite_window(parts_window, window_label="Lista de pecas do CUT-RITE")
    if not maximized_for_save:
        try:
            keyboard.send_keys("{LWIN down}{UP}{LWIN up}", pause=0.05)
        except Exception:
            pass
        _sleep_cutrite(1.2)
        _maximize_cutrite_window(parts_window, window_label="Lista de pecas do CUT-RITE")
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    _report_cutrite_progress(progress_callback, "A gravar a lista de pecas.")
    save_confirmed = _save_cutrite_parts_list(
        parts_window,
        desktop,
        main_window,
        mouse,
        keyboard,
        timeout_seconds=30,
    )
    if not save_confirmed:
        _report_cutrite_progress(
            progress_callback,
            "Nao foi possivel confirmar a janela Guardar em baixo. A validar diretamente os ficheiros do plano.",
        )
    _report_cutrite_progress(progress_callback, "A aguardar os ficheiros finais do plano.")
    _wait_for_cutrite_output_files(context, timeout_seconds=75)
    return cutrite_was_running, launched_cutrite


def _load_cutrite_ui_modules():
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Revert to STA COM threading mode")
            pywinauto_module = importlib.import_module("pywinauto")
            mouse_module = importlib.import_module("pywinauto.mouse")
            keyboard_module = importlib.import_module("pywinauto.keyboard")
    except Exception as exc:
        raise RuntimeError(
            "A automacao da janela CUT-RITE requer o pacote 'pywinauto' instalado.\n\n"
            "Instale-o no Python utilizado pelo Martelo ou gere o executavel com o build atualizado."
        ) from exc
    return pywinauto_module.Desktop, mouse_module, keyboard_module


def _find_cutrite_main_window(desktop) -> Optional[object]:
    for window in desktop.windows():
        title = _normalize_cutrite_ui_text(window.window_text())
        if CUTRITE_MAIN_TITLE_FRAGMENT in title:
            return window
    return None


def _open_cutrite_import_window(main_window, mouse, *, timeout_seconds: int, desktop=None) -> object:
    existing = _find_cutrite_window(desktop, main_window, CUTRITE_IMPORT_WINDOW_NAME)
    if existing is not None:
        return existing

    if _click_cutrite_main_import_button(main_window, mouse):
        import_window = _wait_for_cutrite_window(
            lambda: _find_cutrite_window(desktop, main_window, CUTRITE_IMPORT_WINDOW_NAME),
            timeout_seconds=timeout_seconds,
            error_message="",
            raise_on_timeout=False,
        )
        if import_window is not None:
            return import_window

    parts_window = _ensure_cutrite_parts_window(desktop, main_window, mouse, timeout_seconds=timeout_seconds)
    _maximize_cutrite_window(parts_window)
    import_window = _trigger_cutrite_import(parts_window, main_window, mouse, timeout_seconds=15)
    if import_window is not None:
        return import_window

    raise RuntimeError("Nao foi possivel abrir a janela 'Importacao - Pecas' no CUT-RITE.")


def _click_cutrite_main_import_button(main_window, mouse) -> bool:
    for button in main_window.descendants(control_type="Button"):
        text = _collect_cutrite_element_text(button)
        if CUTRITE_IMPORT_WINDOW_NAME in text:
            _click_cutrite_element(button, mouse)
            return True

    try:
        _click_cutrite_toolbar_button(main_window, CUTRITE_MAIN_IMPORT_TOOLBAR_INDEX, mouse)
        return True
    except Exception:
        return False


def _ensure_cutrite_parts_window(desktop, main_window, mouse, *, timeout_seconds: int) -> object:
    existing = _find_cutrite_window(desktop, main_window, CUTRITE_PARTS_WINDOW_NAME)
    if existing is not None:
        return existing

    _activate_cutrite_window(main_window)
    _sleep_cutrite(CUTRITE_WINDOW_OPEN_DELAY_SECONDS)
    part_buttons = _find_cutrite_parts_buttons(main_window)
    if not part_buttons:
        raise RuntimeError("Nao foi encontrado o botao 'Pecas' na janela principal do CUT-RITE.")

    for button in part_buttons:
        _click_cutrite_element(button, mouse)
        _sleep_cutrite(CUTRITE_WINDOW_OPEN_DELAY_SECONDS)
        existing = _wait_for_cutrite_window(
            lambda: _find_cutrite_window(desktop, main_window, CUTRITE_PARTS_WINDOW_NAME),
            timeout_seconds=max(6, timeout_seconds),
            error_message="",
            raise_on_timeout=False,
        )
        if existing is not None:
            return existing

    raise RuntimeError("Nao foi possivel abrir a janela 'Lista de pecas' no CUT-RITE.")


def _wait_for_cutrite_parts_window_ready(parts_window, *, timeout_seconds: int) -> None:
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        try:
            edit_controls = [
                child
                for child in parts_window.descendants()
                if child.element_info.control_type == "Edit"
            ]
        except Exception:
            edit_controls = []
        if edit_controls:
            return
        _activate_cutrite_window(parts_window)
        _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)


def _trigger_cutrite_import(parts_window, main_window, mouse, *, timeout_seconds: int) -> Optional[object]:
    _click_cutrite_toolbar_button(parts_window, CUTRITE_PARTS_TOOLBAR_IMPORT_INDEX, mouse)
    return _wait_for_cutrite_window(
        lambda: _find_cutrite_descendant_window(main_window, CUTRITE_IMPORT_WINDOW_NAME),
        timeout_seconds=timeout_seconds,
        error_message="Nao foi possivel abrir a janela de importacao CUT-RITE.",
        raise_on_timeout=False,
    )


def _select_cutrite_import_file(import_window, filename: str) -> None:
    filename_norm = _normalize_cutrite_ui_text(filename)
    list_controls = [
        child
        for child in import_window.descendants()
        if child.element_info.control_type in {"List", "Table", "DataGrid", "Tree"}
    ]
    if not list_controls:
        raise RuntimeError("A janela de importacao CUT-RITE nao expos a lista de ficheiros.")
    for control in list_controls:
        for item in control.descendants():
            if item.element_info.control_type not in {"ListItem", "DataItem", "TreeItem"}:
                continue
            if _normalize_cutrite_ui_text(item.element_info.name) == filename_norm:
                _click_cutrite_element(item, None)
                return
    raise RuntimeError(f"O ficheiro CUT-RITE nao apareceu na lista de importacao:\n{filename}")


def _submit_cutrite_import_window(import_window, mouse) -> None:
    if _click_cutrite_named_button_if_present(import_window, "ok", mouse):
        return
    _click_cutrite_toolbar_button(import_window, CUTRITE_IMPORT_WINDOW_TOOLBAR_IMPORT_INDEX, mouse)


def _dismiss_cutrite_error_dialog(main_window, mouse, *, timeout_seconds: int) -> None:
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        error_window = _find_cutrite_descendant_window(main_window, CUTRITE_ERROR_WINDOW_NAME)
        if error_window is None:
            return
        if not _click_cutrite_named_button_if_present(error_window, "ok", mouse):
            return
        time.sleep(0.3)


def _close_cutrite_desenho_window(main_window, mouse, *, timeout_seconds: int) -> bool:
    deadline = time.monotonic() + timeout_seconds
    closed_any = False
    while time.monotonic() < deadline:
        desenho_window = _find_cutrite_descendant_window(main_window, CUTRITE_DESENHO_WINDOW_NAME)
        if desenho_window is None:
            return closed_any
        closed_any = True
        try:
            desenho_window.close()
        except Exception:
            rect = desenho_window.rectangle()
            mouse.click(
                button="left",
                coords=(rect.right - 16, rect.top + 16),
            )
        _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    return closed_any


def _set_cutrite_parts_title(parts_window, plan_name: str, keyboard) -> None:
    edit_controls = [child for child in parts_window.descendants() if child.element_info.control_type == "Edit"]
    if not edit_controls:
        raise RuntimeError("Nao foi encontrado o campo 'Titulo' na janela 'Lista de pecas'.")
    title_field = edit_controls[0]
    try:
        title_field.set_focus()
        _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
        title_field.set_edit_text(plan_name)
    except Exception:
        _click_cutrite_element(title_field, None)
        keyboard.send_keys("^a{BACKSPACE}", pause=0.02)
        keyboard.send_keys(plan_name, with_spaces=True, pause=0.02)
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)


def _navigate_to_cutrite_parts_table(parts_window, keyboard) -> None:
    # Usar SHIFT+ENTER 3 vezes para navegar para a tabela
    for _ in range(3):
        keyboard.send_keys("+{ENTER}", pause=0.1)
        _sleep_cutrite(0.1)


def _set_cutrite_clipboard_text(text: str, *, retries: int = 10, delay_seconds: float = 0.2) -> None:
    errors: list[Exception] = []

    try:
        _set_cutrite_clipboard_text_win32(text, retries=retries, delay_seconds=delay_seconds)
        return
    except Exception as exc:
        errors.append(exc)

    try:
        _set_cutrite_clipboard_text_qt(text)
        return
    except Exception as exc:
        errors.append(exc)

    detail = ""
    if errors:
        detail = f" Ultimo erro: {errors[-1]}"
    raise RuntimeError(
        "Nao foi possivel preparar o clipboard para colar as pecas no CUT-RITE."
        f"{detail}"
    ) from (errors[-1] if errors else None)


def _set_cutrite_clipboard_text_win32(
    text: str,
    *,
    retries: int = 10,
    delay_seconds: float = 0.2,
) -> None:
    win32clipboard = importlib.import_module("win32clipboard")

    last_error: Optional[Exception] = None
    for attempt in range(retries):
        try:
            win32clipboard.OpenClipboard()
            try:
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardText(text, win32clipboard.CF_UNICODETEXT)
            finally:
                win32clipboard.CloseClipboard()
            return
        except Exception as exc:
            last_error = exc
            if attempt == retries - 1:
                break
            time.sleep(delay_seconds)
    if last_error is not None:
        raise last_error


def _set_cutrite_clipboard_text_qt(text: str) -> None:
    from PySide6 import QtGui, QtWidgets

    app = QtGui.QGuiApplication.instance() or QtWidgets.QApplication.instance()
    if app is None:
        raise RuntimeError("QApplication indisponivel.")

    clipboard = app.clipboard()
    if clipboard is None:
        raise RuntimeError("Clipboard Qt indisponivel.")

    clipboard.clear(mode=clipboard.Clipboard)
    clipboard.setText(text, mode=clipboard.Clipboard)
    app.processEvents()

    current = clipboard.text(mode=clipboard.Clipboard)
    if current != text:
        raise RuntimeError("Clipboard Qt nao confirmou o texto esperado.")


def _focus_cutrite_first_data_cell(parts_window, mouse, keyboard) -> None:
    """Foca a primeira celula de dados na coluna Descricao da lista de pecas."""

    _activate_cutrite_window(parts_window)
    used_fallback_target = False

    global_element = None
    row_one_element = None
    description_header = None
    try:
        for child in parts_window.descendants():
            if child.element_info.control_type not in {"Text", "DataItem", "Edit"}:
                continue
            child_name = _normalize_cutrite_ui_text(child.element_info.name)
            if child_name == "descricao" and description_header is None:
                description_header = child
            elif child_name == "1." and row_one_element is None:
                row_one_element = child
            elif "global" in child_name and global_element is None:
                global_element = child
    except Exception:
        global_element = None
        row_one_element = None
        description_header = None

    if description_header is not None and row_one_element is not None:
        header_rect = description_header.rectangle()
        row_rect = row_one_element.rectangle()
        x = header_rect.left + max(14, min(28, header_rect.width() // 6))
        y = row_rect.top + max(6, row_rect.height() // 2)
        mouse.click(button="left", coords=(x, y))
        _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)
        mouse.click(button="left", coords=(x, y))
        return

    if description_header is not None and global_element is not None:
        header_rect = description_header.rectangle()
        global_rect = global_element.rectangle()
        row_height = max(22, global_rect.height())
        x = header_rect.left + max(14, min(28, header_rect.width() // 6))
        y = global_rect.bottom + max(6, row_height // 2)
        mouse.click(button="left", coords=(x, y))
        _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)
        mouse.click(button="left", coords=(x, y))
        used_fallback_target = True
    elif global_element is not None:
        rect = global_element.rectangle()
        row_height = max(22, rect.height())
        x = rect.right + 24
        y = rect.bottom + max(6, row_height // 2)
        mouse.click(button="left", coords=(x, y))
        _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)
        mouse.click(button="left", coords=(x, y))
        used_fallback_target = True
    else:
        table = None
        try:
            for ctrl_type in ("Table", "DataGrid", "List", "ListView"):
                candidates = [
                    child
                    for child in parts_window.descendants(control_type=ctrl_type)
                    if getattr(child, "element_info", None) is not None
                ]
                if candidates:
                    table = candidates[0]
                    break
        except Exception:
            table = None

        if table is not None:
            rect = table.rectangle()
            x = rect.left + max(CUTRITE_FIRST_DATA_CELL_X_MIN, int(rect.width() * CUTRITE_FIRST_DATA_CELL_X_RATIO))
            table_top = rect.top + max(
                CUTRITE_FIRST_DATA_CELL_TABLE_TOP_MIN,
                int(rect.height() * CUTRITE_FIRST_DATA_CELL_TABLE_TOP_RATIO),
            )
            row_height = max(
                CUTRITE_FIRST_DATA_CELL_ROW_HEIGHT_MIN,
                int(rect.height() * CUTRITE_FIRST_DATA_CELL_ROW_HEIGHT_RATIO),
            )
            y = table_top + row_height + row_height // 2
        else:
            rect = parts_window.rectangle()
            x = rect.left + max(CUTRITE_FIRST_DATA_CELL_X_MIN, int(rect.width() * CUTRITE_FIRST_DATA_CELL_X_RATIO))
            table_top = rect.top + max(
                CUTRITE_FIRST_DATA_CELL_TABLE_TOP_MIN,
                int(rect.height() * CUTRITE_FIRST_DATA_CELL_TABLE_TOP_RATIO),
            )
            row_height = max(
                CUTRITE_FIRST_DATA_CELL_ROW_HEIGHT_MIN,
                int(rect.height() * CUTRITE_FIRST_DATA_CELL_ROW_HEIGHT_RATIO),
            )
            y = table_top + row_height + row_height // 2
        mouse.click(button="left", coords=(x, y))
        _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)
        mouse.click(button="left", coords=(x, y))
        used_fallback_target = True

    if used_fallback_target:
        try:
            keyboard.send_keys("{DOWN}", pause=0.02)
        except Exception:
            pass
        _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)

def _dismiss_cutrite_confirmation_dialog(desktop, main_window, mouse, *, button_name: str, timeout_seconds: int) -> None:
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        dialog = _find_cutrite_window_with_button(main_window, button_name) or _find_cutrite_desktop_window_with_button(
            desktop,
            button_name,
        )
        if dialog is None:
            return
        if not _click_cutrite_named_button_if_present(dialog, button_name, mouse):
            return
        _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)
        return


def _find_cutrite_window_with_button(root_window, button_name: str) -> Optional[object]:
    needle = _normalize_cutrite_ui_text(button_name)
    for element in root_window.descendants():
        try:
            if element.element_info.control_type != "Window":
                continue
            buttons = [
                child
                for child in element.children()
                if child.element_info.control_type == "Button"
                and _normalize_cutrite_ui_text(child.element_info.name) == needle
            ]
        except Exception:
            continue
        if buttons:
            return element
    return None


def _find_cutrite_desktop_window_with_button(desktop, button_name: str) -> Optional[object]:
    if desktop is None:
        return None
    needle = _normalize_cutrite_ui_text(button_name)
    for window in desktop.windows():
        try:
            buttons = [
                child
                for child in window.children()
                if child.element_info.control_type == "Button"
                and _normalize_cutrite_ui_text(child.element_info.name) == needle
            ]
        except Exception:
            continue
        if buttons:
            return window
    return None


def _save_cutrite_parts_list(parts_window, desktop, main_window, mouse, keyboard, *, timeout_seconds: int) -> bool:
    _activate_cutrite_window(parts_window)
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    try:
        keyboard.send_keys("{ESC}", pause=0.02)
    except Exception:
        pass
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    save_window = _save_cutrite_cutlist(
        parts_window,
        desktop,
        main_window,
        mouse,
        keyboard,
        timeout_seconds=timeout_seconds,
    )
    if save_window is None:
        return False
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    _click_cutrite_named_button(save_window, "ok", mouse)
    _sleep_cutrite(CUTRITE_WINDOW_OPEN_DELAY_SECONDS)
    return True


def _save_cutrite_cutlist(cutlist_window, desktop, main_window, mouse, keyboard, *, timeout_seconds: int) -> Optional[object]:
    _activate_cutrite_window(cutlist_window)
    _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
    attempts: list[Callable[[], None]] = [
        *[
            (lambda button=button: _click_cutrite_element(button, mouse))
            for button in _find_cutrite_save_button_candidates(cutlist_window)
        ],
        lambda: _click_cutrite_save_via_toolbar(cutlist_window, mouse),
        lambda: _click_cutrite_save_fallback(cutlist_window, mouse),
        lambda: _click_cutrite_save_fallback(cutlist_window, mouse),
        lambda: keyboard.send_keys("^s", pause=0.05),
    ]
    per_attempt_timeout = min(max(timeout_seconds // max(len(attempts), 1), 3), 8)
    for attempt in attempts:
        try:
            attempt()
        except Exception:
            continue
        _sleep_cutrite(CUTRITE_SAVE_WINDOW_DELAY_SECONDS)
        save_window = _wait_for_cutrite_window(
            lambda: _find_cutrite_window_with_button(main_window, "ok")
            or _find_cutrite_window(desktop, main_window, CUTRITE_SAVE_WINDOW_NAME)
            or _find_cutrite_desktop_window(desktop, CUTRITE_SAVE_WINDOW_NAME),
            timeout_seconds=per_attempt_timeout,
            error_message="",
            raise_on_timeout=False,
        )
        if save_window is not None:
            return save_window
        _activate_cutrite_window(cutlist_window)
        _sleep_cutrite(CUTRITE_WINDOW_OPEN_DELAY_SECONDS)

    return None


def _maximize_cutrite_window(window, *, window_label: str = "CUT-RITE") -> bool:
    window_name = getattr(getattr(window, "element_info", None), "name", "") or window_label
    if _is_cutrite_window_maximized(window) or _looks_cutrite_window_maximized(window):
        return True
    for _attempt in range(CUTRITE_MAXIMIZE_RETRIES):
        _activate_cutrite_window(window)
        handle = _get_cutrite_window_handle(window)
        if _is_cutrite_window_maximized(window) or _looks_cutrite_window_maximized(window):
            return True
        if handle and _is_cutrite_window_minimized(handle):
            try:
                ctypes.windll.user32.ShowWindow(handle, SW_RESTORE)
            except Exception:
                pass
            try:
                window.restore()
            except Exception:
                pass
        _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
        if handle:
            try:
                ctypes.windll.user32.ShowWindow(handle, SW_MAXIMIZE)
            except Exception:
                pass
        try:
            window.maximize()
        except Exception:
            pass
        _activate_cutrite_window(window)
        _sleep_cutrite(CUTRITE_ACTION_DELAY_SECONDS)
        if _is_cutrite_window_maximized(window) or _looks_cutrite_window_maximized(window):
            return True
    return False


def _activate_cutrite_window(window) -> None:
    handle = _get_cutrite_window_handle(window)
    if handle and _is_cutrite_window_minimized(handle):
        try:
            ctypes.windll.user32.ShowWindow(handle, SW_RESTORE)
        except Exception:
            pass
        try:
            window.restore()
        except Exception:
            pass
        _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)
    try:
        window.set_focus()
    except Exception:
        pass
    _sleep_cutrite(CUTRITE_POST_FOCUS_DELAY_SECONDS)


def _sleep_cutrite(seconds: float) -> None:
    time.sleep(seconds)


def _wait_for_cutrite_output_files(
    context: CutRiteImportContext,
    *,
    timeout_seconds: int,
) -> tuple[Path, ...]:
    expected_core_paths = tuple(
        context.cutrite_data_dir / f"{context.plan_name}{ext}" for ext in CUTRITE_CORE_OUTPUT_EXTENSIONS
    )
    deadline = time.monotonic() + timeout_seconds
    last_snapshot: Optional[tuple[tuple[str, int, int], ...]] = None
    stable_since: Optional[float] = None
    last_found_paths: tuple[Path, ...] = tuple()

    while time.monotonic() < deadline:
        found_paths = sorted(context.cutrite_data_dir.glob(f"{context.plan_name}.*"))
        last_found_paths = tuple(found_paths)
        if all(path.exists() for path in expected_core_paths):
            snapshot = tuple(
                (path.name, path.stat().st_size, int(path.stat().st_mtime))
                for path in found_paths
            )
            if snapshot == last_snapshot:
                if stable_since is None:
                    stable_since = time.monotonic()
                elif time.monotonic() - stable_since >= 1.5:
                    return tuple(found_paths)
            else:
                last_snapshot = snapshot
                stable_since = time.monotonic()
        time.sleep(0.5)

    if all(path.exists() for path in expected_core_paths):
        return last_found_paths

    found_names = [path.name for path in last_found_paths]
    missing = [path.name for path in expected_core_paths if not path.exists()]
    raise RuntimeError(
        "A importacao CUT-RITE terminou mas nao ficaram gravados os ficheiros base esperados.\n\n"
        f"Em falta: {', '.join(missing) or 'nenhum'}\n"
        f"Encontrados: {', '.join(found_names) or 'nenhum'}"
    )


def _find_cutrite_descendant_window(
    root_window,
    name_fragment: str,
    *,
    suffix_required: bool = False,
    extra_fragment: Optional[str] = None,
) -> Optional[object]:
    needle = _normalize_cutrite_ui_text(name_fragment)
    extra = _normalize_cutrite_ui_text(extra_fragment) if extra_fragment else ""
    for element in root_window.descendants():
        try:
            if element.element_info.control_type != "Window":
                continue
            name = _normalize_cutrite_ui_text(element.element_info.name)
        except Exception:
            continue
        if needle not in name:
            continue
        if suffix_required and name == needle:
            continue
        if extra and extra not in name:
            continue
        return element
    return None


def _find_cutrite_desktop_window(
    desktop,
    name_fragment: str,
    *,
    suffix_required: bool = False,
    extra_fragment: Optional[str] = None,
) -> Optional[object]:
    if desktop is None:
        return None
    needle = _normalize_cutrite_ui_text(name_fragment)
    extra = _normalize_cutrite_ui_text(extra_fragment) if extra_fragment else ""
    for window in desktop.windows():
        try:
            name = _normalize_cutrite_ui_text(window.window_text())
        except Exception:
            try:
                name = _normalize_cutrite_ui_text(window.element_info.name)
            except Exception:
                continue
        if needle not in name:
            continue
        if suffix_required and name == needle:
            continue
        if extra and extra not in name:
            continue
        return window
    return None


def _find_cutrite_window(
    desktop,
    root_window,
    name_fragment: str,
    *,
    suffix_required: bool = False,
    extra_fragment: Optional[str] = None,
) -> Optional[object]:
    return _find_cutrite_descendant_window(
        root_window,
        name_fragment,
        suffix_required=suffix_required,
        extra_fragment=extra_fragment,
    ) or _find_cutrite_desktop_window(
        desktop,
        name_fragment,
        suffix_required=suffix_required,
        extra_fragment=extra_fragment,
    )


def _find_cutrite_parts_buttons(main_window) -> list[object]:
    scored_buttons: list[tuple[int, object]] = []
    for button in main_window.descendants(control_type="Button"):
        score = _score_cutrite_parts_button(button)
        if score > 0:
            scored_buttons.append((score, button))
    scored_buttons.sort(key=lambda item: item[0], reverse=True)
    return [button for _score, button in scored_buttons]


def _score_cutrite_parts_button(button) -> int:
    text = _collect_cutrite_element_text(button)
    if not text or CUTRITE_IMPORT_WINDOW_NAME in text:
        return 0
    score = 0
    if CUTRITE_PARTS_WINDOW_NAME in text:
        score += 200
    if "lista pecas" in text:
        score += 120
    if text == "pecas":
        score += 90
    elif text.startswith("pecas "):
        score += 60
    elif " pecas " in f" {text} ":
        score += 40
    if "pain" in text:
        score -= 100
    return score


def _collect_cutrite_element_text(element) -> str:
    parts: list[str] = []
    try:
        parts.append(str(element.element_info.name or ""))
    except Exception:
        pass
    for attr_name in ("legacy_name", "help_text", "automation_id", "class_name"):
        try:
            parts.append(str(getattr(element.element_info, attr_name, "") or ""))
        except Exception:
            pass
    return _normalize_cutrite_ui_text(" ".join(parts))


def _find_cutrite_save_button_candidates(window) -> list[object]:
    scored_buttons: list[tuple[int, object]] = []
    for button in window.descendants(control_type="Button"):
        text = _collect_cutrite_element_text(button)
        if not text:
            continue
        score = 0
        if "save" in text or "guardar" in text:
            score += 200
        if "disquete" in text:
            score += 120
        if "import" in text:
            score -= 100
        if score > 0:
            scored_buttons.append((score, button))
    scored_buttons.sort(key=lambda item: item[0], reverse=True)
    return [button for _score, button in scored_buttons]


def _click_cutrite_save_fallback(window, mouse) -> None:
    rect = window.rectangle()
    x = rect.left + max(CUTRITE_SAVE_BUTTON_X_MIN, int(rect.width() * CUTRITE_SAVE_BUTTON_X_RATIO))
    y = rect.top + max(CUTRITE_SAVE_BUTTON_Y_MIN, int(rect.height() * CUTRITE_SAVE_BUTTON_Y_RATIO))
    mouse.click(button="left", coords=(x, y))


def _click_cutrite_save_via_toolbar(window, mouse) -> None:
    toolbar_buttons = _get_cutrite_toolbar_buttons(window)
    for button in toolbar_buttons:
        if _normalize_cutrite_ui_text(button.element_info.name) == "save":
            _click_cutrite_element(button, mouse)
            return
    _click_cutrite_toolbar_button(window, CUTRITE_PARTS_TOOLBAR_SAVE_INDEX, mouse)


def _get_cutrite_window_handle(window) -> int:
    for attr_name in ("handle",):
        try:
            handle = int(getattr(window, attr_name, 0) or 0)
            if handle:
                return handle
        except Exception:
            continue
    try:
        handle = int(getattr(window.element_info, "handle", 0) or 0)
        if handle:
            return handle
    except Exception:
        pass
    return 0


def _is_cutrite_window_maximized(window) -> bool:
    handle = _get_cutrite_window_handle(window)
    if handle:
        try:
            return bool(ctypes.windll.user32.IsZoomed(handle))
        except Exception:
            pass
        try:
            placement = WINDOWPLACEMENT()
            placement.length = ctypes.sizeof(WINDOWPLACEMENT)
            if ctypes.windll.user32.GetWindowPlacement(handle, ctypes.byref(placement)):
                return placement.showCmd == SW_MAXIMIZE
        except Exception:
            pass
    try:
        return bool(window.is_maximized())
    except Exception:
        return False


def _is_cutrite_window_minimized(handle: int) -> bool:
    try:
        return bool(ctypes.windll.user32.IsIconic(handle))
    except Exception:
        return False


def _looks_cutrite_window_maximized(window) -> bool:
    handle = _get_cutrite_window_handle(window)
    if not handle:
        return False
    window_rect = _get_cutrite_window_rect(window, handle=handle)
    work_rect = _get_cutrite_monitor_work_rect(handle)
    if window_rect is None or work_rect is None:
        return False

    left, top, right, bottom = window_rect
    work_left, work_top, work_right, work_bottom = work_rect

    # TolerÃ¢ncia baseada no tamanho da monitorizaÃ§Ã£o, para evitar falsos negativos
    work_width = work_right - work_left
    work_height = work_bottom - work_top
    tolerance = max(
        CUTRITE_MAXIMIZED_BOUNDS_TOLERANCE,
        int(work_width * 0.05),
        int(work_height * 0.05),
    )

    return (
        abs(left - work_left) <= tolerance
        and abs(top - work_top) <= tolerance
        and abs(right - work_right) <= tolerance
        and abs(bottom - work_bottom) <= tolerance
    )


def _get_cutrite_window_rect(window, *, handle: Optional[int] = None) -> Optional[tuple[int, int, int, int]]:
    resolved_handle = handle if handle is not None else _get_cutrite_window_handle(window)
    if resolved_handle:
        try:
            rect = wintypes.RECT()
            if ctypes.windll.user32.GetWindowRect(resolved_handle, ctypes.byref(rect)):
                return rect.left, rect.top, rect.right, rect.bottom
        except Exception:
            pass
    try:
        rect = window.rectangle()
        return rect.left, rect.top, rect.right, rect.bottom
    except Exception:
        return None


def _get_cutrite_monitor_work_rect(handle: int) -> Optional[tuple[int, int, int, int]]:
    try:
        monitor = ctypes.windll.user32.MonitorFromWindow(handle, MONITOR_DEFAULTTONEAREST)
        if not monitor:
            return None
        info = MONITORINFO()
        info.cbSize = ctypes.sizeof(MONITORINFO)
        if not ctypes.windll.user32.GetMonitorInfoW(monitor, ctypes.byref(info)):
            return None
        rect = info.rcWork
        return rect.left, rect.top, rect.right, rect.bottom
    except Exception:
        return None


def _report_cutrite_progress(progress_callback: Optional[Callable[[str], None]], message: str) -> None:
    if progress_callback is None:
        return
    try:
        progress_callback(message)
    except Exception:
        pass


def _get_cutrite_toolbar_buttons(window) -> list[object]:
    toolbars = [child for child in window.children() if child.element_info.control_type == "ToolBar"]
    if not toolbars:
        raise RuntimeError(f"A janela CUT-RITE '{window.element_info.name}' nao expÃ´s a toolbar esperada.")
    return list(toolbars[0].children())


def _click_cutrite_toolbar_button(window, index: int, mouse) -> None:
    buttons = _get_cutrite_toolbar_buttons(window)
    if index >= len(buttons):
        raise RuntimeError(
            f"A toolbar CUT-RITE '{window.element_info.name}' nao tem o botao esperado na posicao {index}."
        )
    _click_cutrite_element(buttons[index], mouse)


def _click_cutrite_named_button(window, button_name: str, mouse) -> None:
    needle = _normalize_cutrite_ui_text(button_name)
    for child in window.children():
        if child.element_info.control_type != "Button":
            continue
        if _normalize_cutrite_ui_text(child.element_info.name) == needle:
            _click_cutrite_element(child, mouse)
            return
    raise RuntimeError(f"Nao foi encontrado o botao '{button_name}' na janela CUT-RITE '{window.element_info.name}'.")


def _click_cutrite_named_button_if_present(window, button_name: str, mouse) -> bool:
    needle = _normalize_cutrite_ui_text(button_name)
    for child in window.children():
        if child.element_info.control_type != "Button":
            continue
        if _normalize_cutrite_ui_text(child.element_info.name) == needle:
            _click_cutrite_element(child, mouse)
            return True
    return False


def _click_cutrite_element(element, mouse) -> None:
    try:
        element.click_input()
        return
    except Exception:
        pass
    rect = element.rectangle()
    point = (rect.left + rect.width() // 2, rect.top + rect.height() // 2)
    if mouse is None:
        raise
    mouse.click(button="left", coords=point)


def _wait_for_cutrite_window(factory, *, timeout_seconds: int, error_message: str, raise_on_timeout: bool = True):
    deadline = time.monotonic() + timeout_seconds
    last_value = None
    while time.monotonic() < deadline:
        try:
            last_value = factory()
        except Exception:
            last_value = None
        if last_value:
            return last_value
        time.sleep(0.5)
    if raise_on_timeout and error_message:
        raise RuntimeError(error_message)
    return None


def _normalize_cutrite_ui_text(value: object) -> str:
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def _normalize_cutrite_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"[\r\n]+", " ", value).strip()
    return value


def _build_cutrite_paste_text(rows: Sequence[Sequence[object]]) -> str:
    return "\r\n".join(
        "\t".join(_format_cutrite_paste_cell(value) for value in row[:CUTRITE_EXPORT_MAX_COL])
        for row in rows
    )


def _normalize_cutrite_macro_paste_text(text: Optional[str]) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    lines = [line.rstrip("\r\n") for line in raw.splitlines() if line.strip()]
    return "\r\n".join(lines)


def _format_cutrite_paste_cell(value: object) -> str:
    normalized = _normalize_cutrite_cell(value)
    if normalized in (None, ""):
        return ""
    if isinstance(normalized, float) and normalized.is_integer():
        return str(int(normalized))
    return str(normalized)


def _select_cutrite_export_columns(row: Sequence[object]) -> list[object]:
    return list(row[:CUTRITE_EXPORT_MAX_COL])


def prepare_cutrite_runtime_workdir(cutrite_root: Path, cutrite_profile_dir: Path) -> Path:
    if _is_local_path(cutrite_profile_dir):
        return cutrite_profile_dir

    local_base = Path(os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()) / "Martelo_Orcamentos_V2" / "cutrite_profiles"
    safe_name = sanitize_cutrite_plan_name(cutrite_profile_dir.name) or "cutrite_profile"
    local_base.mkdir(parents=True, exist_ok=True)
    runtime_dir = Path(tempfile.mkdtemp(prefix=f"{safe_name}_", dir=str(local_base)))

    for item in cutrite_profile_dir.iterdir():
        if item.is_file():
            shutil.copy2(item, runtime_dir / item.name)

    _ensure_cutrite_excel_import_profile(cutrite_root, runtime_dir)
    return runtime_dir


def _ensure_cutrite_excel_import_profile(cutrite_root: Path, runtime_dir: Path) -> None:
    _patch_cutrite_v12defs_for_excel_import(runtime_dir / "v12defs.ctl")
    for filename in ("Excel-XLS.upx", "Excel-CSV.upx"):
        target = runtime_dir / filename
        if target.exists():
            continue
        source = resolve_cutrite_import_template(cutrite_root, filename)
        if source is not None:
            shutil.copy2(source, target)


def resolve_cutrite_import_template(cutrite_root: Path, filename: str) -> Optional[Path]:
    candidates = [
        cutrite_root / "_WORK" / "USER1" / filename,
        cutrite_root / "_WORK" / "USER2" / filename,
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


def _patch_cutrite_v12defs_for_excel_import(v12defs_path: Path) -> None:
    if not v12defs_path.is_file():
        return
    text = v12defs_path.read_text(encoding="utf-8", errors="ignore")
    text = re.sub(r"(?m)^ImportPartFormat=.*$", "ImportPartFormat=10", text)
    text = re.sub(r"(?m)^ImportPartParams=.*$", "ImportPartParams=Excel-XLS", text)
    text = re.sub(r"(?m)^ShowImportFileDialog=.*$", "ShowImportFileDialog=1", text)
    v12defs_path.write_text(text, encoding="utf-8")


def _normalize_existing_dir_path(raw_value: str) -> Optional[Path]:
    text = str(raw_value or "").strip()
    if not text:
        return None
    path = Path(text)
    if path.is_dir():
        return path
    return None


def _is_cutrite_profile_dir(path: Optional[Path]) -> bool:
    return bool(path and path.is_dir() and (path / "systemv12.ctl").is_file())


def _is_local_path(path: Path) -> bool:
    return not str(path).startswith("\\\\")


def resolve_cutrite_workdir(cutrite_root: Path, db: Optional[Session] = None) -> Path:
    configured = resolve_configured_cutrite_workdir(db)
    if configured is not None:
        return configured

    if (cutrite_root / "systemv12.ctl").is_file():
        return cutrite_root

    work_root = cutrite_root / "_WORK"
    user_dirs = sorted(
        path
        for path in work_root.glob("USER*")
        if path.is_dir() and (path / "systemv12.ctl").is_file()
    )
    if len(user_dirs) > 1:
        choices = "\n".join(str(path) for path in user_dirs)
        raise ValueError(
            "Foram encontradas varias pastas de trabalho do CUT-RITE.\n\n"
            "Configure a pasta correta em Configuracoes -> Geral -> Pasta Trabalho CUT-RITE.\n\n"
            f"Opcoes detetadas:\n{choices}"
        )
    if user_dirs:
        return user_dirs[0]
    raise ValueError(
        "Nao foi encontrada nenhuma pasta de trabalho do CUT-RITE com 'systemv12.ctl'.\n\n"
        "Verifique a instalacao em Configuracoes -> Geral -> Executavel CUT-RITE."
    )


def resolve_cutrite_input_dir(cutrite_workdir: Path) -> Path:
    system_file = cutrite_workdir / "systemv12.ctl"
    if system_file.is_file():
        for line in system_file.read_text(encoding="utf-8", errors="ignore").splitlines():
            if line.upper().startswith("SYSIMPPATH,"):
                value = line.split(",", 1)[1].strip()
                if value:
                    return Path(value)
    return cutrite_workdir / "IMPORT"


def resolve_cutrite_data_dir(cutrite_workdir: Path) -> Path:
    system_file = cutrite_workdir / "systemv12.ctl"
    if system_file.is_file():
        for line in system_file.read_text(encoding="utf-8", errors="ignore").splitlines():
            if line.upper().startswith("SYSDATAPATH,"):
                value = line.split(",", 1)[1].strip()
                if value:
                    return Path(value)
    return cutrite_workdir.parent / "DATA"


def _sync_cutrite_generated_data(context: CutRiteImportContext) -> list[Path]:
    generated_paths = sorted(context.cutrite_data_dir.glob(f"{context.plan_name}.*"))
    if not generated_paths:
        return []
    if _paths_refer_to_same_location(context.cutrite_target_data_dir, context.cutrite_data_dir):
        return generated_paths

    context.cutrite_target_data_dir.mkdir(parents=True, exist_ok=True)
    synced_paths: list[Path] = []
    for source_path in generated_paths:
        target_path = context.cutrite_target_data_dir / source_path.name
        if _paths_refer_to_same_location(source_path, target_path):
            synced_paths.append(source_path)
            continue
        _copy_file_with_retry(source_path, target_path)
        synced_paths.append(target_path)
    return synced_paths


def _paths_refer_to_same_location(path_a: Path, path_b: Path) -> bool:
    norm_a = os.path.normcase(os.path.normpath(str(path_a)))
    norm_b = os.path.normcase(os.path.normpath(str(path_b)))
    if norm_a == norm_b:
        return True
    try:
        return os.path.samefile(path_a, path_b)
    except (FileNotFoundError, PermissionError, OSError):
        return False


def _copy_file_with_retry(source_path: Path, target_path: Path, *, retries: int = 8, delay_seconds: float = 0.5) -> None:
    last_error: Optional[Exception] = None
    for attempt in range(retries):
        try:
            shutil.copy2(source_path, target_path)
            return
        except PermissionError as exc:
            last_error = exc
            if getattr(exc, "winerror", None) != 32 or attempt == retries - 1:
                raise
            time.sleep(delay_seconds)
    if last_error is not None:
        raise last_error


def _load_cutrite_source_table(source_workbook_path: Path) -> tuple[list[object], list[list[object]], Optional[str]]:
    if source_workbook_path.suffix.lower() == ".xls":
        raise ValueError(
            "O ficheiro Lista_Material encontrado esta em formato XLS antigo.\n\n"
            "Volte a gerar o Excel em 'Lista Material_IMOS' para obter um ficheiro .xlsm atualizado."
        )

    workbook = None
    try:
        workbook = load_workbook(source_workbook_path, data_only=True, read_only=True)
        worksheet = workbook[LISTAGEM_CUT_RITE_SHEET]
        header_values = [
            worksheet.cell(row=CUTRITE_SOURCE_HEADER_ROW, column=col).value
            for col in range(1, CUTRITE_SOURCE_MAX_COL + 1)
        ]
        headers = build_cutrite_import_row(header_values)
        if not any(str(value).strip() for value in headers):
            headers = build_cutrite_import_headers()

        rows: list[list[object]] = []
        for values in worksheet.iter_rows(
            min_row=CUTRITE_SOURCE_DATA_START_ROW,
            max_col=CUTRITE_SOURCE_MAX_COL,
            values_only=True,
        ):
            raw_row = list(values[:CUTRITE_SOURCE_MAX_COL])
            if len(raw_row) < CUTRITE_SOURCE_MAX_COL:
                raw_row.extend([None] * (CUTRITE_SOURCE_MAX_COL - len(raw_row)))
            if not any(value not in (None, "") for value in raw_row):
                continue
            raw_row[13] = CUTRITE_GRAFICO_ORLAS_VALUE
            rows.append(build_cutrite_import_row(raw_row))
        return headers, rows, None
    except Exception as exc:
        # Se o ficheiro estiver aberto no Excel (ou bloqueado), tentamos usar a macro embutida.
        if source_workbook_path.suffix.lower() == ".xlsm":
            macro_text, rows = _load_cutrite_source_table_from_excel_macro(source_workbook_path)
            if rows:
                headers = build_cutrite_import_headers()
                return headers, rows, macro_text
        raise
    finally:
        if workbook is not None:
            workbook.close()


def _load_cutrite_source_table_from_excel_macro(source_workbook_path: Path) -> tuple[str, list[list[object]]]:
    if source_workbook_path.suffix.lower() != ".xlsm":
        return "", []

    try:
        win32_client = importlib.import_module("win32com.client")
        win32clipboard = importlib.import_module("win32clipboard")
    except Exception:
        return "", []

    excel = None
    workbook = None
    try:
        excel = win32_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 1
        except Exception:
            pass

        workbook = excel.Workbooks.Open(str(source_workbook_path))
        _run_excel_macro(excel, source_workbook_path.name, "GERAR_ResumoOrlas")
        _run_excel_macro(excel, source_workbook_path.name, "Copia_Listagem_Software_Cut_Rite")
        clipboard_text = _read_cutrite_clipboard_text(win32clipboard)
        return clipboard_text, _parse_cutrite_clipboard_rows(clipboard_text)
    except Exception:
        return "", []
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _run_excel_macro(excel, workbook_name: str, macro_name: str) -> None:
    candidates = (
        macro_name,
        f"{workbook_name}!{macro_name}",
        f"'{workbook_name}'!{macro_name}",
    )
    last_error: Optional[Exception] = None
    for candidate in candidates:
        try:
            excel.Run(candidate)
            return
        except Exception as exc:
            last_error = exc
    if last_error is not None:
        raise last_error


def _read_cutrite_clipboard_text(win32clipboard_module, *, retries: int = 10, delay_seconds: float = 0.3) -> str:
    last_error: Optional[Exception] = None
    for attempt in range(retries):
        try:
            win32clipboard_module.OpenClipboard()
            try:
                return win32clipboard_module.GetClipboardData(win32clipboard_module.CF_UNICODETEXT)
            finally:
                win32clipboard_module.CloseClipboard()
        except Exception as exc:
            last_error = exc
            if attempt == retries - 1:
                break
            time.sleep(delay_seconds)
    if last_error is not None:
        raise last_error
    return ""


def _parse_cutrite_clipboard_rows(clipboard_text: str) -> list[list[object]]:
    rows: list[list[object]] = []
    for raw_line in str(clipboard_text or "").splitlines():
        if not raw_line.strip():
            continue
        parts = raw_line.split("\t")
        row = list(parts[:CUTRITE_EXPORT_MAX_COL])
        if len(row) < CUTRITE_EXPORT_MAX_COL:
            row.extend([""] * (CUTRITE_EXPORT_MAX_COL - len(row)))
        rows.append([_coerce_cutrite_clipboard_cell(value) for value in row])
    return rows


def _coerce_cutrite_clipboard_cell(value: object) -> object:
    text = _normalize_cutrite_cell(value)
    if not isinstance(text, str) or not text:
        return text
    if re.fullmatch(r"[+-]?\d+", text):
        try:
            return int(text)
        except Exception:
            return text
    if re.fullmatch(r"[+-]?\d+\.\d+", text):
        try:
            return float(text)
        except Exception:
            return text
    return text


def _write_cutrite_import_workbook_with_win32com(
    output_path: Path,
    table_rows: Sequence[Sequence[object]],
    *,
    timeout_seconds: int = 240,
) -> None:
    try:
        win32_client = importlib.import_module("win32com.client")
    except Exception as exc:
        raise RuntimeError(
            "A criacao do ficheiro CUT-RITE requer o pacote 'pywin32' instalado.\n\n"
            "Instale no mesmo Python/venv e volte a testar."
        ) from exc

    excel = None
    workbook = None
    worksheet = None
    try:
        excel = win32_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 3
        except Exception:
            pass

        workbook = excel.Workbooks.Add()
        worksheet = workbook.Worksheets.Item(1)
        worksheet.Name = "CUTRITE_IMPORT"

        for row_index, row_values in enumerate(table_rows, start=1):
            for col_index, value in enumerate(row_values, start=1):
                worksheet.Cells(row_index, col_index).Value = value

        try:
            worksheet.Columns.AutoFit()
        except Exception:
            pass

        if output_path.exists():
            output_path.unlink()

        workbook.SaveAs(str(output_path), 56)
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _is_process_running(process_name: str) -> bool:
    try:
        result = subprocess.run(
            ["tasklist", "/FI", f"IMAGENAME eq {process_name}"],
            capture_output=True,
            text=True,
            check=False,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0,
        )
    except Exception:
        return False
    return process_name.lower() in (result.stdout or "").lower()


def _describe_process_windows(process_id: int) -> str:
    if os.name != "nt":
        return ""
    try:
        import win32gui
        import win32process
    except Exception:
        return ""

    windows: list[str] = []

    def _enum_window(hwnd, _param):
        if not win32gui.IsWindowVisible(hwnd):
            return
        _thread_id, pid = win32process.GetWindowThreadProcessId(hwnd)
        if pid != process_id:
            return
        title = win32gui.GetWindowText(hwnd).strip()
        parts: list[str] = []

        def _enum_child(child_hwnd, _child_param):
            text = win32gui.GetWindowText(child_hwnd).strip()
            if text:
                parts.append(text)

        win32gui.EnumChildWindows(hwnd, _enum_child, None)
        detail = " | ".join([p for p in parts if p])
        if title and detail:
            windows.append(f"Janela: {title}\n{detail}")
        elif title:
            windows.append(f"Janela: {title}")

    try:
        win32gui.EnumWindows(_enum_window, None)
    except Exception:
        return ""
    return "\n\n".join(windows)

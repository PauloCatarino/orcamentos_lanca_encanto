from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from difflib import SequenceMatcher
from typing import Iterable, List, Optional, Sequence

from openpyxl import load_workbook
from sqlalchemy import cast, select, or_, String
from sqlalchemy.orm import Session

from ..models.materia_prima import MateriaPrima, MateriaPrimaPreference
from ..services.settings import get_setting, set_setting

# Config keys / defaults
KEY_MATERIAS_BASE_PATH = "base_path_materias_primas"
DEFAULT_MATERIAS_BASE_PATH = r"C:\\Users\\Utilizador\\Documents\\ORCAMENTOS_LE\\ORCAMENTOS_LE\\orcamentos_lanca_encanto\\Martelo_Orcamentos_V2\\Tabelas"
MATERIA_PRIMA_FILENAME = "TAB_MATERIAS_PRIMAS.xlsm"

# Column metadata
@dataclass(frozen=True)
class MateriaPrimaColumn:
    header: str
    attr: str
    kind: str = "text"  # "text" | "numeric"


COLUMN_DEFS: Sequence[MateriaPrimaColumn] = (
    MateriaPrimaColumn("ID_MP", "id_mp"),
    MateriaPrimaColumn("REF_PHC", "ref_phc"),
    MateriaPrimaColumn("REF_FORNECEDOR", "ref_fornecedor"),
    MateriaPrimaColumn("Ref_LE", "ref_le"),
    MateriaPrimaColumn("DESCRICAO_do_PHC", "descricao_phc"),
    MateriaPrimaColumn("DESCRICAO_no_ORCAMENTO", "descricao_orcamento"),
    MateriaPrimaColumn("PRECO_TABELA", "preco_tabela", "numeric"),
    MateriaPrimaColumn("MARGEM", "margem", "numeric"),
    MateriaPrimaColumn("DESCONTO", "desconto", "numeric"),
    MateriaPrimaColumn("PLIQ", "pliq", "numeric"),
    MateriaPrimaColumn("UND", "und"),
    MateriaPrimaColumn("DESP", "desp", "numeric"),
    MateriaPrimaColumn("COMP_MP", "comp_mp", "numeric"),
    MateriaPrimaColumn("LARG_MP", "larg_mp", "numeric"),
    MateriaPrimaColumn("ESP_MP", "esp_mp", "numeric"),
    MateriaPrimaColumn("TIPO", "tipo"),
    MateriaPrimaColumn("FAMILIA", "familia"),
    MateriaPrimaColumn("COR", "cor"),
    MateriaPrimaColumn("ORL_0.4", "orl_0_4", "text"),
    MateriaPrimaColumn("ORL_1.0", "orl_1_0", "text"),
    MateriaPrimaColumn("COR_REF_MATERIAL", "cor_ref_material"),
    MateriaPrimaColumn("NOME_FORNECEDOR", "nome_fornecedor"),
    MateriaPrimaColumn("NOME_FABRICANTE", "nome_fabricante"),
    MateriaPrimaColumn("DATA_ULTIMO_PRECO", "data_ultimo_preco"),
    MateriaPrimaColumn("APLICACAO", "aplicacao"),
    MateriaPrimaColumn("STOCK", "stock", "numeric"),
    MateriaPrimaColumn("NOTAS_2", "notas_2"),
    MateriaPrimaColumn("NOTAS_3", "notas_3"),
    MateriaPrimaColumn("NOTAS_4", "notas_4"),
)

DEFAULT_VISIBLE_COLUMNS: Sequence[str] = (
    "ID_MP",
    "DESCRICAO_no_ORCAMENTO",
    "PRECO_TABELA",
    "PLIQ",
    "UND",
    "TIPO",
    "FAMILIA",
    "NOME_FORNECEDOR",
)

FUZZY_LIMIT = 75
FUZZY_THRESHOLD = 0.65


def get_base_path(db: Session) -> str:
    return get_setting(db, KEY_MATERIAS_BASE_PATH, DEFAULT_MATERIAS_BASE_PATH)


def set_base_path(db: Session, path_value: str) -> None:
    set_setting(db, KEY_MATERIAS_BASE_PATH, path_value)
    db.flush()


def get_all_columns() -> Sequence[MateriaPrimaColumn]:
    return COLUMN_DEFS


def get_user_columns(db: Session, user_id: Optional[int]) -> List[str]:
    if not user_id:
        return list(DEFAULT_VISIBLE_COLUMNS)
    pref = db.execute(
        select(MateriaPrimaPreference).where(MateriaPrimaPreference.user_id == user_id)
    ).scalar_one_or_none()
    if pref:
        try:
            data = json.loads(pref.columns)
            if isinstance(data, list) and data:
                return [col for col in data if any(c.header == col for c in COLUMN_DEFS)]
        except json.JSONDecodeError:
            pass
    return list(DEFAULT_VISIBLE_COLUMNS)


def set_user_columns(db: Session, user_id: int, columns: Sequence[str]) -> None:
    valid = [col for col in columns if any(c.header == col for c in COLUMN_DEFS)]
    if not valid:
        valid = list(DEFAULT_VISIBLE_COLUMNS)
    pref = db.execute(
        select(MateriaPrimaPreference).where(MateriaPrimaPreference.user_id == user_id)
    ).scalar_one_or_none()
    payload = json.dumps(valid)
    if pref:
        pref.columns = payload
    else:
        pref = MateriaPrimaPreference(user_id=user_id, columns=payload)
        db.add(pref)
    db.flush()


def list_materias_primas(
    db: Session,
    query: str | None = None,
    *,
    tipo: Optional[str] = None,
    familia: Optional[str] = None,
) -> List[MateriaPrima]:
    query = (query or "").strip()
    stmt = select(MateriaPrima)
    if tipo:
        stmt = stmt.where(MateriaPrima.tipo == tipo)
    if familia:
        stmt = stmt.where(MateriaPrima.familia == familia)
    if query:
        terms = [term.strip() for term in query.split('%') if term.strip()]
        if terms:
            stmt = stmt.where(_build_search_filter(terms))
    stmt = stmt.order_by(MateriaPrima.descricao_orcamento)
    rows = db.execute(stmt).scalars().all()
    if query and not rows:
        rows = _fuzzy_search(db, query, tipo=tipo, familia=familia)
    return rows


def _build_search_filter(terms: Sequence[str]):
    filters = []
    for term in terms:
        like = f"%{term}%"
        term_filters = []
        for col in COLUMN_DEFS:
            column_attr = getattr(MateriaPrima, col.attr)
            if col.kind == "numeric":
                term_filters.append(cast(column_attr, String).ilike(like))
            else:
                term_filters.append(column_attr.ilike(like))
        filters.append(or_(*term_filters))
    combined = filters[0]
    for extra in filters[1:]:
        combined = combined & extra
    return combined


def _fuzzy_search(db: Session, raw_query: str, *, tipo: Optional[str] = None, familia: Optional[str] = None) -> List[MateriaPrima]:
    stmt = select(MateriaPrima)
    if tipo:
        stmt = stmt.where(MateriaPrima.tipo == tipo)
    if familia:
        stmt = stmt.where(MateriaPrima.familia == familia)
    rows = db.execute(stmt).scalars().all()
    if not rows:
        return []
    terms = [t.strip().lower() for t in raw_query.split('%') if t.strip()]
    scored: list[tuple[float, MateriaPrima]] = []
    for row in rows:
        best_score = 0.0
        for col in COLUMN_DEFS:
            value = getattr(row, col.attr)
            if value in (None, ""):
                continue
            text = str(value).lower()
            if terms and not all(term in text for term in terms):
                score = max(_similarity(text, term) for term in terms)
            else:
                score = _similarity(text, raw_query.lower())
            if score > best_score:
                best_score = score
        if best_score >= FUZZY_THRESHOLD:
            scored.append((best_score, row))
    scored.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in scored[:FUZZY_LIMIT]]


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def import_materias_primas(db: Session, base_path: Optional[str] = None) -> int:
    base_dir = Path(base_path or get_base_path(db))
    workbook_path = base_dir / MATERIA_PRIMA_FILENAME
    if not workbook_path.exists():
        raise FileNotFoundError(f"Ficheiro não encontrado: {workbook_path}")

    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[5][: len(COLUMN_DEFS)]]
    expected = [col.header for col in COLUMN_DEFS]
    if header != expected:
        raise ValueError(
            "Cabeçalho do Excel não corresponde ao esperado.\n"
            f"Esperado: {expected}\nObtido: {header}"
        )

    db.query(MateriaPrima).delete()

    inserted = 0
    for row in ws.iter_rows(min_row=6, max_col=len(COLUMN_DEFS)):
        values = [cell.value for cell in row]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        data = {}
        for meta, value in zip(COLUMN_DEFS, values):
            key = meta.attr
            if value is None:
                data[key] = None
                continue
            if meta.kind == "numeric":
                try:
                    data[key] = Decimal(str(value).replace(',', '.'))
                except Exception:
                    data[key] = None
            elif key == "data_ultimo_preco" and isinstance(value, (datetime, )):
                data[key] = value.date().isoformat()
            else:
                data[key] = str(value).strip()
        if not data.get("id_mp"):
            # ignora linhas sem identificador
            continue
        db.add(MateriaPrima(**data))
        inserted += 1
    wb.close()
    db.flush()
    return inserted




def get_materia_prima_by_ref_le(db: Session, ref_le: str | None):
    if not (ref_le or "").strip():
        return None
    return db.execute(
        select(MateriaPrima).where(MateriaPrima.ref_le == ref_le.strip())
    ).scalar_one_or_none()


def get_materia_prima_by_id(db: Session, id_mp: str | None):
    if not (id_mp or "").strip():
        return None
    return db.get(MateriaPrima, id_mp.strip())


def listar_tipos(db: Session) -> List[str]:
    stmt = select(MateriaPrima.tipo).where(MateriaPrima.tipo.isnot(None)).distinct().order_by(MateriaPrima.tipo)
    return [row[0] for row in db.execute(stmt).all() if row[0]]


def listar_familias(db: Session) -> List[str]:
    stmt = select(MateriaPrima.familia).where(MateriaPrima.familia.isnot(None)).distinct().order_by(MateriaPrima.familia)
    return [row[0] for row in db.execute(stmt).all() if row[0]]


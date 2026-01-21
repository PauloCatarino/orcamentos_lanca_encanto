from __future__ import annotations

import html
import warnings
warnings.filterwarnings("ignore", message=".*Failed to disconnect.*dataChanged.*")
import logging
import re
import pandas as pd
import shutil
import tempfile
import math
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Tuple, Any

from PySide6 import QtWidgets, QtGui, QtCore
from PySide6.QtCore import Qt
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from Martelo_Orcamentos_V2.utils_email import get_email_log_path, send_email
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
try:
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    MATPLOTLIB_AVAILABLE = True
except Exception:
    PdfPages = FigureCanvas = Figure = None
    MATPLOTLIB_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True

    class NumberedFooterCanvas(rl_canvas.Canvas):
        """
        Canvas que guarda os estados das páginas e desenha o rodapé com
        a data, o número do orçamento e a numeração X/Y.
        Implementação testada e estável (padrão recomendado).
        """
        def __init__(self, *args, footer_info: Optional[dict] = None, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []
            self.footer_info = footer_info or {}

        def showPage(self):
            # 1) Guarda o estado da página atual (conteúdo já desenhado)
            self._saved_page_states.append(dict(self.__dict__))

            # 2) NÃO grava a página agora; apenas "começa" uma nova página em memória
            #    _startPage() prepara uma nova página, mas não escreve nada no PDF.
            self._startPage()

        def save(self):
            # Agora sim, no final, percorremos todos os estados gravados
            num_pages = len(self._saved_page_states)

            for state in self._saved_page_states:
                # Recupera o estado da página
                self.__dict__.update(state)

                # Desenha o rodapé para esta página
                self._draw_footer(num_pages)

                # Grava a página no PDF (agora com rodapé)
                rl_canvas.Canvas.showPage(self)

            # Por fim, chama o save original para fechar o ficheiro
            rl_canvas.Canvas.save(self)

        def _draw_footer(self, page_count: int):
            # posição vertical do rodapé (ajusta se quiseres mais acima/baixo)
            width, _ = self._pagesize
            y = 6 * mm
            self.setFont("Helvetica", 9)

            # esquerda: apenas a data (sem "Data:")
            self.drawString(18 * mm, y, f"{self.footer_info.get('data', '')}")

            # centro: nº orçamento + versão
            self.drawCentredString(width / 2, y, self.footer_info.get("numero", ""))

            # direita: nº página / total
            self.drawRightString(width - 18 * mm, y, f"{self._pageNumber}/{page_count}")


except Exception:
    colors = A4 = ParagraphStyle = getSampleStyleSheet = mm = Image = Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = NumberedFooterCanvas = None
    REPORTLAB_AVAILABLE = False
from sqlalchemy import select, text
from sqlalchemy.exc import OperationalError

from Martelo_Orcamentos_V2.app.db import SessionLocal
from Martelo_Orcamentos_V2.app.models import Client, Orcamento, OrcamentoItem, CusteioItem, CusteioDespBackup
from Martelo_Orcamentos_V2.app.services.custeio_items import atualizar_orlas_custeio
from Martelo_Orcamentos_V2.app.services.orcamentos import (
    resolve_orcamento_cliente_nome,
    resolve_orcamento_temp_cliente,
)
from Martelo_Orcamentos_V2.app.services.settings import get_setting
from Martelo_Orcamentos_V2.ui.models.qt_table import SimpleTableModel


IVA_RATE = Decimal("0.23")
KEY_BASE_PATH = "base_path_orcamentos"
KEY_ORC_DB_BASE = "base_path_dados_orcamento"
DEFAULT_BASE_PATH = r"\\server_le\_Lanca_Encanto\LancaEncanto\Dep._Orcamentos\MARTELO_ORCAMENTOS_V2"
DEFAULT_BASE_DADOS_ORC = r"\\SERVER_LE\_Lanca_Encanto\LancaEncanto\Dep._Orcamentos\Base_Dados_Orcamento"

logger = logging.getLogger(__name__)


class RichTextDelegate(QtWidgets.QStyledItemDelegate):
    def paint(self, painter: QtGui.QPainter, option: QtWidgets.QStyleOptionViewItem, index: QtCore.QModelIndex) -> None:
        text = index.data(Qt.DisplayRole)
        if not isinstance(text, str):
            super().paint(painter, option, index)
            return
        painter.save()
        doc = QtGui.QTextDocument()
        doc.setDefaultFont(option.font)
        doc.setHtml(text)
        ctx = QtGui.QAbstractTextDocumentLayout.PaintContext()
        if option.state & QtWidgets.QStyle.State_Selected:
            painter.fillRect(option.rect, option.palette.highlight())
            ctx.palette.setColor(QtGui.QPalette.Text, option.palette.highlightedText().color())
        painter.translate(option.rect.topLeft())
        doc.setTextWidth(option.rect.width())
        doc.documentLayout().draw(painter, ctx)
        painter.restore()

    def sizeHint(self, option: QtWidgets.QStyleOptionViewItem, index: QtCore.QModelIndex) -> QtCore.QSize:
        text = index.data(Qt.DisplayRole)
        if not isinstance(text, str):
            return super().sizeHint(option, index)
        doc = QtGui.QTextDocument()
        doc.setDefaultFont(option.font)
        doc.setHtml(text)
        width = option.rect.width() if option.rect.width() > 0 else 400
        doc.setTextWidth(width)
        size = doc.size().toSize()
        return QtCore.QSize(int(doc.idealWidth()), size.height())


class BoldValueDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option: QtWidgets.QStyleOptionViewItem, index: QtCore.QModelIndex) -> None:
        super().initStyleOption(option, index)
        font = option.font
        font.setBold(True)
        option.font = font


class _WheelForwarder(QtCore.QObject):
    """Reencaminha eventos de roda do rato para um widget pai (ex.: QScrollArea)."""

    def __init__(self, target: QtWidgets.QWidget):
        super().__init__(target)
        self._target = target

    def eventFilter(self, obj: QtCore.QObject, event: QtCore.QEvent) -> bool:
        if event.type() == QtCore.QEvent.Wheel and self._target is not None:
            # reenvia para o viewport do scroll e deixa o evento propagar
            target_widget = getattr(self._target, "viewport", lambda: self._target)()
            QtWidgets.QApplication.sendEvent(target_widget, event)
            return False
        return super().eventFilter(obj, event)


@dataclass
class ItemPreview:
    item: str
    codigo: str
    descricao: str
    altura: Optional[Decimal]
    largura: Optional[Decimal]
    profundidade: Optional[Decimal]
    unidade: str
    qt: Optional[Decimal]
    preco_unitario: Optional[Decimal]
    preco_total: Optional[Decimal]


def _is_separator_item(item: ItemPreview) -> bool:
    """
    Itens "vazios" são usados como separadores visuais (ex.: linha em branco).
    Não devem contar para totais de quantidade (Qt).

    Regra: se Código e Descrição estiverem vazios (ou só espaços), é separador.
    """
    codigo = (getattr(item, "codigo", "") or "").strip()
    descricao = (getattr(item, "descricao", "") or "").strip()
    return not codigo and not descricao


def _fmt_decimal(value: Optional[Decimal], decimals: int = 2) -> str:
    if value in (None, ""):
        return ""
    try:
        number = Decimal(value)
    except Exception:
        return str(value)
    fmt = f"{{0:.{decimals}f}}"
    return fmt.format(number)


def _fmt_currency(value: Optional[Decimal]) -> str:
    if value in (None, ""):
        return ""
    try:
        number = Decimal(value)
    except Exception:
        return str(value)
    return f"{number:.2f} €"


class RelatoriosPage(QtWidgets.QWidget):
    """Página dedicada aos relatórios de orçamento."""

    def __init__(self, parent=None, *, current_user=None):
        super().__init__(parent)
        self.current_user = current_user
        self.current_orcamento_id: Optional[int] = None
        self._current_orcamento: Optional[Orcamento] = None
        self._current_client: Optional[Client] = None
        self._current_temp_client = None
        self._current_cliente_nome: str = ""
        self._current_items: List[ItemPreview] = []
        self._dashboard_data = {"placas": [], "orlas": [], "ferr": [], "maq": []}
        self._dash_info_labels: dict[str, QtWidgets.QLabel] = {}
        self._nao_stock_dirty: bool = False
        self._setup_ui()

    def _show_toast(self, widget: Optional[QtWidgets.QWidget], text: str, timeout_ms: int = 3000) -> None:
        if not text:
            return
        try:
            if widget is not None:
                rect = widget.rect()
                center = rect.center()
                global_pos = widget.mapToGlobal(center)
                QtWidgets.QToolTip.showText(global_pos, text, widget, rect, timeout_ms)
                return
        except Exception:
            pass
        QtWidgets.QToolTip.showText(QtGui.QCursor.pos(), text)

    def _set_nao_stock_dirty(self, dirty: bool) -> None:
        """Adiciona '*' ao botão de gravação quando há alterações por gravar."""
        self._nao_stock_dirty = bool(dirty)
        if getattr(self, "btn_nao_stock_save", None):
            prefix = "* " if self._nao_stock_dirty else ""
            self.btn_nao_stock_save.setText(f"{prefix}Gravar Nao Stock")

    # ------------------------------------------------------------------ UI SETUP
    def _setup_ui(self) -> None:
        layout = QtWidgets.QVBoxLayout(self)

        self.tab_widget = QtWidgets.QTabWidget()
        layout.addWidget(self.tab_widget, 1)

        self.tab_orcamento = QtWidgets.QWidget()
        self.tab_widget.addTab(self.tab_orcamento, "Relatório de Orçamento")

        tab_layout = QtWidgets.QVBoxLayout(self.tab_orcamento)
        tab_layout.setContentsMargins(4, 4, 4, 4)
        tab_layout.setSpacing(8)

        top_row = QtWidgets.QHBoxLayout()
        tab_layout.addLayout(top_row)

        self.client_box = QtWidgets.QGroupBox("Dados do Cliente")
        client_form = QtWidgets.QFormLayout(self.client_box)
        client_form.setSpacing(4)
        self.lbl_cliente_nome = QtWidgets.QLabel("-")
        self.lbl_cliente_morada = QtWidgets.QLabel("-")
        self.lbl_cliente_morada.setWordWrap(True)
        self.lbl_cliente_email = QtWidgets.QLabel("-")
        self.lbl_cliente_phc = QtWidgets.QLabel("-")
        self.lbl_cliente_tel = QtWidgets.QLabel("-")
        self.lbl_cliente_telemovel = QtWidgets.QLabel("-")
        for lbl in (
            self.lbl_cliente_nome,
            self.lbl_cliente_morada,
            self.lbl_cliente_email,
            self.lbl_cliente_phc,
            self.lbl_cliente_tel,
            self.lbl_cliente_telemovel,
        ):
            lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)

        client_form.addRow("Nome:", self.lbl_cliente_nome)
        client_form.addRow("Morada:", self.lbl_cliente_morada)
        client_form.addRow("Email:", self.lbl_cliente_email)
        client_form.addRow("N.º Cliente PHC:", self.lbl_cliente_phc)
        client_form.addRow("Telefone:", self.lbl_cliente_tel)
        client_form.addRow("Telemóvel:", self.lbl_cliente_telemovel)
        top_row.addWidget(self.client_box, 3)

        self.orc_box = QtWidgets.QGroupBox("Identificação do Orçamento")
        orc_form = QtWidgets.QFormLayout(self.orc_box)
        orc_form.setSpacing(4)
        self.lbl_orc_data = QtWidgets.QLabel("-")
        self.lbl_orc_num = QtWidgets.QLabel("-")
        self.lbl_orc_versao = QtWidgets.QLabel("-")
        self.lbl_orc_obra = QtWidgets.QLabel("-")
        self.lbl_orc_ref = QtWidgets.QLabel("-")
        self.lbl_orc_obra.setWordWrap(True)
        for lbl in (self.lbl_orc_data, self.lbl_orc_num, self.lbl_orc_versao, self.lbl_orc_obra, self.lbl_orc_ref):
            lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        orc_form.addRow("Data:", self.lbl_orc_data)
        orc_form.addRow("N.º Orçamento:", self.lbl_orc_num)
        orc_form.addRow("Versão:", self.lbl_orc_versao)
        orc_form.addRow("Obra:", self.lbl_orc_obra)
        orc_form.addRow("Ref. Cliente:", self.lbl_orc_ref)
        top_row.addWidget(self.orc_box, 2)

        actions_box = QtWidgets.QGroupBox("Ações")
        actions_layout = QtWidgets.QVBoxLayout(actions_box)

        def _icon(theme_name: str, fallback: QtWidgets.QStyle.StandardPixmap) -> QtGui.QIcon:
            icon = QtGui.QIcon.fromTheme(theme_name)
            if icon.isNull():
                icon = self.style().standardIcon(fallback)
            return icon

        def _letter_icon(text: str, color: QtGui.QColor) -> QtGui.QIcon:
            size = 22
            pm = QtGui.QPixmap(size, size)
            pm.fill(QtCore.Qt.transparent)
            painter = QtGui.QPainter(pm)
            painter.setRenderHint(QtGui.QPainter.Antialiasing)
            pen = QtGui.QPen(color)
            pen.setWidth(2)
            painter.setPen(pen)
            painter.setBrush(QtGui.QBrush(color.lighter(180)))
            painter.drawRoundedRect(1, 1, size - 2, size - 2, 4, 4)
            font = QtGui.QFont(self.font())
            font.setBold(True)
            painter.setFont(font)
            painter.setPen(QtCore.Qt.white)
            painter.drawText(pm.rect(), QtCore.Qt.AlignCenter, text)
            painter.end()
            return QtGui.QIcon(pm)

        def _envelope_icon(color: QtGui.QColor) -> QtGui.QIcon:
            size = 22
            pm = QtGui.QPixmap(size, size)
            pm.fill(QtCore.Qt.transparent)
            painter = QtGui.QPainter(pm)
            painter.setRenderHint(QtGui.QPainter.Antialiasing)
            pen = QtGui.QPen(color)
            pen.setWidth(2)
            painter.setPen(pen)
            painter.setBrush(QtGui.QBrush(color.lighter(190)))
            rect = QtCore.QRect(2, 6, size - 4, size - 10)
            painter.drawRoundedRect(rect, 3, 3)
            mid_x = rect.center().x()
            mid_y = rect.center().y()
            painter.drawLine(rect.left() + 1, rect.top() + 1, mid_x, mid_y)
            painter.drawLine(rect.right() - 1, rect.top() + 1, mid_x, mid_y)
            painter.drawLine(rect.left() + 1, rect.bottom() - 1, mid_x, mid_y)
            painter.drawLine(rect.right() - 1, rect.bottom() - 1, mid_x, mid_y)
            painter.end()
            return QtGui.QIcon(pm)

        def _bold_button(text: str) -> QtWidgets.QPushButton:
            btn = QtWidgets.QPushButton(text)
            f = btn.font()
            f.setBold(True)
            btn.setFont(f)
            return btn

        self.btn_send_email = _bold_button("Enviar Orçamento Email")
        # Ícone de envelope
        self.btn_send_email.setIcon(_envelope_icon(QtGui.QColor("#2d7dd2")))
        self.btn_send_email.setToolTip(
            "Enviar orçamento por email.\n"
            "- Se existir PDF gerado pelo Martelo, é anexado automaticamente.\n"
            "- Se não existir, pode indicar o destinatário e adicionar anexos manualmente."
        )

        self.btn_export_excel = _bold_button("Exportar para Excel")
        self.btn_export_excel.setIcon(_letter_icon("X", QtGui.QColor("#1d6f42")))
        self.btn_export_excel.setToolTip("Exporta o relatório do orçamento para Excel.")

        self.btn_export_pdf = _bold_button("Exportar para PDF")
        self.btn_export_pdf.setIcon(_letter_icon("P", QtGui.QColor("#c62828")))
        self.btn_export_pdf.setToolTip("Exporta o relatório do orçamento para PDF.")

        self.btn_open_folder = _bold_button("Abrir Pasta Orçamento")
        self.btn_open_folder.setIcon(_icon("folder-open", QtWidgets.QStyle.SP_DirOpenIcon))
        self.btn_open_folder.setToolTip("Abre no Explorador a pasta do orçamento (se existir).")

        # Botões que estavam no separador "Resumo de Consumos" (agora agrupados em "Ações")
        self.btn_dash_export = _bold_button("Exportar Dashboard para PDF")
        self.btn_dash_export.setIcon(_letter_icon("D", QtGui.QColor("#6a1b9a")))
        self.btn_dash_export.setToolTip("Exporta o dashboard de consumos para PDF.")

        self.btn_cut_plan = _bold_button("Exportar Plano Corte PDF")
        self.btn_cut_plan.setIcon(_letter_icon("C", QtGui.QColor("#f57c00")))
        self.btn_cut_plan.setToolTip("Gera plano de corte de placas em PDF a partir do Resumo_Custos.")

        actions_layout.addWidget(self.btn_export_excel)
        actions_layout.addWidget(self.btn_export_pdf)
        actions_layout.addWidget(self.btn_open_folder)
        actions_layout.addWidget(self.btn_dash_export)
        actions_layout.addWidget(self.btn_cut_plan)
        actions_layout.addWidget(self.btn_send_email)
        actions_layout.addStretch(1)
        top_row.addWidget(actions_box, 1)

        self.btn_send_email.clicked.connect(self._on_send_email)
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        self.btn_export_pdf.clicked.connect(self.export_to_pdf)
        self.btn_open_folder.clicked.connect(self._open_orcamento_folder)
        self.btn_dash_export.clicked.connect(self._export_dashboard_pdf)
        self.btn_cut_plan.clicked.connect(self.export_cut_plan_pdf)

        # tabela de itens
        self.table = QtWidgets.QTableView()
        self.table_model = SimpleTableModel(
            columns=[
                ("Item", "item"),
                ("Código", "codigo"),
                ("Descrição", "descricao_html", lambda v: v or ""),
                ("Altura", "altura", lambda v: _fmt_decimal(v, 1)),
                ("Largura", "largura", lambda v: _fmt_decimal(v, 1)),
                ("Profundidade", "profundidade", lambda v: _fmt_decimal(v, 1)),
                ("Unidade", "unidade"),
                ("Qt", "qt", lambda v: _fmt_decimal(v, 2)),
                ("Preço Unitário", "preco_unitario", _fmt_currency),
                ("Preço Total", "preco_total", _fmt_currency),
            ]
        )
        self.table.setModel(self.table_model)
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.table.verticalHeader().setVisible(False)
        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Interactive)
        header.resizeSection(1, 120)  # Código
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Interactive)  # Descrição (largura fixa)
        header.resizeSection(2, 530)

        header.setSectionResizeMode(3, QtWidgets.QHeaderView.Interactive)
        header.resizeSection(3, 80)  # Altura
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.Interactive)
        header.resizeSection(4, 80)  # Largura

        for idx in range(5, 9):
            header.setSectionResizeMode(idx, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(9, QtWidgets.QHeaderView.ResizeToContents)
        self.table.verticalHeader().setDefaultSectionSize(80)
        self.table.setStyleSheet(
            """
            QTableView {
                gridline-color: #b5b5b5;
                selection-background-color: #e0f0ff;
                font-size: 12px;
            }
            """
        )
        self.table.setItemDelegateForColumn(2, RichTextDelegate(self.table))
        self.table.setItemDelegateForColumn(9, BoldValueDelegate(self.table))
        header.sectionResized.connect(self._on_header_resized)
        tab_layout.addWidget(self.table, 1)

        totals_frame = QtWidgets.QFrame()
        totals_layout = QtWidgets.QGridLayout(totals_frame)
        totals_layout.setContentsMargins(0, 0, 0, 0)
        totals_layout.setHorizontalSpacing(16)
        totals_layout.setVerticalSpacing(2)
        self.lbl_total_qt = QtWidgets.QLabel("Total Qt: 0")
        self.lbl_subtotal = QtWidgets.QLabel("Subtotal: 0.00 €")
        self.lbl_iva = QtWidgets.QLabel("IVA (23%): 0.00 €")
        self.lbl_total_geral = QtWidgets.QLabel("Total Geral: 0.00 €")
        bold_font = self.lbl_total_geral.font()
        bold_font.setBold(True)
        self.lbl_total_geral.setFont(bold_font)
        totals_layout.addWidget(self.lbl_total_qt, 0, 0, alignment=Qt.AlignRight)
        totals_layout.addWidget(self.lbl_subtotal, 0, 1, alignment=Qt.AlignRight)
        totals_layout.addWidget(self.lbl_iva, 1, 0, alignment=Qt.AlignRight)
        totals_layout.addWidget(self.lbl_total_geral, 1, 1, alignment=Qt.AlignRight)
        tab_layout.addWidget(totals_frame, 0, alignment=Qt.AlignRight)

        # segundo separador: dashboard de consumos
        self.tab_dashboard = QtWidgets.QWidget()
        self._build_dashboard_tab(self.tab_dashboard)
        self.tab_widget.addTab(self.tab_dashboard, "Resumo de Consumos")

        self._update_actions_enabled(False)

    # ------------------------------------------------------------------ DATA FLOW
    def set_orcamento(self, orcamento_id: Optional[int]) -> None:
        self.current_orcamento_id = orcamento_id
        has_orc = orcamento_id is not None
        if not has_orc:
            self._clear_preview()
        self._update_actions_enabled(has_orc)

    def refresh_preview(self) -> None:
        if not self.current_orcamento_id:
            QtWidgets.QMessageBox.information(self, "Relatórios", "Selecione um orçamento para gerar o relatório.")
            return
        try:
            self._load_from_db(self.current_orcamento_id)
            self._apply_to_ui()
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Relatórios", f"Falha ao gerar pré-visualização: {exc}")

    def _load_from_db(self, orcamento_id: int) -> None:
        with SessionLocal() as session:
            orcamento = session.get(Orcamento, orcamento_id)
            if not orcamento:
                raise ValueError("Orçamento não encontrado.")
            client = session.get(Client, orcamento.client_id) if orcamento.client_id else None
            temp_client = resolve_orcamento_temp_cliente(session, orcamento)
            cliente_nome = resolve_orcamento_cliente_nome(session, orcamento, client=client)
            stmt = (
                select(OrcamentoItem)
                .where(OrcamentoItem.id_orcamento == orcamento_id)
                .order_by(OrcamentoItem.item_ord)
            )
            items = session.execute(stmt).scalars().all()

        parsed_rows: List[ItemPreview] = []
        for it in items:
            parsed_rows.append(
                ItemPreview(
                    item=str(it.item or ""),
                    codigo=str(it.codigo or ""),
                    descricao=str(it.descricao or ""),
                    altura=Decimal(it.altura) if it.altura is not None else None,
                    largura=Decimal(it.largura) if it.largura is not None else None,
                    profundidade=Decimal(it.profundidade) if it.profundidade is not None else None,
                    unidade=str(it.und or ""),
                    qt=Decimal(it.qt) if it.qt is not None else None,
                    preco_unitario=Decimal(it.preco_unitario) if it.preco_unitario is not None else None,
                    preco_total=Decimal(it.preco_total) if it.preco_total is not None else None,
                )
            )

        self._current_orcamento = orcamento
        self._current_client = client
        self._current_temp_client = temp_client
        self._current_cliente_nome = cliente_nome
        self._current_items = parsed_rows

    def _apply_to_ui(self) -> None:
        client = self._current_client
        cliente_source = self._current_temp_client or client
        orc = self._current_orcamento
        items = self._current_items
        if not orc:
            self._clear_preview()
            return

        def label_text(value: Optional[str]) -> str:
            text = str(value or "").strip()
            return text if text else "-"

        self.lbl_cliente_nome.setText(label_text(self._cliente_display_nome()))
        self.lbl_cliente_morada.setText(label_text(getattr(cliente_source, "morada", None)))
        self.lbl_cliente_email.setText(label_text(getattr(cliente_source, "email", None)))
        self.lbl_cliente_phc.setText(label_text(getattr(cliente_source, "num_cliente_phc", None)))
        self.lbl_cliente_tel.setText(label_text(getattr(cliente_source, "telefone", None)))
        self.lbl_cliente_telemovel.setText(label_text(getattr(cliente_source, "telemovel", None)))

        self.lbl_orc_data.setText(label_text(orc.data))
        self.lbl_orc_num.setText(label_text(orc.num_orcamento))
        self.lbl_orc_versao.setText(label_text(self._format_versao(orc.versao)))
        self.lbl_orc_obra.setText(label_text(orc.obra))
        self.lbl_orc_ref.setText(label_text(getattr(orc, "ref_cliente", None)))

        table_rows = [
            {
                "item": item.item,
                "codigo": item.codigo,
                "descricao_html": self._format_description_preview(item.descricao),
                "altura": item.altura,
                "largura": item.largura,
                "profundidade": item.profundidade,
                "unidade": item.unidade,
                "qt": item.qt,
                "preco_unitario": item.preco_unitario,
                "preco_total": item.preco_total,
            }
            for item in items
        ]
        self.table_model.set_rows(table_rows)
        QtCore.QTimer.singleShot(0, self._update_row_heights)

        total_qt = sum((row.qt or Decimal("0")) for row in items if not _is_separator_item(row))
        subtotal = sum((row.preco_total or Decimal("0")) for row in items)
        iva = subtotal * IVA_RATE
        total = subtotal + iva

        self.lbl_total_qt.setText(f"Total Qt: {_fmt_decimal(total_qt, 2) or '0'}")
        self.lbl_subtotal.setText(f"Subtotal: {_fmt_currency(subtotal)}")
        self.lbl_iva.setText(f"IVA (23%): {_fmt_currency(iva)}")
        self.lbl_total_geral.setText(f"Total Geral: {_fmt_currency(total)}")
        self._refresh_dashboard()

    def _parse_description(self, text: Optional[str]) -> List[Tuple[str, str]]:
        """
        Converte o texto da descrição em uma lista de (tipo, conteúdo),
        onde tipo pode ser:
        - 'header'  : primeira linha (título do módulo)
        - 'header2' : linhas seguintes totalmente em maiúsculas (sub-títulos)
        - 'dash'    : linhas que começam por '-' (bullets normais)
        - 'star'    : linhas que começam por '*' (bullets especiais)
        - 'plain'   : resto

        Trata também descrições que vêm NUMA ÚNICA LINHA e/ou que começam
        por '-' ou '*' (separando corretamente em bullets).
        """
        lines: List[Tuple[str, str]] = []
        if not text:
            return lines

        # Normalizar quebras de linha
        norm = text.replace("\r\n", "\n").replace("\r", "\n").strip()
        if not norm:
            return lines

        # -------------------------------------------------
        # CASO 1: já existem quebras de linha -> usa-as
        # -------------------------------------------------
        if "\n" in norm:
            raw_lines = norm.splitlines()
            for idx, raw in enumerate(raw_lines):
                stripped = raw.strip()
                if not stripped:
                    continue

                lstripped = raw.lstrip()

                # 1ª linha é sempre o cabeçalho principal
                if idx == 0:
                    lines.append(("header", stripped))
                    continue

                # Bullets
                if lstripped.startswith("-"):
                    lines.append(("dash", lstripped[1:].strip() or stripped))
                elif lstripped.startswith("*"):
                    lines.append(("star", lstripped[1:].strip() or stripped))
                else:
                    # Linha toda em maiúsculas -> "sub-header"
                    letters = [ch for ch in stripped if ch.isalpha()]
                    if letters and stripped.upper() == stripped:
                        lines.append(("header2", stripped))
                    else:
                        lines.append(("plain", stripped))
            return lines

        # -------------------------------------------------
        # CASO 2: tudo numa só linha -> tratar bullets mesmo assim
        #         (verifica se começa por '-' ou '*', ou tem " - " / " * ")
        # -------------------------------------------------
        # Se começa por '-' ou '*', capturar todos os bullets com regex
        first_nonspace = norm.lstrip()[0] if norm.lstrip() else ""
        if first_nonspace in ("-", "*"):
            # encontra todos os blocos "- texto" ou "* texto"
            for m in re.finditer(r'([-*])\s*([^-*]+?)(?=(?:\s[-*]\s)|$)', norm):
                bullet = m.group(1)
                content = m.group(2).strip()
                if not content:
                    continue
                kind = "star" if bullet == "*" else "dash"
                lines.append((kind, content))
            return lines

        # Caso tenha separadores " - " ou " * " mas não comece com bullet:
        parts = re.split(r'( - |\* )', norm)
        if len(parts) == 1:
            # sem separadores: tratar tudo como header
            lines.append(("header", norm))
            return lines

        # primeira parte antes de qualquer "-" ou "*" é o header
        header_text = parts[0].strip()
        if header_text:
            lines.append(("header", header_text))

        # restantes vêm aos pares: [delim, conteúdo, delim, conteúdo, ...]
        for i in range(1, len(parts), 2):
            delim = parts[i]
            if i + 1 >= len(parts):
                break
            content = parts[i + 1].strip()
            if not content:
                continue

            if "*" in delim:
                lines.append(("star", content))
            else:
                lines.append(("dash", content))

        return lines

    def _format_description_preview(self, text: Optional[str]) -> str:
        entries = self._parse_description(text)
        if not entries:
            return ""
        html_parts: List[str] = []
        for kind, content in entries:
            render_text = content.upper() if kind in ("header", "header2") else content
            # Evitar entidades tipo &#x27; / &quot; (ReportLab e alguns parsers não gostam quando ficam em uppercase)
            safe_text = html.escape(render_text, quote=False)
            if kind in ("header", "header2"):
                html_parts.append(f"<div style='font-weight:bold;text-transform:uppercase'>{safe_text}</div>")
            elif kind == "dash":
                html_parts.append(
                    f"<div style='font-style:italic;margin-left:10px'>- {safe_text}</div>"
                )
            elif kind == "star":
                html_parts.append(
                    "<div style='font-style:italic;text-decoration:underline;"
                    "background-color:#d6f6c8;color:#095c1b;padding:1px 2px;margin-left:10px;'>* "
                    f"{safe_text}</div>"
                )
            else:
                html_parts.append(f"<div>{safe_text}</div>")
        return "".join(html_parts)

    @staticmethod
    def _format_versao(versao: Optional[str]) -> str:
        """Formata a versão como duas casas (ex.: '01'); devolve '-' se vazia."""
        if versao in (None, "", "-"):
            return "-"
        try:
            return f"{int(str(versao)) :02d}"
        except Exception:
            return str(versao).strip()

    def _cliente_info_source(self):
        return self._current_temp_client or self._current_client

    def _cliente_display_nome(self) -> str:
        nome = (self._current_cliente_nome or "").strip()
        if nome:
            return nome
        source = self._cliente_info_source()
        return (getattr(source, "nome", "") or "").strip()

    def _clear_preview(self) -> None:
        for lbl in (
            self.lbl_cliente_nome,
            self.lbl_cliente_morada,
            self.lbl_cliente_email,
            self.lbl_cliente_phc,
            self.lbl_cliente_tel,
            self.lbl_cliente_telemovel,
            self.lbl_orc_data,
            self.lbl_orc_num,
            self.lbl_orc_versao,
            self.lbl_orc_obra,
            self.lbl_orc_ref,
        ):
            lbl.setText("-")
        self._current_temp_client = None
        self._current_cliente_nome = ""
        self.table_model.set_rows([])
        self._update_row_heights()
        self.lbl_total_qt.setText("Total Qt: 0")
        self.lbl_subtotal.setText("Subtotal: 0.00 €")
        self.lbl_iva.setText("IVA (23%): 0.00 €")
        self.lbl_total_geral.setText("Total Geral: 0.00 €")

    def _format_description_pdf(self, text: Optional[str]) -> str:
        """
        Formatação HTML usada no PDF do orçamento (títulos em bold/maiusculas e bullets em itálico).
        """
        entries = self._parse_description(text)
        if not entries:
            return ""
        parts: List[str] = []
        for kind, content in entries:
            if kind in ("header", "header2"):
                safe = html.escape(content.upper(), quote=False)
                parts.append(f"<b>{safe}</b>")
            elif kind == "dash":
                safe = html.escape(content, quote=False)
                parts.append(f"<i>- {safe}</i>")
            elif kind == "star":
                safe = html.escape(content, quote=False)
                parts.append(f"<i><u><font color='#0a5c0a'>* {safe}</font></u></i>")
            else:
                safe = html.escape(content, quote=False)
                parts.append(safe)
        return "<br/>".join(parts)

    def _update_actions_enabled(self, enabled: bool) -> None:
        self.btn_send_email.setEnabled(enabled)
        self.btn_export_excel.setEnabled(enabled)
        self.btn_export_pdf.setEnabled(enabled)
        if getattr(self, "btn_open_folder", None):
            self.btn_open_folder.setEnabled(enabled)
        if getattr(self, "btn_dash_export", None):
            self.btn_dash_export.setEnabled(enabled)
        if getattr(self, "btn_cut_plan", None):
            self.btn_cut_plan.setEnabled(enabled)
        if not enabled:
            self._set_nao_stock_dirty(False)

    # ---------------------- DASHBOARD ----------------------
    def _wrap_in_group(self, title: str, widget: QtWidgets.QWidget, wheel_forwarder: Optional[_WheelForwarder] = None) -> QtWidgets.QGroupBox:
        box = QtWidgets.QGroupBox(title)
        lay = QtWidgets.QVBoxLayout(box)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.addWidget(widget)
        if wheel_forwarder is not None:
            box.installEventFilter(wheel_forwarder)
            widget.installEventFilter(wheel_forwarder)
        return box

    def _create_canvas(self) -> Tuple[Any, Any]:
        fig = Figure(figsize=(5, 3))
        ax = fig.add_subplot(111)
        canvas = FigureCanvas(fig)
        canvas.setFocusPolicy(QtCore.Qt.NoFocus)
        return canvas, ax

    def _build_dashboard_tab(self, parent: QtWidgets.QWidget) -> None:
        if not MATPLOTLIB_AVAILABLE:
            layout = QtWidgets.QVBoxLayout(parent)
            layout.addStretch(1)
            layout.addWidget(
                QtWidgets.QLabel(
                    "Dashboard indisponível: instale 'matplotlib' para ver gráficos.",
                    alignment=Qt.AlignCenter,
                )
            )
            layout.addStretch(1)
            return
        outer = QtWidgets.QVBoxLayout(parent)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        content = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(content)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)
        self._wheel_forwarder_dashboard = _WheelForwarder(scroll)

        # barra de info orçamento
        info_frame = QtWidgets.QFrame()
        info_layout = QtWidgets.QHBoxLayout(info_frame)
        info_layout.setContentsMargins(4, 2, 4, 2)
        info_layout.setSpacing(12)
        def _mk_info(label: str) -> QtWidgets.QLabel:
            lbl = QtWidgets.QLabel("-")
            lbl.setStyleSheet("font-weight: bold;")
            title = QtWidgets.QLabel(label)
            title.setStyleSheet("color: #555;")
            container = QtWidgets.QHBoxLayout()
            w = QtWidgets.QWidget()
            container.setSpacing(4)
            container.setContentsMargins(0, 0, 0, 0)
            container.addWidget(title)
            container.addWidget(lbl)
            container.addStretch(1)
            w.setLayout(container)
            info_layout.addWidget(w)
            return lbl
        self._dash_info_labels = {
            "ano": _mk_info("Ano:"),
            "cliente": _mk_info("Cliente:"),
            "orc": _mk_info("Nº Orçamento:"),
            "versao": _mk_info("Versão:"),
            "user": _mk_info("Utilizador:"),
        }
        info_layout.addStretch(1)
        layout.addWidget(info_frame)

        actions = QtWidgets.QHBoxLayout()
        actions.setSpacing(10)

        def _dash_icon(theme: str, fallback: QtWidgets.QStyle.StandardPixmap) -> QtGui.QIcon:
            icon = QtGui.QIcon.fromTheme(theme)
            if icon.isNull():
                icon = self.style().standardIcon(fallback)
            return icon

        def _make_dash_button(text: str, icon: QtGui.QIcon) -> QtWidgets.QPushButton:
            btn = QtWidgets.QPushButton(text)
            font = btn.font()
            font.setBold(True)
            font.setPointSize(max(font.pointSize(), 11))
            btn.setFont(font)
            btn.setIcon(icon)
            btn.setIconSize(QtCore.QSize(22, 22))
            return btn

        self.btn_dash_refresh = _make_dash_button(
            "Atualizar Dashboard",
            _dash_icon("view-refresh", QtWidgets.QStyle.SP_BrowserReload),
        )
        self.btn_dash_refresh.setToolTip(
            "Recalcula e atualiza o Resumo de Consumos do orçamento/versão selecionados.\n"
            "- Recarrega dados de custeio (placas, orlas, ferragens e máquinas).\n"
            "- Atualiza as tabelas e os gráficos.\n"
            "- Nota: se alterou 'Nao Stock' e ainda não gravou, clique primeiro em 'Gravar Nao Stock'."
        )
        self.btn_nao_stock_save = _make_dash_button(
            "Gravar Nao Stock",
            _dash_icon("document-save", QtWidgets.QStyle.SP_DialogSaveButton),
        )
        self.btn_nao_stock_save.setToolTip(
            "Grava na base de dados o estado da coluna 'Nao Stock' no Resumo de Placas.\n"
            "- Quando ativo, o custeio passa a considerar placas inteiras (ajusta desperdício/BLK conforme necessário).\n"
            "- Pode desfazer: desmarque e grave novamente.\n"
            "- Dica: o botão mostra '* ' quando existem alterações por gravar.\n"
            "- Atalho: Ctrl+G."
        )
        actions.addWidget(self.btn_dash_refresh)
        actions.addWidget(self.btn_nao_stock_save)
        actions.addStretch(1)
        layout.addLayout(actions)

        # tabelas principais (mais altura)
        top_split = QtWidgets.QSplitter(Qt.Horizontal)
        self.tbl_dash_placas = QtWidgets.QTableView()
        self.tbl_dash_orlas = QtWidgets.QTableView()
        top_split.addWidget(self._wrap_in_group("Resumo de Placas", self.tbl_dash_placas, self._wheel_forwarder_dashboard))
        top_split.addWidget(self._wrap_in_group("Resumo de Orlas", self.tbl_dash_orlas, self._wheel_forwarder_dashboard))
        top_split.setStretchFactor(0, 1)
        top_split.setStretchFactor(1, 1)
        layout.addWidget(top_split, 3)

        mid_split = QtWidgets.QSplitter(Qt.Horizontal)
        self.tbl_dash_ferr = QtWidgets.QTableView()
        self.tbl_dash_maq = QtWidgets.QTableView()
        mid_split.addWidget(self._wrap_in_group("Resumo de Ferragens", self.tbl_dash_ferr, self._wheel_forwarder_dashboard))
        mid_split.addWidget(self._wrap_in_group("Resumo de Máquinas / MO", self.tbl_dash_maq, self._wheel_forwarder_dashboard))
        mid_split.setStretchFactor(0, 1)
        mid_split.setStretchFactor(1, 1)
        layout.addWidget(mid_split, 3)

        bottom_split = QtWidgets.QSplitter(Qt.Horizontal)
        self.canvas_placas, self.ax_placas = self._create_canvas()
        self.canvas_orlas, self.ax_orlas = self._create_canvas()
        for c in (self.canvas_placas, self.canvas_orlas):
            c.setMinimumHeight(1075)
            c.installEventFilter(self._wheel_forwarder_dashboard)
        bottom_split.addWidget(self._wrap_in_group("Comparativo de Custos por Placa", self.canvas_placas, self._wheel_forwarder_dashboard))
        bottom_split.addWidget(self._wrap_in_group("Consumo de Orlas (ml)", self.canvas_orlas, self._wheel_forwarder_dashboard))
        bottom_split.setStretchFactor(0, 1)
        bottom_split.setStretchFactor(1, 1)
        layout.addWidget(bottom_split, 3)

        bottom2_split = QtWidgets.QSplitter(Qt.Horizontal)
        self.canvas_ferr, self.ax_ferr = self._create_canvas()
        self.canvas_ops, self.ax_ops = self._create_canvas()
        self.canvas_pie, self.ax_pie = self._create_canvas()
        for c in (self.canvas_ferr, self.canvas_ops, self.canvas_pie):
            c.setMinimumHeight(875)
            c.installEventFilter(self._wheel_forwarder_dashboard)
        bottom2_split.addWidget(self._wrap_in_group("Custos por Ferragem", self.canvas_ferr, self._wheel_forwarder_dashboard))
        bottom2_split.addWidget(self._wrap_in_group("Custos por Operação", self.canvas_ops, self._wheel_forwarder_dashboard))
        layout.addWidget(bottom2_split, 3)

        pie_box = self._wrap_in_group("Distribuição de Custos (Placas / Orlas / Ferragens / Máquinas / Margens)", self.canvas_pie, self._wheel_forwarder_dashboard)
        layout.addWidget(pie_box, 2)
        bottom2_split.setStretchFactor(0, 1)
        bottom2_split.setStretchFactor(1, 1)
        layout.addWidget(pie_box, 2)

        self.btn_dash_refresh.clicked.connect(self._refresh_dashboard)
        self.btn_nao_stock_save.clicked.connect(self._save_nao_stock_states)
        self._shortcut_save = QtGui.QShortcut(QtGui.QKeySequence("Ctrl+G"), self)
        self._shortcut_save.setContext(Qt.WidgetWithChildrenShortcut)
        self._shortcut_save.activated.connect(self._save_nao_stock_states)

        # Garantir que as figuras n�o "prendem" o scroll do rato
        for canvas in (self.canvas_placas, self.canvas_orlas, self.canvas_ferr, self.canvas_ops, self.canvas_pie):
            canvas.installEventFilter(self._wheel_forwarder_dashboard)

        scroll.setWidget(content)
        outer.addWidget(scroll)

    def _compute_resumos_dashboard(self) -> dict:
        result = {"placas": [], "orlas": [], "ferr": [], "maq": []}
        orc = self._current_orcamento
        if not orc:
            return result
        try:
            with SessionLocal() as session:
                cust_rows = (
                    session.query(CusteioItem)
                    .filter(CusteioItem.orcamento_id == orc.id, CusteioItem.versao == orc.versao)
                    .order_by(CusteioItem.ordem)
                    .all()
                )
                
                # ✅ Carregar também os OrcamentoItem para ter acesso à quantidade real
                orc_items = (
                    session.query(OrcamentoItem)
                    .filter(OrcamentoItem.id_orcamento == orc.id)
                    .all()
                )
                # Criar mapa: item_id -> OrcamentoItem.qt
                orc_item_qt_map = {item.id_item: float(item.qt or 1) for item in orc_items}

                backup_map: dict[int, CusteioDespBackup] = {}
                item_ids_all = [getattr(ci, "id", None) for ci in cust_rows if getattr(ci, "id", None)]
                if item_ids_all:
                    backups = (
                        session.query(CusteioDespBackup)
                        .filter(
                            CusteioDespBackup.custeio_item_id.in_(item_ids_all),
                            CusteioDespBackup.orcamento_id == orc.id,
                            CusteioDespBackup.versao == orc.versao,
                        )
                        .all()
                    )
                    backup_map = {bak.custeio_item_id: bak for bak in backups}

        except Exception:
            return result

        def _orla_width_from_esp(esp_val: float) -> float:
            try:
                esp = float(esp_val)
            except Exception:
                return 0.0
            if esp <= 0:
                return 0.0
            if esp < 20:
                return 23.0
            if esp < 31:
                return 35.0
            if esp < 40:
                return 45.0
            return 60.0

        def _safe_float(value: Any) -> float:
            try:
                return float(value)
            except Exception:
                return 0.0

        def _fmt_percent_value(value: Any) -> str:
            try:
                num = float(value)
            except Exception:
                return "-"
            if abs(num) <= 1.0:
                num *= 100.0
            return f"{num:.2f} %"

        def _fmt_currency_value(value: Any) -> str:
            try:
                num = float(value)
            except Exception:
                return "0.00 EUR"
            return f"{num:.2f} EUR"

        # ------------ Placas ------------
        placas_map: dict[tuple, dict] = {}
        for ci in cust_rows:
            und = (ci.und or "").upper()
            if und != "M2":
                continue
            
            # ✅ Obter a quantidade real do item de orçamento
            item_qt = orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0)
            
            comp_mp = float(ci.comp_mp or 0)
            larg_mp = float(ci.larg_mp or 0)
            try:
                desp_pct = float(ci.desp or 0)
            except Exception:
                desp_pct = 0.0
            desp_fraction = desp_pct / 100.0 if abs(desp_pct) > 1 else desp_pct
            qt_total = float(ci.qt_total or 0)
            pliq = float(ci.pliq or 0)
            area_placa = (comp_mp / 1000.0) * (larg_mp / 1000.0) if comp_mp and larg_mp else 0
            m2_total_pecas = float(ci.area_m2_und or 0) * qt_total * item_qt  # ✅ Multiplicar também por quantidade do item
            m2_consumidos = m2_total_pecas * (1 + desp_fraction)
            key = (ci.ref_le, ci.descricao_no_orcamento)
            if key not in placas_map:
                placas_map[key] = {
                    "ref_le": ci.ref_le,
                    "descricao_no_orcamento": ci.descricao_no_orcamento,
                    "pliq": pliq,
                    "und": und,
                    "desp": desp_pct,
                    "comp_mp": comp_mp,
                    "larg_mp": larg_mp,
                    "esp_mp": float(ci.esp_mp or 0),
                    "qt_placas_utilizadas": 0.0,
                    "area_placa": area_placa,
                    "m2_consumidos": 0.0,
                    "m2_total_pecas": 0.0,
                    "custo_mp_total": 0.0,
                    "custo_placas_utilizadas": 0.0,
                    "item_ids": [],
                    "nao_stock": False,
                    "_desp_original_values": [],
                    "_blk_original_values": [],
                    "_blk_current_values": [],
                }
            if area_placa > 0:
                placas_map[key]["area_placa"] = area_placa
            if pliq:
                placas_map[key]["pliq"] = pliq
            placas_map[key]["desp"] = desp_pct
            placas_map[key]["m2_consumidos"] += m2_consumidos
            placas_map[key]["m2_total_pecas"] += m2_total_pecas
            placas_map[key]["custo_mp_total"] += float(ci.custo_mp_total or 0) * item_qt  # ✅ Multiplicar por quantidade do item
            placas_map[key]["item_ids"].append(getattr(ci, "id", None))
            ci_id = getattr(ci, "id", None)
            placas_map[key]["_blk_current_values"].append(bool(getattr(ci, "blk", False)))
            if ci_id:
                bak = backup_map.get(ci_id)
                if bak:
                    if bak.desp_original is not None:
                        placas_map[key]["_desp_original_values"].append(float(bak.desp_original))
                    placas_map[key]["_blk_original_values"].append(bool(getattr(bak, "blk_original", False)))
                    if getattr(bak, "nao_stock_active", False):
                        placas_map[key]["nao_stock"] = True
        placas_rows: List[dict] = []
        for data in placas_map.values():
            area = data.get("area_placa") or 0
            total_m2 = data.get("m2_consumidos") or 0
            if area > 0 and total_m2 > 0:
                ratio = total_m2 / area
                qt_placas = math.ceil(ratio)
            else:
                qt_placas = 0
            data["qt_placas_utilizadas"] = float(qt_placas)
            data["custo_placas_utilizadas"] = qt_placas * area * (data.get("pliq") or 0)
            orig_list = [v for v in data.get("_desp_original_values", []) if v is not None]
            blk_orig_list = [v for v in data.get("_blk_original_values", []) if v is not None]
            blk_cur_list = [v for v in data.get("_blk_current_values", []) if v is not None]
            data["_desp_original"] = orig_list[0] if orig_list else None
            data["_blk_original"] = blk_orig_list[0] if blk_orig_list else None
            data["_blk_atual"] = blk_cur_list[-1] if blk_cur_list else None
            data.pop("_desp_original_values", None)
            data.pop("_blk_original_values", None)
            data.pop("_blk_current_values", None)

            tooltip_lines = [
                f"m2 de pecas: {_safe_float(data.get('m2_total_pecas')):.2f}",
                f"Desperdicio aplicado: {_fmt_percent_value(data.get('desp'))} -> m2 consumidos: {_safe_float(data.get('m2_consumidos')):.2f}",
            ]
            if area > 0 and qt_placas > 0:
                tooltip_lines.append(f"Area placa: {area:.2f} m2 | Qt placas: {qt_placas}")
            tooltip_lines.append(f"C.MP Tot (custeio): {_fmt_currency_value(data.get('custo_mp_total'))}")
            tooltip_lines.append(f"C.Placa Usada: {_fmt_currency_value(data.get('custo_placas_utilizadas'))}")
            if data.get("nao_stock") and data.get("_desp_original") is not None:
                tooltip_lines.append(f"Desp. original (NST): {_fmt_percent_value(data.get('_desp_original'))}")
            if data.get("_blk_original") is not None:
                tooltip_lines.append(f"BLK original: {'Ativo' if data.get('_blk_original') else 'Inativo'}")
            if data.get("_blk_atual") is not None:
                tooltip_lines.append(f"BLK atual: {'Ativo' if data.get('_blk_atual') else 'Inativo'}")
            data["_placa_tooltip"] = "\n".join(tooltip_lines)
            placas_rows.append(data)
        result["placas"] = placas_rows

        # ------------ Orlas ------------
        orlas_map: dict[tuple, dict] = {}
        for ci in cust_rows:
            # ✅ Obter a quantidade real do item de orçamento
            item_qt = orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0)
            
            largura_orla = _orla_width_from_esp(getattr(ci, "esp_res", None) or getattr(ci, "esp_mp", None))
            ref_map = {
                1: getattr(ci, "orl_0_4", None) or getattr(ci, "corres_orla_0_4", None) or "",
                2: getattr(ci, "orl_1_0", None) or getattr(ci, "corres_orla_1_0", None) or "",
            }
            sides = [
                (getattr(ci, "orl_c1", 0), getattr(ci, "ml_orl_c1", 0), getattr(ci, "custo_orl_c1", 0)),
                (getattr(ci, "orl_c2", 0), getattr(ci, "ml_orl_c2", 0), getattr(ci, "custo_orl_c2", 0)),
                (getattr(ci, "orl_l1", 0), getattr(ci, "ml_orl_l1", 0), getattr(ci, "custo_orl_l1", 0)),
                (getattr(ci, "orl_l2", 0), getattr(ci, "ml_orl_l2", 0), getattr(ci, "custo_orl_l2", 0)),
            ]
            for code_raw, ml_raw, custo_raw in sides:
                try:
                    code_val = float(code_raw or 0)
                except Exception:
                    code_val = 0.0
                if code_val <= 0:
                    continue
                if code_val >= 0.9:
                    code = 2
                    esp_descr = "1.0mm"
                else:
                    code = 1
                    esp_descr = "0.4mm"
                ref_orl = ref_map.get(code) or ""
                if not ref_orl:
                    continue
                ml_val = float(ml_raw or 0)
                custo_val = float(custo_raw or 0)
                
                # ✅ Multiplicar ML e custo pela quantidade do item
                ml_val *= item_qt
                custo_val *= item_qt
                
                if ml_val == 0 and custo_val == 0:
                    continue
                key = (ref_orl, ci.descricao_no_orcamento, esp_descr)
                if key not in orlas_map:
                    orlas_map[key] = {
                        "ref_orla": ref_orl,
                        "descricao_material": ci.descricao_no_orcamento,
                        "espessura_orla": esp_descr,
                        "largura_orla": largura_orla,
                        "ml_total": 0.0,
                        "custo_total": 0.0,
                    }
                orlas_map[key]["ml_total"] += ml_val
                orlas_map[key]["custo_total"] += custo_val
        result["orlas"] = list(orlas_map.values())

        # ------------ Ferragens ------------
        ferr_map: dict[tuple, dict] = {}
        for ci in cust_rows:
            ref = (ci.ref_le or "").upper()
            if not ref.startswith("FER"):
                continue
            
            # ✅ Obter a quantidade real do item de orçamento
            item_qt = orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0)
            
            key = (ci.ref_le, ci.descricao_no_orcamento)
            qt_total = float(ci.qt_total or 0)
            spp_ml_total = float(ci.spp_ml_und or 0) * qt_total
            if key not in ferr_map:
                ferr_map[key] = {
                    "ref_le": ci.ref_le,
                    "descricao_no_orcamento": ci.descricao_no_orcamento,
                    "pliq": float(ci.pliq or 0),
                    "und": ci.und,
                    "desp": float(ci.desp or 0),
                    "comp_mp": float(ci.comp_mp or 0),
                    "larg_mp": float(ci.larg_mp or 0),
                    "esp_mp": float(ci.esp_mp or 0),
                    "qt_total": 0.0,
                    "spp_ml_total": 0.0,
                    "custo_mp_und": float(ci.custo_mp_und or 0),
                    "custo_mp_total": 0.0,
                }
            ferr_map[key]["qt_total"] += qt_total * item_qt  # ✅ Multiplicar também por quantidade do item
            ferr_map[key]["spp_ml_total"] += spp_ml_total * item_qt  # ✅ Multiplicar também por quantidade do item
            ferr_map[key]["custo_mp_total"] += float(ci.custo_mp_total or 0) * item_qt  # ✅ Multiplicar por quantidade do item
        result["ferr"] = list(ferr_map.values())

        # ------------ Máquinas / MO ------------
        def get_cost(filter_attr: str) -> float:
            return sum(
                float(getattr(ci, f"{filter_attr}_und") or 0) * float(ci.qt_total or 0) * orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0)  # ✅ Multiplicar por quantidade do item
                for ci in cust_rows
                if getattr(ci, filter_attr) and float(getattr(ci, filter_attr) or 0) > 0
            )

        def get_ml_corte() -> float:
            total = 0.0
            for ci in cust_rows:
                if float(ci.cp01_sec or 0) <= 0:
                    continue
                # ✅ Multiplicar por quantidade do item
                item_qt = orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0)
                total += ((float(ci.comp_res or 0) * 2 + float(ci.larg_res or 0) * 2) * float(ci.qt_total or 0) * item_qt) / 1000.0
            return round(total, 2)

        def get_ml_orla() -> float:
            total = 0.0
            for ci in cust_rows:
                if float(ci.cp02_orl or 0) <= 0:
                    continue
                # ✅ Multiplicar por quantidade do item
                item_qt = orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0)
                total += item_qt * sum(float(x or 0) for x in [
                    getattr(ci, "ml_orl_c1", 0),
                    getattr(ci, "ml_orl_c2", 0),
                    getattr(ci, "ml_orl_l1", 0),
                    getattr(ci, "ml_orl_l2", 0),
                ])
            return round(total, 2)

        maq_rows = []
        maq_rows.append({
            "operacao": "Seccionadora (Corte)",
            "custo_total": round(get_cost("cp01_sec"), 2),
            "ml_corte": get_ml_corte(),
            "ml_orlado": "",
            "num_pecas": int(sum(float(ci.qt_total or 0) * orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0) for ci in cust_rows if float(ci.cp01_sec or 0) > 0)),  # ✅ Multiplicar por quantidade do item
        })
        maq_rows.append({
            "operacao": "Orladora (Orlagem)",
            "custo_total": round(get_cost("cp02_orl"), 2),
            "ml_corte": "",
            "ml_orlado": get_ml_orla(),
            "num_pecas": int(sum(float(ci.qt_total or 0) * orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0) for ci in cust_rows if float(ci.cp02_orl or 0) > 0)),  # ✅ Multiplicar por quantidade do item
        })
        maq_rows.append({
            "operacao": "CNC (Mecanizações)",
            "custo_total": round(get_cost("cp03_cnc"), 2),
            "ml_corte": "",
            "ml_orlado": "",
            "num_pecas": int(sum(float(ci.qt_total or 0) * orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0) for ci in cust_rows if float(ci.cp03_cnc or 0) > 0)),  # ✅ Multiplicar por quantidade do item
        })
        maq_rows.append({
            "operacao": "ABD (Mecanizações)",
            "custo_total": round(get_cost("cp04_abd"), 2),
            "ml_corte": "",
            "ml_orlado": "",
            "num_pecas": int(sum(float(ci.qt_total or 0) * orc_item_qt_map.get(getattr(ci, "item_id", None), 1.0) for ci in cust_rows if float(ci.cp04_abd or 0) > 0)),  # ✅ Multiplicar por quantidade do item
        })
        maq_rows.append({"operacao": "Prensa (Montagem)", "custo_total": round(get_cost("cp05_prensa"), 2), "ml_corte": "", "ml_orlado": "", "num_pecas": ""})
        maq_rows.append({"operacao": "Esquadrejadora (Cortes Manuais)", "custo_total": round(get_cost("cp06_esquad"), 2), "ml_corte": "", "ml_orlado": "", "num_pecas": ""})
        maq_rows.append({"operacao": "Embalamento (Paletização)", "custo_total": round(get_cost("cp07_embalagem"), 2), "ml_corte": "", "ml_orlado": "", "num_pecas": ""})
        maq_rows.append({"operacao": "Mão de Obra (MO geral)", "custo_total": round(get_cost("cp08_mao_de_obra"), 2), "ml_corte": "", "ml_orlado": "", "num_pecas": ""})
        result["maq"] = maq_rows
        return result

    def _refresh_dashboard(self) -> None:
        if not MATPLOTLIB_AVAILABLE:
            QtWidgets.QMessageBox.information(self, "Resumo de Consumos", "Instale 'matplotlib' para usar o dashboard.")
            return
        orc = self._current_orcamento
        cli = self._current_client
        if self._dash_info_labels:
            self._dash_info_labels["ano"].setText(str(getattr(orc, "ano", "") or "-"))
            self._dash_info_labels["cliente"].setText(self._cliente_display_nome() or "-")
            self._dash_info_labels["orc"].setText(str(getattr(orc, "num_orcamento", "") or "-"))
            self._dash_info_labels["versao"].setText(self._format_versao(getattr(orc, "versao", "")) if orc else "-")
            self._dash_info_labels["user"].setText(str(getattr(self.current_user, "username", "") or "-"))
        try:
            self._dashboard_data = self._compute_resumos_dashboard()
            self._update_dashboard_ui()
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "Resumo de Consumos", f"Falha ao atualizar dashboard: {exc}")

    def _calc_custo_mp_total_estimate(self, ci: CusteioItem) -> float:
        """
        Estima o custo MP total replicando a logica principal (base * (1+desp) * pliq * qt_total).
        """
        try:
            und = (ci.und or "").upper()
        except Exception:
            return 0.0

        try:
            desp_pct = float(ci.desp or 0) or 0.0
        except Exception:
            desp_pct = 0.0
        desp = desp_pct / 100.0 if abs(desp_pct) > 1 else desp_pct

        try:
            pliq = float(ci.pliq or 0) or 0.0
        except Exception:
            pliq = 0.0

        try:
            qt_total = float(ci.qt_total or 0) or 0.0
        except Exception:
            qt_total = 0.0

        if pliq <= 0 or qt_total <= 0:
            return 0.0

        base = 0.0
        if und == "M2":
            base = float(ci.area_m2_und or 0) or 0.0
        elif und == "ML":
            base = float(ci.spp_ml_und or 0) or 0.0
        elif und == "UND":
            base = 1.0

        if base <= 0:
            return 0.0

        custo_und = base * (1.0 + desp) * pliq
        return custo_und * qt_total

    def _apply_nao_stock_toggle(self, row_data: dict, checked: bool) -> None:
        """
        Ajusta o valor de 'desp' das linhas de PLACAS para aproximar C.MP Tot de C.Placa Usad.
        """
        orc = self._current_orcamento
        if not orc:
            return

        item_ids = [iid for iid in row_data.get("item_ids", []) if iid]
        if not item_ids:
            return

        target_total = float(row_data.get("custo_placas_utilizadas") or 0) or 0.0

        try:
            with SessionLocal() as session:
                itens = (
                    session.query(CusteioItem)
                    .filter(CusteioItem.id.in_(item_ids))
                    .all()
                )
                placas_items = [ci for ci in itens if (ci.familia or "").upper() == "PLACAS"]
                if not placas_items:
                    return

                touched_items = set()

                if checked:
                    current_total = sum(self._calc_custo_mp_total_estimate(ci) for ci in placas_items)

                    # Step 1: Backup original values
                    for ci in placas_items:
                        blk_before = bool(getattr(ci, "blk", False))
                        existing = (
                            session.query(CusteioDespBackup)
                            .filter(CusteioDespBackup.custeio_item_id == ci.id)
                            .first()
                        )
                        if not existing:
                            session.add(
                                CusteioDespBackup(
                                    orcamento_id=orc.id,
                                    versao=orc.versao,
                                    user_id=getattr(self.current_user, "id", None),
                                    custeio_item_id=ci.id,
                                    desp_original=ci.desp or 0,
                                    blk_original=blk_before,
                                )
                            )
                        else:
                            if existing.blk_original is None:
                                existing.blk_original = blk_before
                            existing.nao_stock_active = True

                    if target_total <= 0 or current_total <= 0:
                        session.commit()
                        return

                    # Step 2: Calculate aggregate qt_total (NEW - considering item quantities)
                    qt_total_agregada = sum(
                        float(ci.qt_total or 0) or 0.0 
                        for ci in placas_items
                    )

                    # Step 3: Calculate custo_excesso and distribute proportionally
                    custo_excesso = max(target_total - current_total, 0.0)
                    
                    for ci in placas_items:
                        # Mark BLK flag
                        blk_before = bool(getattr(ci, "blk", False))
                        if not blk_before:
                            try:
                                ci.blk = True
                            except Exception:
                                pass

                        # Calculate new desp considering item quantity
                        try:
                            qt_item = float(ci.qt_total or 0) or 0.0
                        except Exception:
                            qt_item = 0.0

                        if qt_total_agregada > 0 and custo_excesso > 0:
                            # Distribute excess cost proportionally to item quantity
                            desp_absoluto_item = (custo_excesso * qt_item) / qt_total_agregada
                            
                            # Get unit cost without new desp (to convert from absolute to percentage)
                            try:
                                und = (ci.und or "").upper()
                            except Exception:
                                und = "UND"
                            
                            base = 0.0
                            if und == "M2":
                                base = float(ci.area_m2_und or 0) or 0.0
                            elif und == "ML":
                                base = float(ci.spp_ml_und or 0) or 0.0
                            elif und == "UND":
                                base = 1.0

                            try:
                                pliq = float(ci.pliq or 0) or 0.0
                            except Exception:
                                pliq = 0.0

                            if base > 0 and pliq > 0:
                                custo_unitario = base * pliq
                                novo_desp_pct = (desp_absoluto_item / custo_unitario) * 100.0
                                novo_desp_pct = max(novo_desp_pct, 0.0)  # Prevent negative
                            else:
                                novo_desp_pct = 0.0
                        else:
                            # Fallback: no excess cost or invalid data
                            novo_desp_pct = 0.0

                        # Apply new desp percentage
                        try:
                            ci.desp = Decimal(str(novo_desp_pct)).quantize(Decimal("0.0001"))
                        except Exception:
                            ci.desp = novo_desp_pct
                        
                        item_id = getattr(ci, "item_id", None)
                        if item_id:
                            touched_items.add(item_id)
                else:
                    backups = (
                        session.query(CusteioDespBackup)
                        .filter(CusteioDespBackup.custeio_item_id.in_([ci.id for ci in placas_items]))
                        .all()
                    )
                    backup_map = {b.custeio_item_id: b for b in backups}
                    for ci in placas_items:
                        bak = backup_map.get(ci.id)
                        if not bak:
                            continue
                        ci.desp = bak.desp_original
                        bak.nao_stock_active = False
                        blk_restore = bak.blk_original
                        if blk_restore is not None:
                            try:
                                ci.blk = bool(blk_restore)
                            except Exception:
                                ci.blk = blk_restore
                        item_id = getattr(ci, "item_id", None)
                        if item_id:
                            touched_items.add(item_id)

                for item_id in touched_items:
                    try:
                        atualizar_orlas_custeio(session, orcamento_id=orc.id, item_id=item_id)
                    except Exception:
                        logger.exception("Falha ao recalcular custos para item %s", item_id)

                session.commit()
        except OperationalError as exc:
            QtWidgets.QMessageBox.warning(
                self,
                "Resumo de Consumos",
                f"Falha ao aplicar Nao Stock: {exc.orig if hasattr(exc, 'orig') else exc}",
            )
        except Exception:
            logger.exception("Erro ao aplicar Nao Stock")
            QtWidgets.QMessageBox.warning(
                self,
                "Resumo de Consumos",
                "Falha ao aplicar Nao Stock. Verifique os dados e tente novamente.",
            )

    def _on_dash_placas_changed(self, top_left: QtCore.QModelIndex, bottom_right: QtCore.QModelIndex, roles=None) -> None:
        """
        Handler para edicao da coluna 'Nao Stock' no resumo de placas.
        """
        model = self.tbl_dash_placas.model()
        if not isinstance(model, SimpleTableModel):
            return

        cols = getattr(model, "columns", getattr(model, "_columns", []))
        if not cols:
            return

        nao_stock_col = None
        for idx, col in enumerate(cols):
            spec = model._col_spec(col)
            if spec.get("attr") == "nao_stock":
                nao_stock_col = idx
                break

        if nao_stock_col is None:
            return

        if top_left.column() > nao_stock_col or bottom_right.column() < nao_stock_col:
            return

        try:
            row_obj = model.get_row(top_left.row())
        except Exception:
            return

        checked = False
        try:
            checked = bool(row_obj.get("nao_stock"))
        except Exception:
            checked = False

        self._apply_nao_stock_toggle(row_obj, checked)
        self._set_nao_stock_dirty(True)
        self._refresh_dashboard()

    def _save_nao_stock_states(self) -> None:
        """
        Grava em base de dados o estado atual da coluna Nao Stock para todas as linhas exibidas.
        """
        model = self.tbl_dash_placas.model()
        if not isinstance(model, SimpleTableModel):
            return
        rows = getattr(model, "_rows", [])
        for row in rows:
            try:
                checked = bool(row.get("nao_stock"))
            except Exception:
                checked = False
            self._apply_nao_stock_toggle(row, checked)
        self._refresh_dashboard()
        self._set_nao_stock_dirty(False)


    def _update_dashboard_ui(self) -> None:
        data = self._dashboard_data or {}
        # tabelas
        def fmt_auto(v):
            if v is None or v == "":
                return ""
            try:
                val = float(v)
            except Exception:
                return v
            if abs(val - round(val)) < 0.005:
                return f"{round(val):.0f}"
            if abs(val*10 - round(val*10)) < 0.05:
                return f"{val:.1f}"
            return f"{val:.2f}"

        def fmt_unit(unit: str, decimals: Optional[int] = None):
            def _f(v):
                if v is None or v == "":
                    return ""
                try:
                    num = float(v)
                except Exception:
                    return v
                if decimals is None:
                    try:
                        val = float(f"{num:.2f}")
                    except Exception:
                        val = num
                    if abs(val - round(val)) < 0.005:
                        txt = f"{round(val):.0f}"
                    elif abs(val * 10 - round(val * 10)) < 0.05:
                        txt = f"{val:.1f}"
                    else:
                        txt = f"{val:.2f}"
                else:
                    txt = f"{num:.{decimals}f}"
                return f"{txt} {unit}" if unit else txt
            return _f

        moedas = fmt_unit("€", 2)
        m2_fmt = fmt_unit("m2", 2)
        ml_fmt = fmt_unit("ml", 2)
        mm_fmt = fmt_unit("mm", 0)


        def _percent(val: Any) -> str:
            try:
                num = float(val)
            except Exception:
                return "" if val in (None, "") else str(val)
            # sempre 2 casas decimais para percentagem
            if abs(num) <= 1.0:
                return f"{num * 100:.2f} %"
            return f"{num:.2f} %"

        def _placa_row_tooltip(row: dict, value: Any, spec=None) -> Optional[str]:
            return row.get("_placa_tooltip")

        def _desp_original_tooltip(row: dict, value: Any, spec=None) -> Optional[str]:
            base_tip = row.get("_placa_tooltip")
            if row.get("nao_stock") and row.get("_desp_original") is not None:
                original_txt = _percent(row.get("_desp_original"))
                extra = f"Desp. original: {original_txt}"
                if base_tip:
                    return f"{extra}\n{base_tip}"
                return extra
            return base_tip

        placas_cols = [
            ("Ref.", "ref_le", None, _placa_row_tooltip),
            ("Descricao", "descricao_no_orcamento", None, _placa_row_tooltip),
            ("P.Liq", "pliq", moedas, _placa_row_tooltip),
            ("Und", "und", None, _placa_row_tooltip),
            ("Desp.", "desp", _percent, _desp_original_tooltip),
            ("Comp.", "comp_mp", mm_fmt, _placa_row_tooltip),
            ("Larg.", "larg_mp", mm_fmt, _placa_row_tooltip),
            ("Esp.", "esp_mp", mm_fmt, _placa_row_tooltip),
            ("Qt.Pla.", "qt_placas_utilizadas", fmt_auto, _placa_row_tooltip),
            ("Area", "area_placa", m2_fmt, _placa_row_tooltip),
            ("m2 Usad.", "m2_consumidos", m2_fmt, _placa_row_tooltip),
            ("m2_total_pecas", "m2_total_pecas", m2_fmt, _placa_row_tooltip),
            ("C.MP Tot", "custo_mp_total", moedas, _placa_row_tooltip),
            ("C.Placa Usad.", "custo_placas_utilizadas", moedas, _placa_row_tooltip),
            {"header": "Nao Stock", "attr": "nao_stock", "type": "bool", "editable": True, "tooltip": _placa_row_tooltip},
        ]
        orlas_cols = [
            ("Ref. Orla", "ref_orla"),
            ("Descr. Mat.", "descricao_material"),
            ("Esp.", "espessura_orla"),
            ("Larg.", "largura_orla", mm_fmt),
            ("ML Tot.", "ml_total", ml_fmt),
            ("Custo Tot", "custo_total", moedas),
        ]
        def _percent(val: Any) -> str:
            try:
                num = float(val)
            except Exception:
                return "" if val in (None, "") else str(val)
            if abs(num) < 1.0:
                return f"{num:.0%}"
            return f"{num:.0f} %"

        ferr_cols = [
            ("Ref.", "ref_le"),
            ("Descrição", "descricao_no_orcamento"),
            ("P.Liq", "pliq", moedas),
            ("Und", "und"),
            ("Desp.", "desp", _percent),
            ("Comp.", "comp_mp", mm_fmt),
            ("Larg.", "larg_mp", mm_fmt),
            ("Esp.", "esp_mp", mm_fmt),
            ("Qt", "qt_total", fmt_auto),
            ("ML Sup.", "spp_ml_total", ml_fmt),
            ("Custo Und", "custo_mp_und", moedas),
            ("Custo Tot", "custo_mp_total", moedas),
        ]
        maq_cols = [
            ("Operação", "operacao"),
            ("Custo Total (€)", "custo_total", moedas),
            ("ML Corte", "ml_corte", ml_fmt),
            ("ML Orlado", "ml_orlado", ml_fmt),
            ("Nº Peças", "num_pecas", fmt_auto),
        ]

        model_placas = SimpleTableModel(data.get("placas", []), placas_cols, self.tbl_dash_placas)
        old_model = getattr(self, "_dash_placas_model", None)
        if old_model is not None and old_model is not model_placas:
            try:
                old_model.dataChanged.disconnect(self._on_dash_placas_changed)
            except Exception:
                pass
        self._dash_placas_model = model_placas
        self.tbl_dash_placas.setModel(model_placas)
        self.tbl_dash_orlas.setModel(SimpleTableModel(data.get("orlas", []), orlas_cols, self.tbl_dash_orlas))
        self.tbl_dash_ferr.setModel(SimpleTableModel(data.get("ferr", []), ferr_cols, self.tbl_dash_ferr))
        self.tbl_dash_maq.setModel(SimpleTableModel(data.get("maq", []), maq_cols, self.tbl_dash_maq))

        tooltips = {
            "Resumo de Placas": "Placas (M2): área usada, despesa, custo teórico (C.MP Tot) e custo real consumido (C.Placa Usad.). Origem: custeio_items.",
            "Resumo de Orlas": "Orlas por referência/espessura, ML total e custo total. Origem: custeio_items (ml_orl_*, custo_total_orla).",
            "Resumo de Ferragens": "Ferragens com quantidades/ML/custos unitário e total. Origem: custeio_items (custo_mp_total, spp_ml_und).",
            "Resumo de Máquinas / MO": "Custos por operação e ML corte/orlado; nº peças processadas. Origem: custeio_items (cp01..cp08).",
        }
        for tbl, title in (
            (self.tbl_dash_placas, "Resumo de Placas"),
            (self.tbl_dash_orlas, "Resumo de Orlas"),
            (self.tbl_dash_ferr, "Resumo de Ferragens"),
            (self.tbl_dash_maq, "Resumo de Maquinas / MO"),
        ):
            tbl.setAlternatingRowColors(True)
            header = tbl.horizontalHeader()
            font = header.font()
            font.setBold(True)
            header.setFont(font)
            header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
            header.setStretchLastSection(False)
            tbl.verticalHeader().setVisible(False)
            tbl.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
            tbl.setMinimumHeight(460)
            if title in tooltips:
                header.setToolTip(tooltips[title])

        # ligar toggles da coluna Não Stock
        model_placas = self.tbl_dash_placas.model()
        if isinstance(model_placas, SimpleTableModel):
            try:
                model_placas.dataChanged.disconnect(self._on_dash_placas_changed)
            except Exception:
                pass
            try:
                model_placas.dataChanged.connect(self._on_dash_placas_changed, QtCore.Qt.UniqueConnection)
            except Exception:
                pass


        # gráficos
        self._plot_dashboard_charts()

    def _plot_dashboard_charts(self) -> None:
        data = self._dashboard_data or {}
        def wrap_label(text: str, max_len: int = 18) -> str:
            if not text:
                return ""
            if len(text) <= max_len:
                return text
            parts = text.split()
            lines = []
            current = ""
            for p in parts:
                if len(current) + len(p) + 1 <= max_len:
                    current = f"{current} {p}".strip()
                else:
                    if current:
                        lines.append(current)
                    current = p
            if current:
                lines.append(current)
            return "\n".join(lines) if lines else text

        # Placas
        placas = data.get("placas", [])
        self.ax_placas.clear()
        if placas:
            labels = [wrap_label(p.get("descricao_no_orcamento") or p.get("ref_le") or "") for p in placas]
            x = list(range(len(labels)))
            teor = [float(p.get("custo_mp_total", 0) or 0) for p in placas]
            real = [float(p.get("custo_placas_utilizadas", 0) or 0) for p in placas]
            width = 0.35
            teor_colors = []
            tol = 2.0
            for t, r in zip(teor, real):
                teor_colors.append("#4caf50" if abs(t - r) <= tol else "#7ec0ee")
            bars_teor = self.ax_placas.bar(
                [xi - width / 2 for xi in x], teor, width=width, label="Custo Teórico", color=teor_colors
            )
            bars_real = self.ax_placas.bar(
                [xi + width / 2 for xi in x], real, width=width, label="Custo Real", color="#ef5350"
            )
            self.ax_placas.set_xticks(x)
            self.ax_placas.set_xticklabels(labels, rotation=20, ha="right")
            self.ax_placas.set_ylabel("Custo (€)")
            self.ax_placas.tick_params(axis="x", labelsize=8)
            self.ax_placas.tick_params(axis="y", labelsize=8)
            self.ax_placas.legend(fontsize=9)
            for bar, val in zip(bars_teor, teor):
                self.ax_placas.text(
                    bar.get_x() + bar.get_width() / 2, val, f"{val:.2f}", ha="center", va="bottom", fontsize=8
                )
            for bar, val in zip(bars_real, real):
                self.ax_placas.text(
                    bar.get_x() + bar.get_width() / 2, val, f"{val:.2f}", ha="center", va="bottom", fontsize=8
                )
            self.ax_placas.set_title("Comparativo de Custos por Placa")
        else:
            self.ax_placas.text(0.5, 0.5, "Sem dados", ha="center", va="center")
        self.canvas_placas.draw_idle()

        # Orlas
        orlas = data.get("orlas", [])
        self.ax_orlas.clear()
        if orlas:
            labels = [wrap_label(f"{o.get('ref_orla','')} ({o.get('espessura_orla','')})") for o in orlas]
            x = list(range(len(labels)))
            ml = [float(o.get("ml_total", 0) or 0) for o in orlas]
            bars = self.ax_orlas.bar(x, ml, color="#ffa726")
            self.ax_orlas.set_xticks(x)
            self.ax_orlas.set_xticklabels(labels, rotation=25, ha="right")
            self.ax_orlas.set_ylabel("Metros Lineares (ml)")
            self.ax_orlas.tick_params(axis="x", labelsize=8)
            self.ax_orlas.tick_params(axis="y", labelsize=8)
            self.ax_orlas.set_title("Consumo de Orlas (ml)")
            for rect, val in zip(bars, ml):
                self.ax_orlas.text(rect.get_x() + rect.get_width() / 2, rect.get_height(), f"{val:.2f}", ha="center", va="bottom", fontsize=8)
        else:
            self.ax_orlas.text(0.5, 0.5, "Sem dados", ha="center", va="center")
        self.canvas_orlas.draw_idle()

        # Ferragens
        ferr = data.get("ferr", [])
        self.ax_ferr.clear()
        if ferr:
            labels = [wrap_label(f.get("descricao_no_orcamento") or f.get("ref_le") or "") for f in ferr]
            x = list(range(len(labels)))
            custos = [float(f.get("custo_mp_total", 0) or 0) for f in ferr]
            bars = self.ax_ferr.bar(x, custos, color="#ef5350")
            self.ax_ferr.set_xticks(x)
            self.ax_ferr.set_xticklabels(labels, rotation=25, ha="right")
            self.ax_ferr.set_ylabel("Custo (€)")
            self.ax_ferr.tick_params(axis="x", labelsize=8)
            self.ax_ferr.tick_params(axis="y", labelsize=8)
            self.ax_ferr.set_title("Custos por Ferragem")
            for rect, val in zip(bars, custos):
                self.ax_ferr.text(rect.get_x() + rect.get_width() / 2, rect.get_height(), f"{val:.2f}", ha="center", va="bottom", fontsize=8)
        else:
            self.ax_ferr.text(0.5, 0.5, "Sem dados", ha="center", va="center")
        self.canvas_ferr.draw_idle()

        # Operações
        maq = data.get("maq", [])
        self.ax_ops.clear()
        if maq:
            labels = [wrap_label(m.get("operacao", "")) for m in maq]
            x = list(range(len(labels)))
            custos = [float(m.get("custo_total", 0) or 0) for m in maq]
            bars = self.ax_ops.bar(x, custos, color="#42a5f5")
            self.ax_ops.set_xticks(x)
            self.ax_ops.set_xticklabels(labels, rotation=20, ha="right")
            self.ax_ops.set_ylabel("Custo (€)")
            self.ax_ops.tick_params(axis="x", labelsize=8)
            self.ax_ops.tick_params(axis="y", labelsize=8)
            self.ax_ops.set_title("Custos por Operação")
            for rect, val in zip(bars, custos):
                self.ax_ops.text(rect.get_x() + rect.get_width() / 2, rect.get_height(), f"{val:.2f}", ha="center", va="bottom", fontsize=8)
        else:
            self.ax_ops.text(0.5, 0.5, "Sem dados", ha="center", va="center")
        self.canvas_ops.draw_idle()

        # Pizza
        self.ax_pie.clear()

        # Margens (a partir de `orcamento_items`)
        margem_lucro_total = 0.0
        margem_acab_total = 0.0
        margem_mo_total = 0.0
        custos_admin_total = 0.0
        margem_mp_total = 0.0
        base_custo_total = 0.0
        total_venda = 0.0
        orc = self._current_orcamento
        if orc is not None:
            try:
                with SessionLocal() as session:
                    itens_orc = (
                        session.query(OrcamentoItem)
                        .filter(
                            OrcamentoItem.id_orcamento == orc.id,
                            OrcamentoItem.versao == orc.versao,
                        )
                        .all()
                    )
                    for it in itens_orc:
                        try:
                            qt = float(getattr(it, "qt", 0) or 0)
                        except Exception:
                            qt = 0.0

                        try:
                            base_custo_total += float(getattr(it, "custo_produzido", 0) or 0) * qt
                        except Exception:
                            pass

                        try:
                            preco_total = float(getattr(it, "preco_total", 0) or 0)
                        except Exception:
                            preco_total = 0.0
                        if preco_total == 0.0 and qt:
                            try:
                                preco_total = float(getattr(it, "preco_unitario", 0) or 0) * qt
                            except Exception:
                                preco_total = 0.0
                        total_venda += preco_total

                        def _acc(attr: str) -> float:
                            try:
                                return float(getattr(it, attr, 0) or 0) * qt
                            except Exception:
                                return 0.0

                        margem_lucro_total += _acc("valor_margem")
                        margem_acab_total += _acc("valor_acabamentos")
                        margem_mo_total += _acc("valor_mao_obra")
                        custos_admin_total += _acc("valor_custos_admin")
                        margem_mp_total += _acc("valor_mp_orlas")
            except Exception:
                pass

        margens_total = (
            margem_lucro_total
            + margem_acab_total
            + margem_mo_total
            + custos_admin_total
            + margem_mp_total
        )
        total_placas = sum(float(p.get("custo_placas_utilizadas", 0) or 0) for p in placas)
        total_orlas = sum(float(o.get("custo_total", 0) or 0) for o in orlas)
        total_ferr = sum(float(f.get("custo_mp_total", 0) or 0) for f in ferr)
        total_maq = sum(float(m.get("custo_total", 0) or 0) for m in maq)
        valores = [total_placas, total_orlas, total_ferr, total_maq]
        labels_pie = ["Placas", "Orlas", "Ferragens", "Máquinas/MO"]
        if margens_total > 0:
            valores.append(margens_total)
            labels_pie.append("Margens")
        if sum(valores) > 0:
            def _autopct_eur(pct: float) -> str:
                total = sum(valores)
                val = (pct / 100.0) * total
                if pct <= 0 or val <= 0:
                    return ""
                return f"{val:.2f}€\n{pct:.1f}%"

            wedges, texts, autotexts = self.ax_pie.pie(
                valores,
                labels=labels_pie,
                autopct=_autopct_eur,
                startangle=90,
                pctdistance=0.8,
            )
            for t in texts + autotexts:
                t.set_fontsize(8)
            self.ax_pie.set_title("Distribuição de Custos")
            # tabela ao lado com valores em €
            total_sum = sum(valores)
            table_data = [["Tipo", "€", "%"]]

            def _fmt_eur(val: float) -> str:
                try:
                    return f"{float(val):.2f}€"
                except Exception:
                    return "0.00€"

            def _fmt_pct(val: float, total: float) -> str:
                if not total:
                    return "0.0%"
                return f"{(val / total * 100.0):.1f}%"

            def _fmt_rate(val: float, base: float) -> str:
                if not base:
                    return "0.0%"
                return f"{(val / base * 100.0):.1f}%"

            # Linhas do gráfico (custos + margens total)
            for label, val in zip(labels_pie, valores):
                table_data.append([label, _fmt_eur(val), _fmt_pct(val, total_sum)])

            # Detalhe de margens (valor em € + taxa efetiva % sobre custo produzido total)
            table_data.extend(
                [
                    ["Margens", _fmt_eur(margens_total), _fmt_pct(margens_total, total_sum)] if "Margens" not in labels_pie else ["", "", ""],
                    ["Margem Lucro", _fmt_eur(margem_lucro_total), _fmt_rate(margem_lucro_total, base_custo_total)],
                    ["Margem Acabamentos", _fmt_eur(margem_acab_total), _fmt_rate(margem_acab_total, base_custo_total)],
                    ["Margem Mão de Obra", _fmt_eur(margem_mo_total), _fmt_rate(margem_mo_total, base_custo_total)],
                    ["Custos Administrativos", _fmt_eur(custos_admin_total), _fmt_rate(custos_admin_total, base_custo_total)],
                    ["Margem Materias Primas", _fmt_eur(margem_mp_total), _fmt_rate(margem_mp_total, base_custo_total)],
                ]
            )

            # Total de venda do orçamento
            if total_venda:
                table_data.append(["Total Venda", _fmt_eur(total_venda), ""])

            # remover linha "Margens" extra se já existir como fatia no gráfico
            table_data = [row for row in table_data if row != ["", "", ""]]

            table = self.ax_pie.table(
                cellText=table_data,
                colWidths=[0.8, 0.45, 0.35],
                cellLoc="center",
                bbox=[1.12, 0.04, 0.98, 0.92],
            )
            table.auto_set_font_size(False)
            table.set_fontsize(8)

            bold_rows = {0}
            if "Margens" in labels_pie:
                bold_rows.add(1 + labels_pie.index("Margens"))
            if total_venda:
                bold_rows.add(len(table_data) - 1)
            for (row, col), cell in table.get_celld().items():
                if row in bold_rows:
                    cell.set_text_props(weight="bold")
        else:
            self.ax_pie.text(0.5, 0.5, "Sem dados", ha="center", va="center")
        self.canvas_pie.draw_idle()

    def _export_dashboard_pdf(self) -> None:
        if not MATPLOTLIB_AVAILABLE:
            self._show_toast(self.btn_dash_export, "Instale 'matplotlib' para exportar o dashboard.", timeout_ms=3000)
            return
        if not self._current_orcamento:
            self._show_toast(self.btn_dash_export, "Selecione um orçamento.", timeout_ms=3000)
            return
        # garantir dados atualizados
        self._refresh_dashboard()
        data = self._dashboard_data or {}
        orc = self._current_orcamento
        client = self._current_client
        export_dir = self._require_export_folder(orc, client, title="Dashboard")
        if export_dir is None:
            return
        fname = f"Resumo_Custos_Dashboard_{orc.num_orcamento or 'ORC'}_{self._format_versao(orc.versao)}.pdf"
        dest = export_dir / fname

        cliente_nome = self._cliente_display_nome() or "-"
        header_line = f"Ano: {getattr(orc, 'ano', '') or '-'}  |  Cliente: {cliente_nome}  |  Nº Orçamento: {orc.num_orcamento or ''}  |  Versão: {self._format_versao(orc.versao)}  |  Utilizador: {getattr(self.current_user, 'username', '') or '-'}"
        today = QtCore.QDate.currentDate().toString("dd/MM/yyyy")

        def add_header_footer(fig: Figure, page_idx: int, total: int) -> None:
            fig.text(0.02, 0.98, header_line, va="top", ha="left", fontsize=9, fontweight="bold")
            fig.text(0.02, 0.02, today, ha="left", va="bottom", fontsize=8)
            fig.text(0.98, 0.02, f"{page_idx}/{total}", ha="right", va="bottom", fontsize=8)

        # Definir novamente colunas/formatos para uso no PDF (evita NameError)
        def fmt_auto(v):
            if v is None or v == "":
                return ""
            try:
                val = float(v)
            except Exception:
                return v
            if abs(val - round(val)) < 0.005:
                return f"{round(val):.0f}"
            if abs(val * 10 - round(val * 10)) < 0.05:
                return f"{val:.1f}"
            return f"{val:.2f}"

        def fmt_unit(unit: str, decimals: Optional[int] = None):
            def _f(v):
                if v is None or v == "":
                    return ""
                try:
                    num = float(v)
                except Exception:
                    return v
                if decimals is None:
                    if abs(num - round(num)) < 0.005:
                        txt = f"{round(num):.0f}"
                    elif abs(num * 10 - round(num * 10)) < 0.05:
                        txt = f"{num:.1f}"
                    else:
                        txt = f"{num:.2f}"
                else:
                    txt = f"{num:.{decimals}f}"
                return f"{txt} {unit}" if unit else txt
            return _f

        moedas = fmt_unit("€", 2)
        m2_fmt = fmt_unit("m2", 2)
        ml_fmt = fmt_unit("ml", 2)
        mm_fmt = fmt_unit("mm", 0)

        placas_cols = [
            ("Ref.", "ref_le", None),
            ("Descrição", "descricao_no_orcamento", None),
            ("P.Liq", "pliq", moedas),
            ("Und", "und", None),
            ("Desp.", "desp", fmt_auto),
            ("Comp.", "comp_mp", mm_fmt),
            ("Larg.", "larg_mp", mm_fmt),
            ("Esp.", "esp_mp", mm_fmt),
            ("Qt.Pla.", "qt_placas_utilizadas", fmt_auto),
            ("Área", "area_placa", m2_fmt),
            ("m2 Usad.", "m2_consumidos", m2_fmt),
            ("m2_total_pecas", "m2_total_pecas", m2_fmt),
            ("C.MP Tot", "custo_mp_total", moedas),
            ("C.Placa Usad.", "custo_placas_utilizadas", moedas),
        ]
        orlas_cols = [
            ("Ref. Orla", "ref_orla", None),
            ("Descr. Mat.", "descricao_material", None),
            ("Esp.", "espessura_orla", None),
            ("Larg.", "largura_orla", mm_fmt),
            ("ML Tot.", "ml_total", ml_fmt),
            ("Custo Tot", "custo_total", moedas),
        ]
        ferr_cols = [
            ("Ref.", "ref_le", None),
            ("Descrição", "descricao_no_orcamento", None),
            ("P.Liq", "pliq", moedas),
            ("Und", "und", None),
            ("Desp.", "desp", fmt_auto),
            ("Comp.", "comp_mp", mm_fmt),
            ("Larg.", "larg_mp", mm_fmt),
            ("Esp.", "esp_mp", mm_fmt),
            ("Qt", "qt_total", fmt_auto),
            ("ML Sup.", "spp_ml_total", ml_fmt),
            ("Custo Und", "custo_mp_und", moedas),
            ("Custo Tot", "custo_mp_total", moedas),
        ]
        maq_cols = [
            ("Operação", "operacao", None),
            ("Custo Total (€)", "custo_total", moedas),
            ("ML Corte", "ml_corte", ml_fmt),
            ("ML Orlado", "ml_orlado", ml_fmt),
            ("Nº Peças", "num_pecas", fmt_auto),
        ]

        # Páginas de gráficos: 2 gráficos por página
        def build_chart_page(pairs: List[Tuple[str, List[Any], str]]) -> Figure:
            fig = Figure(figsize=(11.7, 8.3))
            fig.subplots_adjust(top=0.9, hspace=0.4, wspace=0.25, left=0.08, right=0.95)
            axes = [fig.add_subplot(121), fig.add_subplot(122)]
            def wrap_label(text: str, max_len: int = 18) -> str:
                if not text:
                    return ""
                if len(text) <= max_len:
                    return text
                parts = text.split()
                lines = []
                current = ""
                for p in parts:
                    if len(current) + len(p) + 1 <= max_len:
                        current = f"{current} {p}".strip()
                    else:
                        if current:
                            lines.append(current)
                        current = p
                if current:
                    lines.append(current)
                return "\n".join(lines) if lines else text
            for ax, (chart_type, dataset, title_txt) in zip(axes, pairs):
                if chart_type == "placas":
                    lbls = [wrap_label(p.get("descricao_no_orcamento") or p.get("ref_le") or "") for p in dataset]
                    xvals = list(range(len(lbls)))
                    teor = [float(p.get("custo_mp_total", 0) or 0) for p in dataset]
                    real = [float(p.get("custo_placas_utilizadas", 0) or 0) for p in dataset]
                    width = 0.35
                    ax.bar([xi - width / 2 for xi in xvals], teor, width=width, label="Custo Teórico", color="#7ec0ee")
                    ax.bar([xi + width / 2 for xi in xvals], real, width=width, label="Custo Real", color="#ef5350")
                    ax.set_xticks(xvals)
                    ax.set_xticklabels(lbls, rotation=20, ha="right", fontsize=7)
                    ax.set_ylabel("Custo (€)", fontsize=9)
                    ax.tick_params(axis="y", labelsize=8)
                    ax.legend(fontsize=8)
                    ax.set_title(title_txt, fontsize=11, fontweight="bold")
                elif chart_type == "orlas":
                    lbls = [wrap_label(f"{o.get('ref_orla','')} ({o.get('espessura_orla','')})") for o in dataset]
                    xvals = list(range(len(lbls)))
                    vals = [float(o.get("ml_total", 0) or 0) for o in dataset]
                    bars = ax.bar(xvals, vals, color="#ffa726")
                    ax.set_xticks(xvals)
                    ax.set_xticklabels(lbls, rotation=25, ha="right", fontsize=7)
                    ax.set_ylabel("Metros Lineares (ml)", fontsize=9)
                    ax.tick_params(axis="y", labelsize=8)
                    ax.set_title(title_txt, fontsize=11, fontweight="bold")
                    for rect, val in zip(bars, vals):
                        ax.text(rect.get_x() + rect.get_width() / 2, rect.get_height(), f"{val:.2f}", ha="center", va="bottom", fontsize=7)
                elif chart_type == "ferr":
                    lbls = [wrap_label(f.get("descricao_no_orcamento") or f.get("ref_le") or "", 22) for f in dataset]
                    xvals = list(range(len(lbls)))
                    vals = [float(f.get("custo_mp_total", 0) or 0) for f in dataset]
                    bars = ax.bar(xvals, vals, color="#ef5350")
                    ax.set_xticks(xvals)
                    ax.set_xticklabels(lbls, rotation=25, ha="right", fontsize=7)
                    ax.set_ylabel("Custo (€)", fontsize=9)
                    ax.tick_params(axis="y", labelsize=8)
                    ax.set_title(title_txt, fontsize=11, fontweight="bold")
                    for rect, val in zip(bars, vals):
                        ax.text(rect.get_x() + rect.get_width() / 2, rect.get_height(), f"{val:.2f}", ha="center", va="bottom", fontsize=7)
                elif chart_type == "ops":
                    lbls = [wrap_label(m.get("operacao", ""), 20) for m in dataset]
                    xvals = list(range(len(lbls)))
                    vals = [float(m.get("custo_total", 0) or 0) for m in dataset]
                    bars = ax.bar(xvals, vals, color="#42a5f5")
                    ax.set_xticks(xvals)
                    ax.set_xticklabels(lbls, rotation=20, ha="right", fontsize=7)
                    ax.set_ylabel("Custo (€)", fontsize=9)
                    ax.tick_params(axis="y", labelsize=8)
                    ax.set_title(title_txt, fontsize=11, fontweight="bold")
                    for rect, val in zip(bars, vals):
                        ax.text(rect.get_x() + rect.get_width() / 2, rect.get_height(), f"{val:.2f}", ha="center", va="bottom", fontsize=7)
                elif chart_type == "pie":
                    total_placas = sum(float(p.get("custo_placas_utilizadas", 0) or 0) for p in dataset or [])
                    total_orlas = sum(float(o.get("custo_total", 0) or 0) for o in data.get("orlas", []))
                    total_ferr = sum(float(f.get("custo_mp_total", 0) or 0) for f in data.get("ferr", []))
                    total_maq = sum(float(m.get("custo_total", 0) or 0) for m in data.get("maq", []))
                    valores_local = [total_placas, total_orlas, total_ferr, total_maq]
                    labels_pie = ["Placas", "Orlas", "Ferragens", "Máquinas/MO"]
                    if sum(valores_local) > 0:
                        def _autopct_eur(pct: float) -> str:
                            total = sum(valores_local)
                            val = (pct / 100.0) * total
                            if pct <= 0 or val <= 0:
                                return ""
                            return f"{val:.2f}€\n{pct:.1f}%"

                        wedges, texts, autotexts = ax.pie(
                            valores_local,
                            labels=labels_pie,
                            autopct=_autopct_eur,
                            startangle=90,
                            pctdistance=0.8,
                        )
                        for t in texts + autotexts:
                            t.set_fontsize(7)
                        total_sum = sum(valores_local)
                        table_data = [["Tipo", "€", "%"]]
                        for lbl, val in zip(labels_pie, valores_local):
                            pct = (val / total_sum * 100) if total_sum else 0
                            table_data.append([lbl, f"{val:.2f}", f"{pct:.1f}%"])
                        table = ax.table(cellText=table_data, colWidths=[0.6, 0.5, 0.4], cellLoc="center", bbox=[1.3, 0.1, 0.85, 0.8])
                        table.auto_set_font_size(False)
                        table.set_fontsize(8)
                        for (row, col), cell in table.get_celld().items():
                            if row == 0:
                                cell.set_text_props(weight="bold")
                        ax.set_title(title_txt, fontsize=11, fontweight="bold")
                    else:
                        ax.text(0.5, 0.5, "Sem dados", ha="center", va="center")
            return fig

        fig1 = build_chart_page([
            ("placas", data.get("placas", []), "Comparativo de Custos por Placa"),
            ("orlas", data.get("orlas", []), "Consumo de Orlas (ml)"),
        ])
        fig2 = build_chart_page([
            ("ferr", data.get("ferr", []), "Custos por Ferragem"),
            ("ops", data.get("maq", []), "Custos por Operação"),
        ])

        add_header_footer(fig1, 1, 3)

        # Página de tabelas Placas / Orlas
        def build_table_fig(title: str, headers_rows: List[Tuple[str, List[List[str]]]]) -> Figure:
            """
            Constrói uma figura com 1 coluna e N linhas (cada linha = um 'Resumo ...')
            Ajusta dinamicamente top / hspace / pad do título para evitar sobreposição
            entre o suptitle (título da figura) e os títulos dos subplots.
            """
            fig = Figure(figsize=(11.7, 8.3))

            # --- Parâmetros dinâmicos para evitar sobreposição ---
            n_blocks = len(headers_rows)
            # Se tivermos 2 (ou mais) blocos, precisamos de mais espaço no topo
            if n_blocks >= 2:
                top = 0.75         # área de subplot termina mais em baixo -> mais folga acima
                suptitle_y = 0.92  # suptitle colocado abaixo do header (header está ~0.98)
                hspace = 0.50      # mais espaço entre blocos
                title_pad = 12     # distância do título do subplot ao conteúdo (padrão/ligeramente menor)
                suptitle_fontsize = 13  # ligeiramente menor para ocupar menos espaço vertical
            else:
                # caso 1 bloco; espaços mais modestos
                top = 0.92
                suptitle_y = 0.96
                hspace = 0.25
                title_pad = 14
                suptitle_fontsize = 14

            # Aplica os ajustes de layout
            fig.subplots_adjust(top=top, left=0.03, right=0.97, hspace=hspace)

            # Criar grid para N blocos (1 coluna)
            gs = fig.add_gridspec(n_blocks, 1)

            for idx, (title_sub, rows) in enumerate(headers_rows):
                ax = fig.add_subplot(gs[idx, 0])
                ax.axis("off")
                if not rows:
                    ax.text(0.5, 0.5, "Sem dados", ha="center", va="center")
                    continue

                header = rows[0]
                body = rows[1:]

                # usa width_map definido no scope exterior (mantive a lógica original)
                widths = width_map.get(title_sub)
                if widths:
                    total = sum(widths)
                    col_widths = [w / total for w in widths]
                else:
                    col_widths = None

                # cria a tabela (igual à implementação anterior)
                table = ax.table(cellText=body, colLabels=header, loc="upper center", cellLoc="center", colWidths=col_widths)
                table.auto_set_font_size(False)
                table.set_fontsize(7)
                for key, cell in table.get_celld().items():
                    if key[0] == 0:
                        cell.set_text_props(weight="bold")

                # Define o título do subplot COM UM PAD MAIOR para afastar do conteúdo
                ax.set_title(title_sub, fontsize=10, fontweight="bold", pad=title_pad)

            # Suptitle (título geral) posicionado mais baixo quando há vários blocos
            fig.suptitle(title, fontsize=11, fontweight="bold", y=suptitle_y)

            return fig

        def to_table_rows(data_list: list, cols: list) -> list:
            header = [c[0] for c in cols]
            rows = [header]
            for row in data_list:
                rows.append([(c[2](row.get(c[1])) if len(c) > 2 and callable(c[2]) else row.get(c[1], "")) for c in cols])
            return rows
        width_map = {
            "Resumo de Placas": [0.7, 3.0, 0.7, 0.5, 0.5, 0.8, 0.8, 0.6, 0.6, 0.7, 0.8, 0.8, 1.0, 1.0],
            "Resumo de Orlas": [0.9, 2.5, 0.7, 0.7, 0.8, 0.8],
            "Resumo de Ferragens": [0.9, 2.8, 0.6, 0.5, 0.5, 0.7, 0.7, 0.6, 0.6, 0.7, 0.8, 0.9],
            "Resumo de Máquinas / MO": [1.6, 0.9, 0.9, 0.9, 0.7],
        }

        rows_placas = to_table_rows(data.get("placas", []), placas_cols)
        rows_orlas = to_table_rows(data.get("orlas", []), orlas_cols)
        fig_tables1 = build_table_fig("Tabelas - Placas e Orlas", [("Resumo de Placas", rows_placas), ("Resumo de Orlas", rows_orlas)])

        rows_ferr = to_table_rows(data.get("ferr", []), ferr_cols)
        rows_maq = to_table_rows(data.get("maq", []), maq_cols)
        fig_tables2 = build_table_fig("Tabelas - Ferragens e Máquinas/MO", [("Resumo de Ferragens", rows_ferr), ("Resumo de Máquinas / MO", rows_maq)])

        # Pág. extra: Distribuição de Custos (pizza + tabela) em A4 horizontal
        def build_pie_fig() -> Figure:
            EUR = "\u20ac"
            fig = Figure(figsize=(11.7, 8.3))
            fig.subplots_adjust(top=0.9, bottom=0.08, left=0.05, right=0.95, wspace=0.05)
            gs = fig.add_gridspec(1, 2, width_ratios=[2.2, 1.1])
            ax_pie = fig.add_subplot(gs[0, 0])
            ax_tbl = fig.add_subplot(gs[0, 1])
            ax_tbl.axis("off")

            placas = data.get("placas", []) or []
            orlas = data.get("orlas", []) or []
            ferr = data.get("ferr", []) or []
            maq = data.get("maq", []) or []

            total_placas = sum(float(p.get("custo_placas_utilizadas", 0) or 0) for p in placas)
            total_orlas = sum(float(o.get("custo_total", 0) or 0) for o in orlas)
            total_ferr = sum(float(f.get("custo_mp_total", 0) or 0) for f in ferr)
            total_maq = sum(float(m.get("custo_total", 0) or 0) for m in maq)

            margem_lucro_total = 0.0
            margem_acab_total = 0.0
            margem_mo_total = 0.0
            custos_admin_total = 0.0
            margem_mp_total = 0.0
            base_custo_total = 0.0
            total_venda = 0.0

            try:
                with SessionLocal() as session:
                    itens_orc = (
                        session.query(OrcamentoItem)
                        .filter(
                            OrcamentoItem.id_orcamento == orc.id,
                            OrcamentoItem.versao == orc.versao,
                        )
                        .all()
                    )
                    for it in itens_orc:
                        try:
                            qt = float(getattr(it, "qt", 0) or 0)
                        except Exception:
                            qt = 0.0

                        try:
                            base_custo_total += float(getattr(it, "custo_produzido", 0) or 0) * qt
                        except Exception:
                            pass

                        try:
                            preco_total = float(getattr(it, "preco_total", 0) or 0)
                        except Exception:
                            preco_total = 0.0
                        if preco_total == 0.0 and qt:
                            try:
                                preco_total = float(getattr(it, "preco_unitario", 0) or 0) * qt
                            except Exception:
                                preco_total = 0.0
                        total_venda += preco_total

                        def _acc(attr: str) -> float:
                            try:
                                return float(getattr(it, attr, 0) or 0) * qt
                            except Exception:
                                return 0.0

                        margem_lucro_total += _acc("valor_margem")
                        margem_acab_total += _acc("valor_acabamentos")
                        margem_mo_total += _acc("valor_mao_obra")
                        custos_admin_total += _acc("valor_custos_admin")
                        margem_mp_total += _acc("valor_mp_orlas")
            except Exception:
                pass

            margens_total = (
                margem_lucro_total
                + margem_acab_total
                + margem_mo_total
                + custos_admin_total
                + margem_mp_total
            )

            valores = [total_placas, total_orlas, total_ferr, total_maq]
            labels_pie = ["Placas", "Orlas", "Ferragens", "M\u00e1quinas/MO"]
            if margens_total > 0:
                valores.append(margens_total)
                labels_pie.append("Margens")

            if sum(valores) <= 0:
                ax_pie.text(0.5, 0.5, "Sem dados", ha="center", va="center")
                ax_pie.set_title("Distribui\u00e7\u00e3o de Custos", fontsize=11, fontweight="bold")
                return fig

            total_sum = sum(valores)

            def _autopct_eur(pct: float) -> str:
                val = (pct / 100.0) * total_sum
                if pct <= 0 or val <= 0:
                    return ""
                return f"{val:.2f}{EUR}\n{pct:.1f}%"

            wedges, texts, autotexts = ax_pie.pie(
                valores,
                labels=labels_pie,
                autopct=_autopct_eur,
                startangle=90,
                pctdistance=0.8,
            )
            for t in texts + autotexts:
                t.set_fontsize(8)
            ax_pie.set_title("Distribui\u00e7\u00e3o de Custos", fontsize=11, fontweight="bold")

            def _fmt_eur(val: float) -> str:
                try:
                    return f"{float(val):.2f}{EUR}"
                except Exception:
                    return f"0.00{EUR}"

            def _fmt_pct(val: float, total: float) -> str:
                if not total:
                    return "0.0%"
                return f"{(val / total * 100.0):.1f}%"

            def _fmt_rate(val: float, base: float) -> str:
                if not base:
                    return "0.0%"
                return f"{(val / base * 100.0):.1f}%"

            table_data = [["Tipo", EUR, "%"]]
            for label, val in zip(labels_pie, valores):
                table_data.append([label, _fmt_eur(val), _fmt_pct(val, total_sum)])

            table_data.extend(
                [
                    ["Margem Lucro", _fmt_eur(margem_lucro_total), _fmt_rate(margem_lucro_total, base_custo_total)],
                    ["Margem Acabamentos", _fmt_eur(margem_acab_total), _fmt_rate(margem_acab_total, base_custo_total)],
                    ["Margem M\u00e3o de Obra", _fmt_eur(margem_mo_total), _fmt_rate(margem_mo_total, base_custo_total)],
                    ["Custos Administrativos", _fmt_eur(custos_admin_total), _fmt_rate(custos_admin_total, base_custo_total)],
                    ["Margem Mat\u00e9rias Primas", _fmt_eur(margem_mp_total), _fmt_rate(margem_mp_total, base_custo_total)],
                ]
            )
            if total_venda:
                table_data.append(["Total Venda", _fmt_eur(total_venda), ""])

            table = ax_tbl.table(
                cellText=table_data,
                colWidths=[0.70, 0.40, 0.28],
                cellLoc="center",
                bbox=[0.02, 0.04, 0.96, 0.92],
            )
            table.auto_set_font_size(False)
            table.set_fontsize(8)

            bold_rows = {0}
            if "Margens" in labels_pie:
                bold_rows.add(1 + labels_pie.index("Margens"))
            if total_venda:
                bold_rows.add(len(table_data) - 1)

            for (row, col), cell in table.get_celld().items():
                if row in bold_rows:
                    cell.set_text_props(weight="bold")

            return fig

        fig_pie = build_pie_fig()

        add_header_footer(fig1, 1, 5)
        add_header_footer(fig2, 2, 5)
        add_header_footer(fig_pie, 3, 5)
        add_header_footer(fig_tables1, 4, 5)
        add_header_footer(fig_tables2, 5, 5)

        with PdfPages(dest) as pdf:
            pdf.savefig(fig1)
            pdf.savefig(fig2)
            pdf.savefig(fig_pie)
            pdf.savefig(fig_tables1)
            pdf.savefig(fig_tables2)

        self._show_toast(self.btn_dash_export, f"Dashboard guardado em:\n{dest}", timeout_ms=3000)

    def _open_orcamento_folder(self) -> None:
        if not self.current_orcamento_id:
            self._show_toast(self.btn_open_folder, "Selecione um orçamento antes de abrir a pasta.", timeout_ms=3000)
            return
        if not self._current_orcamento or getattr(self._current_orcamento, "id", None) != self.current_orcamento_id:
            self.refresh_preview()
            if not self._current_orcamento:
                return

        orc = self._current_orcamento
        client = self._current_client
        export_dir = self._find_existing_export_folder(orc, client) or self._determine_export_folder(orc, client)

        target: Optional[Path] = export_dir if export_dir.exists() else None
        if target is None:
            base_dir = export_dir.parent
            if base_dir.exists():
                target = base_dir
        if target is None:
            year_dir = export_dir.parent.parent
            prefix = f"{(orc.num_orcamento or '').strip()}_"
            ver_dir = self._format_versao(getattr(orc, "versao", None))
            alt_ver = str(getattr(orc, "versao", "") or "").strip()
            if prefix != "_" and year_dir.exists():
                try:
                    fallback_base: Optional[Path] = None
                    for entry in year_dir.iterdir():
                        if not (entry.is_dir() and entry.name.startswith(prefix)):
                            continue
                        dir_ver = entry / ver_dir
                        alt_dir_ver = entry / alt_ver if alt_ver else None
                        if dir_ver.is_dir():
                            target = dir_ver
                            break
                        if alt_dir_ver is not None and alt_dir_ver.is_dir():
                            target = alt_dir_ver
                            break
                        fallback_base = fallback_base or entry
                    if target is None and fallback_base is not None and fallback_base.exists():
                        target = fallback_base
                except Exception:
                    pass

        if target is None:
            self._show_toast(
                self.btn_open_folder,
                "A pasta ainda não existe.\n"
                "Crie a pasta no separador 'Orçamentos' (Criar Pasta do Orçamento).\n\n"
                f"Caminho esperado:\n{export_dir}",
                timeout_ms=3000,
            )
            return

        ok = False
        try:
            ok = bool(QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(str(target))))
        except Exception:
            ok = False
        if not ok:
            try:
                detached = QtCore.QProcess.startDetached("explorer.exe", [str(target)])
                ok = bool(detached[0] if isinstance(detached, tuple) else detached)
            except Exception:
                ok = False

        if not ok:
            self._show_toast(self.btn_open_folder, f"Falha ao abrir a pasta:\n{target}", timeout_ms=3000)
            return

        if target != export_dir:
            self._show_toast(self.btn_open_folder, f"A pasta da versão não existe. A abrir:\n{target}", timeout_ms=3000)

    # -- envio de email -------------------------------------------------
    def _on_send_email(self) -> None:
        if not self.current_orcamento_id:
            self._show_toast(self.btn_send_email, "Selecione um orçamento antes de enviar.", timeout_ms=3000)
            return
        if not self._current_orcamento or getattr(self._current_orcamento, "id", None) != self.current_orcamento_id:
            self.refresh_preview()
            if not self._current_orcamento:
                return

        orc = self._current_orcamento
        client = self._current_client

        total = Decimal(0)
        try:
            if getattr(orc, "preco_total", None) is not None:
                total = Decimal(str(orc.preco_total))
        except Exception:
            total = Decimal(0)
        if total <= 0:
            total = sum((it.preco_total or Decimal(0)) for it in self._current_items)

        export_dir = self._find_existing_export_folder(orc, client) or self._determine_export_folder(orc, client)
        pdf_path = export_dir / f"{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}.pdf"
        anexos: list[str] = []
        pdf_filename = ""
        pasta_inicial = str(export_dir)

        if pdf_path.exists():
            anexos = [str(pdf_path)]
            pdf_filename = pdf_path.name
        elif self._current_items:
            # Se houver itens, tentamos gerar o PDF automaticamente (mas não bloqueia o envio se falhar).
            try:
                if not REPORTLAB_AVAILABLE:
                    raise RuntimeError("Biblioteca reportlab não encontrada.")
                _build_pdf_full(self, pdf_path)
                try:
                    self._export_resumo_custos(export_dir, orc, client)
                except Exception:
                    pass
            except Exception as exc:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Email",
                    "Não foi possível gerar/anexar o PDF automaticamente.\n"
                    "Pode continuar e adicionar anexos manualmente.\n\n"
                    f"Detalhe: {exc}",
                )
            else:
                if pdf_path.exists():
                    anexos = [str(pdf_path)]
                    pdf_filename = pdf_path.name
                    pasta_inicial = str(export_dir)

        remetente_email = getattr(self.current_user, "email", None)
        remetente_nome = getattr(self.current_user, "username", None)
        cliente_source = self._cliente_info_source()
        dialog = EmailOrcamentoDialog(
            parent=self,
            destinatario=getattr(cliente_source, "email", "") or "",
            cc=str(remetente_email or ""),
            assunto=self._format_email_subject(orc),
            corpo=self._default_email_body(total, orc, client, pdf_filename=pdf_filename),
            anexos=anexos,
            pasta_inicial=pasta_inicial,
        )
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return
        log_path = get_email_log_path()
        try:
            logger.info(
                "email.send start orcamento_id=%s versao=%s destinatario=%s anexos=%s user_id=%s",
                getattr(orc, "id", None),
                getattr(orc, "versao", None),
                dialog.destinatario(),
                dialog.anexos(),
                getattr(self.current_user, "id", None),
            )
            send_email(
                dialog.destinatario(),
                dialog.assunto(),
                dialog.corpo_html(),
                dialog.anexos(),
                remetente_email=remetente_email,
                remetente_nome=remetente_nome,
                cc=dialog.cc(),
            )
            self._marcar_enviado(orc.id)
        except Exception as exc:
            logger.exception(
                "email.send erro orcamento_id=%s versao=%s destinatario=%s user_id=%s",
                getattr(orc, "id", None),
                getattr(orc, "versao", None),
                dialog.destinatario(),
                getattr(self.current_user, "id", None),
            )
            QtWidgets.QMessageBox.critical(
                self,
                "Email",
                f"Falha ao enviar email: {exc}\nConsulte o log em:\n{log_path}",
            )
            return
        logger.info(
            "email.send ok orcamento_id=%s versao=%s destinatario=%s anexos=%s user_id=%s",
            getattr(orc, "id", None),
            getattr(orc, "versao", None),
            dialog.destinatario(),
            dialog.anexos(),
            getattr(self.current_user, "id", None),
        )
        self._show_toast(self.btn_send_email, "Email enviado com sucesso.", timeout_ms=3000)

    def _marcar_enviado(self, orc_id: int) -> None:
        try:
            with SessionLocal() as session:
                orc = session.get(Orcamento, orc_id)
                if not orc:
                    return
                orc.status = "Enviado"
                session.commit()
        except Exception:
            pass

    def _format_email_subject(self, orc: Orcamento) -> str:
        num = orc.num_orcamento or ""
        ver = self._format_versao(orc.versao)
        ref = getattr(orc, "ref_cliente", "") or ""
        obra = getattr(orc, "obra", "") or ""
        parts = [f"{num}_{ver}"]
        if ref:
            parts.append(f"Ref. Cliente: {ref}")
        if obra:
            parts.append(f"Obra: {obra}")
        return " | ".join(parts)

    def _default_email_body(
        self,
        total: Decimal,
        orc: Orcamento,  # noqa: ARG002
        client: Optional[Client],  # noqa: ARG002
        *,
        pdf_filename: str,
    ) -> str:
        total_val = None
        try:
            if total is not None and float(total) > 0:
                total_val = float(total)
        except Exception:
            total_val = None

        valor_html = ""
        if total_val is not None:
            total_fmt = f"{total_val:.2f}".replace(".", ",")
            valor_html = f"<p style='margin:0 0 12px;'><b>Valor sem IVA:</b> {total_fmt} €</p>"

        pdf_name = html.escape(str(pdf_filename or "").strip())
        pdf_part = f" (<b>{pdf_name}</b>)" if pdf_name else ""
        return (
            "<div style='font-family: Arial, sans-serif; color:#333;'>"
            "<p style='margin:0 0 12px;'>Exmo(a). Sr(a),</p>"
            "<p style='margin:0 0 12px;'>Agradecemos o seu contacto.<br>"
            f"Em anexo segue o orçamento{pdf_part} solicitado.</p>"
            f"{valor_html}"
            "<p style='margin:0 0 16px;'>Se tiver alguma dúvida ou necessitar de mais informação, não hesite em contactar-nos.</p>"
            "<p style='margin:0 0 4px;'>Com os melhores cumprimentos,</p>"
            "<p style='margin:0;'>{{assinatura}}</p>"
            "</div>"
        )

    # wrappers para helpers definidos abaixo (mantem chamadas existentes)
    def _build_workbook(self, output_path: Path):
        return _build_workbook_full(self, output_path)

    def _export_resumo_custos(self, export_dir: Path, orc: Orcamento, client: Optional[Client]) -> None:
        return _export_resumo_custos_full(self, export_dir, orc, client)

    def _export_excel_phc(self, export_dir: Path, orc: Orcamento) -> None:
        return _export_excel_phc_full(self, export_dir, orc)

    def _on_header_resized(self, section: int, _old: int, _new: int) -> None:
        if section == 2:
            QtCore.QTimer.singleShot(0, self._update_row_heights)

    def _update_row_heights(self) -> None:
        vh = self.table.verticalHeader()
        if vh is None:
            return
        model = self.table_model
        row_count = model.rowCount()
        if row_count <= 0:
            vh.setDefaultSectionSize(80)
            return
        width = max(150, self.table.columnWidth(2) - 12)
        doc = QtGui.QTextDocument()
        doc.setDefaultFont(self.table.font())
        for row in range(row_count):
            index = model.index(row, 2)
            html_text = index.data(Qt.DisplayRole) or ""
            doc.setHtml(html_text)
            doc.setTextWidth(width)
            height = doc.size().height() + 24
            vh.resizeSection(row, max(70, int(height)))

    # ---------------------- EXPORT HELPERS ----------------------
    def _get_setting_value(self, key: str, default: str) -> str:
        try:
            with SessionLocal() as session:
                return get_setting(session, key, default) or default
        except Exception:
            return default

    def _determine_export_folder(self, orc: Orcamento, client: Optional[Client]) -> Path:
        base_path = Path(self._get_setting_value(KEY_BASE_PATH, DEFAULT_BASE_PATH))
        ano = str(orc.ano or "")
        cliente_nome = (getattr(client, "nome_simplex", None) or getattr(client, "nome", None) or "CLIENTE")
        simplex = str(cliente_nome).upper().replace(" ", "_")
        pasta = f"{orc.num_orcamento or 'ORC'}_{simplex}"
        versao = self._format_versao(orc.versao)
        if not versao or versao == "-":
            versao = "01"
        return base_path / ano / pasta / versao

    def _find_existing_export_folder(self, orc: Orcamento, client: Optional[Client]) -> Optional[Path]:
        base_path = Path(self._get_setting_value(KEY_BASE_PATH, DEFAULT_BASE_PATH))
        ano = str(orc.ano or "")
        num_orc = str(orc.num_orcamento or "").strip()
        if not (str(base_path).strip() and ano and num_orc):
            return None
        yy_path = base_path / ano
        if not yy_path.is_dir():
            return None

        cliente_nome = (getattr(client, "nome_simplex", None) or getattr(client, "nome", None) or "CLIENTE")
        simplex = str(cliente_nome).upper().replace(" ", "_")
        expected_dir = yy_path / f"{num_orc}_{simplex}"

        ver_dir = self._format_versao(getattr(orc, "versao", None))
        alt_ver = str(getattr(orc, "versao", "") or "").strip()
        ver_candidates = []
        if ver_dir and ver_dir != "-":
            ver_candidates.append(ver_dir)
        if alt_ver and alt_ver not in ver_candidates:
            ver_candidates.append(alt_ver)

        if expected_dir.is_dir():
            for ver in ver_candidates:
                if (expected_dir / ver).is_dir():
                    return expected_dir / ver
            return expected_dir

        prefix = f"{num_orc}_"
        fallback_base = None
        try:
            for entry in yy_path.iterdir():
                if not (entry.is_dir() and entry.name.startswith(prefix)):
                    continue
                base_dir = entry
                for ver in ver_candidates:
                    if (base_dir / ver).is_dir():
                        return base_dir / ver
                if fallback_base is None:
                    fallback_base = base_dir
        except Exception:
            pass

        if fallback_base and fallback_base.is_dir():
            return fallback_base
        return None

    def _require_export_folder(self, orc: Orcamento, client: Optional[Client], *, title: str) -> Optional[Path]:
        export_dir = self._find_existing_export_folder(orc, client) or self._determine_export_folder(orc, client)
        if export_dir is None:
            return
        existing_dir = self._find_existing_export_folder(orc, client)
        if existing_dir and existing_dir.exists():
            return existing_dir
        QtWidgets.QMessageBox.warning(
            self,
            title,
            "A pasta do orcamento ainda nao existe.\n"
            "Crie a pasta no separador \'Orcamentos\' (Criar Pasta do Orcamento).\n\n"
            f"Caminho esperado:\n{export_dir}",
        )
        return None

    def _resolve_logo_path(self) -> Optional[Path]:
        base = self._get_setting_value(KEY_ORC_DB_BASE, DEFAULT_BASE_DADOS_ORC)
        candidate = Path(base) / "LE_Logotipo.png"
        return candidate if candidate.exists() else None

    # ---------------------- EXPORTAÇÃO SIMPLIFICADA ----------------------
    def export_to_excel(self) -> None:
        if not self.current_orcamento_id:
            self._show_toast(self.btn_export_excel, "Selecione um orçamento antes de exportar.", timeout_ms=3000)
            return
        if not self._current_orcamento or not self._current_items:
            self.refresh_preview()
            if not self._current_orcamento or not self._current_items:
                return
        orc = self._current_orcamento
        client = self._current_client
        export_dir = self._require_export_folder(orc, client, title="Exportar Excel")
        if export_dir is None:
            return
        file_path = export_dir / f"{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}.xlsx"
        try:
            self._build_workbook(file_path)
            self._export_resumo_custos(export_dir, orc, client)
            self._export_excel_phc(export_dir, orc)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Exportar Excel", f"Falha ao exportar: {exc}")
            return
        self._show_toast(self.btn_export_excel, f"Relatório guardado em:\n{file_path}", timeout_ms=3000)

    def export_to_pdf(self) -> None:
        if not REPORTLAB_AVAILABLE:
            QtWidgets.QMessageBox.warning(self, "Exportar PDF", "Biblioteca reportlab não encontrada.")
            return
        if not self.current_orcamento_id:
            self._show_toast(self.btn_export_pdf, "Selecione um orçamento antes de exportar.", timeout_ms=3000)
            return
        if not self._current_orcamento or not self._current_items:
            self.refresh_preview()
            if not self._current_orcamento or not self._current_items:
                return
        orc = self._current_orcamento
        client = self._current_client
        export_dir = self._require_export_folder(orc, client, title="Exportar PDF")
        if export_dir is None:
            return
        output_path = export_dir / f"{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}.pdf"
        try:
            _build_pdf_full(self, output_path)
            self._export_resumo_custos(export_dir, orc, client)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Exportar PDF", f"Falha ao exportar: {exc}")
            return
        self._show_toast(self.btn_export_pdf, f"Relatorio guardado em:\n{output_path}", timeout_ms=3000)

        return None


class EmailOrcamentoDialog(QtWidgets.QDialog):
    def __init__(
        self,
        parent=None,
        *,
        destinatario: str = "",
        cc: str = "",
        assunto: str = "",
        corpo: str = "",
        anexos: Optional[List[str]] = None,
        pasta_inicial: Optional[str] = None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Enviar Orçamento por Email")
        self.resize(820, 520)
        self._pasta_inicial = pasta_inicial or ""

        layout = QtWidgets.QVBoxLayout(self)

        form = QtWidgets.QFormLayout()
        self.ed_dest = QtWidgets.QLineEdit(destinatario)
        self.ed_cc = QtWidgets.QLineEdit(cc)
        self.ed_assunto = QtWidgets.QLineEdit(assunto)
        form.addRow("Destinatário:", self.ed_dest)
        form.addRow("CC:", self.ed_cc)
        form.addRow("Assunto:", self.ed_assunto)
        layout.addLayout(form)

        layout.addWidget(QtWidgets.QLabel("Corpo do Email:"))
        self.txt_corpo = QtWidgets.QTextEdit()
        self.txt_corpo.setAcceptRichText(True)
        self.txt_corpo.setHtml(corpo or "")
        layout.addWidget(self.txt_corpo, 1)

        layout.addWidget(QtWidgets.QLabel("Anexos:"))
        self.list_anexos = QtWidgets.QListWidget()
        for path in anexos or []:
            self.list_anexos.addItem(path)
        layout.addWidget(self.list_anexos, 1)

        btns_anexo = QtWidgets.QHBoxLayout()
        self.btn_add = QtWidgets.QPushButton("Adicionar anexo(s)")
        self.btn_remove = QtWidgets.QPushButton("Remover selecionado")
        btns_anexo.addWidget(self.btn_add)
        btns_anexo.addWidget(self.btn_remove)
        layout.addLayout(btns_anexo)

        self.btn_add.clicked.connect(self._on_add_anexo)
        self.btn_remove.clicked.connect(self._on_remove_anexo)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _on_add_anexo(self) -> None:
        start_dir = self._pasta_inicial or QtWidgets.QFileDialog.directory().path()
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(self, "Selecionar anexos", start_dir)
        for f in files:
            if not f:
                continue
            if not any(self.list_anexos.item(i).text() == f for i in range(self.list_anexos.count())):
                self.list_anexos.addItem(f)

    def _on_remove_anexo(self) -> None:
        for item in self.list_anexos.selectedItems():
            row = self.list_anexos.row(item)
            self.list_anexos.takeItem(row)

    def destinatario(self) -> str:
        return self.ed_dest.text().strip()

    def cc(self) -> str:
        return self.ed_cc.text().strip()

    def assunto(self) -> str:
        return self.ed_assunto.text().strip()

    def corpo_html(self) -> str:
        return self.txt_corpo.toHtml()

    def anexos(self) -> List[str]:
        return [self.list_anexos.item(i).text() for i in range(self.list_anexos.count())]


# ----------------------------------------------------------------------
# Funções de exportação completas (restauradas)
def _build_workbook_full(self, output_path: Path) -> None:
    """
    Gera o ficheiro Excel (usando xlsxwriter) com formatação semelhante ao PDF.
    """
    client = self._current_client
    cliente_source = self._cliente_info_source()
    orc = self._current_orcamento
    rows = self._current_items
    if not orc:
        raise ValueError("Nenhum orçamento disponível para exportar.")

    import xlsxwriter
    try:
        from PIL import Image as PILImage
        PIL_AVAILABLE = True
    except Exception:
        PIL_AVAILABLE = False

    wb = xlsxwriter.Workbook(str(output_path))
    ws = wb.add_worksheet("Relatório")

    azul_escuro = '#184ca7'
    fmt_title = wb.add_format({'bold': True, 'font_size': 15, 'align': 'center', 'valign': 'vcenter'})
    fmt_cliente = wb.add_format({'bold': True, 'font_size': 12, 'align': 'left', 'valign': 'vcenter'})
    fmt_info = wb.add_format({'bold': True, 'font_size': 12, 'align': 'right', 'valign': 'vcenter'})
    fmt_small = wb.add_format({'font_size': 10, 'align': 'left', 'valign': 'vcenter'})

    fmt_header = wb.add_format({'bold': True, 'bg_color': '#D9D9D9', 'align': 'center', 'border': 1, 'valign': 'vcenter', 'font_size': 9})
    fmt_cell = wb.add_format({'align': 'center', 'font_size': 10, 'border': 1, 'valign': 'vcenter'})
    fmt_descr_titulo = wb.add_format({'bold': True, 'font_size': 11, 'align': 'left', 'valign': 'top', 'border': 1})
    fmt_descr_sub = wb.add_format({'italic': True, 'font_size': 10, 'align': 'left', 'valign': 'top', 'border': 1})
    fmt_descr_container = wb.add_format({'border': 1, 'text_wrap': True, 'valign': 'top'})
    fmt_number = wb.add_format({'num_format': '0.00', 'align': 'right', 'border': 1, 'valign': 'vcenter'})
    fmt_money = wb.add_format({'num_format': '€ #,##0.00', 'align': 'right', 'valign': 'vcenter', 'border': 1})
    fmt_money_bold = wb.add_format({'num_format': '€ #,##0.00', 'align': 'right', 'bold': True, 'valign': 'vcenter', 'border': 1})
    fmt_row_border = wb.add_format({'border': 1})
    fmt_total_label_nb = wb.add_format({'align': 'right', 'bold': True, 'font_size': 10, 'valign': 'vcenter', 'border': 0})
    fmt_total_val_nb = wb.add_format({'align': 'right', 'bold': True, 'font_size': 11, 'color': '#002060', 'valign': 'vcenter', 'border': 0})
    fmt_money_nb = wb.add_format({'num_format': '€ #,##0.00', 'align': 'right', 'valign': 'vcenter', 'border': 0})
    fmt_date_right = wb.add_format({'font_size': 10, 'align': 'right', 'valign': 'vcenter', 'border': 0})

    ws.set_margins(left=0.2, right=0.2, top=0.2, bottom=0.25)
    ws.fit_to_pages(1, 0)
    ws.set_row(1, 50)

    logo_path = self._resolve_logo_path()
    if logo_path:
        try:
            desired_w_cm = 13.32
            desired_h_cm = 5.61
            if PIL_AVAILABLE:
                pil_img = PILImage.open(str(logo_path))
                img_w_px, img_h_px = pil_img.size
                dpi = 96.0
                desired_w_px = (desired_w_cm / 2.54) * dpi
                desired_h_px = (desired_h_cm / 2.54) * dpi
                x_scale = desired_w_px / img_w_px
                y_scale = desired_h_px / img_h_px
                ws.insert_image(0, 0, str(logo_path), {
                    'x_scale': x_scale, 'y_scale': y_scale, 'x_offset': 8, 'y_offset': 6, 'positioning': 1
                })
            else:
                ws.insert_image(0, 0, str(logo_path), {
                    'x_scale': 0.18, 'y_scale': 0.18, 'x_offset': 8, 'y_offset': 6, 'positioning': 1
                })
        except Exception:
            pass

    cliente_nome = (self._cliente_display_nome() or "").upper()
    ws.merge_range(2, 0, 2, 2, cliente_nome, fmt_cliente)
    ws.merge_range(0, 3, 0, 9, "Relatório de Orçamento", fmt_title)
    ws.merge_range(1, 3, 1, 9, f"Nº Orçamento: {orc.num_orcamento or ''}_{self._format_versao(orc.versao)}", fmt_info)
    ws.merge_range(2, 3, 2, 9, f"Data: {orc.data or ''}", fmt_date_right)

    ws.merge_range(3, 0, 3, 2, getattr(cliente_source, "morada", "") or "", fmt_small)
    ws.merge_range(4, 0, 4, 2, getattr(cliente_source, "email", "") or "", fmt_small)
    contactos = (
        f"Telefone: {(getattr(cliente_source, 'telefone', '') or '')}  |  "
        f"Telemóvel: {(getattr(cliente_source, 'telemovel', '') or '')}"
    )
    ws.merge_range(5, 0, 5, 2, contactos, fmt_small)
    ws.merge_range(6, 0, 6, 2, f"Ref.: {orc.ref_cliente or '-'}", wb.add_format({'bold': True, 'font_size': 12, 'font_color': azul_escuro, 'border': 0}))
    ws.merge_range(6, 3, 6, 9, f"Obra: {orc.obra or '-'}", wb.add_format({'bold': True, 'font_size': 12, 'font_color': 'red', 'align': 'right', 'border': 0}))

    headers = ["Item", "Código", "Descrição", "Alt", "Larg", "Prof", "Und", "Qt", "Preço Unit", "Preço Total"]
    start_row = 9
    for col, h in enumerate(headers):
        ws.write(start_row, col, h, fmt_header)
    col_widths = [6, 12, 44, 8, 8, 8, 6, 6, 14, 16]
    for idx, w in enumerate(col_widths):
        ws.set_column(idx, idx, w)

    row = start_row + 1
    for item in rows:
        entries = self._parse_description(item.descricao)
        text_lines = []
        if entries:
            text_lines.append(entries[0][1].upper())
            for kind, text in entries[1:]:
                if kind == 'dash':
                    text_lines.append(f"- {text}")
                elif kind == 'star':
                    text_lines.append(f"* {text}")
                elif kind == 'header2':
                    text_lines.append(text.upper())
                else:
                    text_lines.append(text)
        else:
            text_lines = [""]
        num_lines = max(1, len(text_lines))
        line_height = max(20, 14 * num_lines)
        ws.set_row(row, line_height)

        ws.write(row, 0, item.item, fmt_cell)
        ws.write(row, 1, item.codigo, fmt_cell)

        if not entries:
            ws.write(row, 2, "", fmt_descr_container)
        else:
            if len(entries) == 1:
                ws.write(row, 2, entries[0][1].upper(), fmt_descr_titulo)
            else:
                rich_args = [fmt_descr_titulo, entries[0][1].upper()]
                for kind, text in entries[1:]:
                    rich_args.append('\n')
                    if kind == 'dash':
                        rich_args.extend([fmt_descr_sub, f"- {text}"])
                    elif kind == 'star':
                        rich_args.extend([fmt_descr_sub, f"* {text}"])
                    elif kind == 'header2':
                        rich_args.extend([fmt_descr_titulo, str(text).upper()])
                    else:
                        rich_args.extend([fmt_descr_sub, text])
                try:
                    ws.write_rich_string(row, 2, *rich_args, fmt_descr_container)
                except Exception:
                    joined = "\n".join(
                        [entries[0][1].upper()]
                        + [("- " + t) if k == 'dash' else ("* " + t) if k == 'star' else t for k, t in entries[1:]]
                    )
                    ws.write(row, 2, joined, fmt_descr_container)

        ws.write_number(row, 3, float(item.altura) if item.altura is not None else None, fmt_cell)
        ws.write_number(row, 4, float(item.largura) if item.largura is not None else None, fmt_cell)
        ws.write_number(row, 5, float(item.profundidade) if item.profundidade is not None else None, fmt_cell)
        ws.write(row, 6, item.unidade, fmt_cell)
        ws.write_number(row, 7, float(item.qt) if item.qt is not None else 0.0, fmt_number)
        ws.write_number(row, 8, float(item.preco_unitario) if item.preco_unitario is not None else 0.0, fmt_money)
        ws.write_number(row, 9, float(item.preco_total) if item.preco_total is not None else 0.0, fmt_money_bold)
        row += 1

    totals_row = row + 1
    total_qt = sum((item.qt or Decimal("0")) for item in rows if not _is_separator_item(item))
    subtotal = sum((item.preco_total or Decimal("0")) for item in rows)
    iva = subtotal * IVA_RATE
    total = subtotal + iva

    ws.write(totals_row, 8, "Total Qt:", fmt_total_label_nb)
    ws.write_number(totals_row, 9, float(total_qt), fmt_total_val_nb)
    ws.write(totals_row + 1, 8, "SubTotal:", fmt_total_label_nb)
    ws.write_number(totals_row + 1, 9, float(subtotal), fmt_total_val_nb)
    ws.write(totals_row + 2, 8, "IVA (23%):", fmt_total_label_nb)
    ws.write_number(totals_row + 2, 9, float(iva), fmt_total_val_nb)
    ws.write(totals_row + 3, 8, "Total Geral:", fmt_total_label_nb)
    ws.write_number(totals_row + 3, 9, float(total), fmt_money_nb)
    wb.close()


def _export_resumo_custos_full(self, export_dir: Path, orc: Orcamento, client: Optional[Client]) -> None:
    """
    Gera 'Resumo_Custos_<num>_<ver>.xlsx' a partir do modelo MODELO_Resumo_Custos.xlsx
    preenchendo o separador 'Resumo Geral' com os registos de custeio_items.
    """
    col_order = [
        "id",
        "descricao_livre",
        "def_peca",
        "descricao",
        "qt_total",
        "comp_res",
        "larg_res",
        "esp_res",
        "mps",
        "mo",
        "orla",
        "blk",
        "nst",
        "mat_default",
        "acabamento_sup",
        "acabamento_inf",
        "ref_le",
        "descricao_no_orcamento",
        "pliq",
        "und",
        "desp",
        "orl_0_4",
        "orl_1_0",
        "tipo",
        "familia",
        "comp_mp",
        "larg_mp",
        "esp_mp",
        "orl_c1",
        "orl_c2",
        "orl_l1",
        "orl_l2",
        "ml_orl_c1",
        "ml_orl_c2",
        "ml_orl_l1",
        "ml_orl_l2",
        "custo_orl_c1",
        "custo_orl_c2",
        "custo_orl_l1",
        "custo_orl_l2",
        "gravar_modulo",
        "custo_total_orla",
        "soma_total_ml_orla",
        "area_m2_und",
        "perimetro_und",
        "spp_ml_und",
        "cp01_sec",
        "cp01_sec_und",
        "cp02_orl",
        "cp02_orl_und",
        "cp03_cnc",
        "cp03_cnc_und",
        "cp04_abd",
        "cp04_abd_und",
        "cp05_prensa",
        "cp05_prensa_und",
        "cp06_esquad",
        "cp06_esquad_und",
        "cp07_embalagem",
        "cp07_embalagem_und",
        "cp08_mao_de_obra",
        "cp08_mao_de_obra_und",
        "cp09_colagem",
        "cp09_colagem_und",
        "custo_mp_und",
        "custo_mp_total",
        "soma_custo_orla_total",
        "soma_custo_und",
        "soma_custo_total",
        "soma_custo_acb",
    ]

    base_dados_path = Path(self._get_setting_value(KEY_ORC_DB_BASE, DEFAULT_BASE_DADOS_ORC))
    template_candidates = [
        base_dados_path / "MODELO_Resumo_Custos.xlsx",
        base_dados_path / "MODELO_Resumo_Custos_V2.xlsx",
    ]
    template_path = next((p for p in template_candidates if p.exists()), None)
    if not template_path:
        QtWidgets.QMessageBox.warning(self, "Resumo Custos", "Modelo não encontrado (MODELO_Resumo_Custos.xlsx).")
        return

    nome_base = f"Resumo_Custos_{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}.xlsx"
    destino = export_dir / nome_base
    try:
        shutil.copyfile(template_path, destino)
    except Exception as exc:
        QtWidgets.QMessageBox.warning(self, "Resumo Custos", f"Falha ao copiar modelo: {exc}")
        return

    try:
        wb = load_workbook(destino)
    except Exception as exc:
        QtWidgets.QMessageBox.warning(self, "Resumo Custos", f"Não foi possível abrir o modelo: {exc}")
        return

    ws = wb["Resumo Geral"] if "Resumo Geral" in wb.sheetnames else wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for col_idx, name in enumerate(col_order, start=1):
        ws.cell(row=1, column=col_idx, value=name)
    if ws.max_column > len(col_order):
        ws.delete_cols(len(col_order) + 1, ws.max_column - len(col_order))

    try:
        with SessionLocal() as session:
            cust_rows = (
                session.query(CusteioItem)
                .filter(CusteioItem.orcamento_id == orc.id, CusteioItem.versao == orc.versao)
                .order_by(CusteioItem.ordem)
                .all()
            )
            
            # ✅ Carregar também os OrcamentoItem para ter acesso à quantidade real de cada item
            orc_items = (
                session.query(OrcamentoItem)
                .filter(OrcamentoItem.id_orcamento == orc.id)
                .all()
            )
            # Criar mapa: item_id -> OrcamentoItem.qt
            orc_item_qt_map = {item.id_item: float(item.qt or 1) for item in orc_items}
    except Exception as exc:
        QtWidgets.QMessageBox.warning(self, "Resumo Custos", f"Falha ao carregar custeio: {exc}")
        return

    def _val(obj, attr):
        v = getattr(obj, attr, None)
        if isinstance(v, Decimal):
            try:
                return float(v)
            except Exception:
                return None
        if isinstance(v, bool):
            return 1 if v else 0
        return v

    # ✅ Lista de campos que precisam ser multiplicados pela quantidade do item
    # Estes são valores que estão por unidade no custeio e precisam ser ajustados
    FIELDS_TO_MULTIPLY = {
        "qt_total",  # Quantidade local do custeio deve refletir quantidade total
        "ml_orl_c1", "ml_orl_c2", "ml_orl_l1", "ml_orl_l2",  # ML de orlas por unidade
        "custo_orl_c1", "custo_orl_c2", "custo_orl_l1", "custo_orl_l2",  # Custos de orlas por unidade
        "custo_total_orla",  # Custo total de orlas por unidade
        "custo_mp_und",  # Custo MP por unidade (não multiplicar custo_mp_total pois pode já estar ajustado)
        "spp_ml_und",  # ML de borda por unidade
        "cp01_sec_und", "cp02_orl_und", "cp03_cnc_und", "cp04_abd_und",  # Custos de operações por unidade
        "cp05_prensa_und", "cp06_esquad_und", "cp07_embalagem_und", "cp08_mao_de_obra_und",
        "cp09_colagem_und",
        "soma_custo_und",  # Soma de custos por unidade
    }

    row_idx = 2
    for c_item in cust_rows:
        # ✅ Obter a quantidade real do item de orçamento
        item_qt = orc_item_qt_map.get(getattr(c_item, "item_id", None), 1.0)
        
        for col_idx, attr in enumerate(col_order, start=1):
            value = _val(c_item, attr)
            
            # ✅ Multiplicar por quantidade do item se for um campo que precisa
            if attr in FIELDS_TO_MULTIPLY and value is not None and isinstance(value, (int, float)):
                value = value * item_qt
            
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1

    def _orla_width_from_esp(esp_val: float) -> float:
        try:
            esp = float(esp_val)
        except Exception:
            return 0.0
        if esp <= 0:
            return 0.0
        if esp < 20:
            return 23.0
        if esp < 31:
            return 35.0
        if esp < 40:
            return 45.0
        return 60.0

    ws_resumo_orlas = wb.create_sheet("Resumo_Orlas")
    ws_resumo_orlas.append(["Ref Orla", "Descr Material", "Espessura", "Largura", "ML Total", "Custo Total"])
    for c_item in cust_rows:
        # ✅ Obter a quantidade real do item de orçamento
        item_qt = orc_item_qt_map.get(getattr(c_item, "item_id", None), 1.0)
        
        largura_orla = _orla_width_from_esp(getattr(c_item, "esp_res", None) or getattr(c_item, "esp_mp", None))
        ref_map = {
            1: getattr(c_item, "orl_0_4", None) or getattr(c_item, "corres_orla_0_4", None) or "",
            2: getattr(c_item, "orl_1_0", None) or getattr(c_item, "corres_orla_1_0", None) or "",
        }
        sides = [
            (getattr(c_item, "orl_c1", 0), getattr(c_item, "ml_orl_c1", 0), getattr(c_item, "custo_orl_c1", 0)),
            (getattr(c_item, "orl_c2", 0), getattr(c_item, "ml_orl_c2", 0), getattr(c_item, "custo_orl_c2", 0)),
            (getattr(c_item, "orl_l1", 0), getattr(c_item, "ml_orl_l1", 0), getattr(c_item, "custo_orl_l1", 0)),
            (getattr(c_item, "orl_l2", 0), getattr(c_item, "ml_orl_l2", 0), getattr(c_item, "custo_orl_l2", 0)),
        ]
        for code_raw, ml_raw, custo_raw in sides:
            try:
                code_val = float(code_raw or 0)
            except Exception:
                code_val = 0.0
            if code_val <= 0:
                continue
            esp_descr = "1.0mm" if code_val >= 0.9 else "0.4mm"
            ref_orl = ref_map.get(2 if code_val >= 0.9 else 1) or ""
            if not ref_orl:
                continue
            ml_val = float(ml_raw or 0)
            custo_val = float(custo_raw or 0)
            
            # ✅ Multiplicar ML e custo pela quantidade do item
            ml_val *= item_qt
            custo_val *= item_qt
            
            if ml_val == 0 and custo_val == 0:
                continue
            ws_resumo_orlas.append([ref_orl, c_item.descricao_no_orcamento, esp_descr, largura_orla, ml_val, custo_val])

    try:
        wb.save(destino)
    except Exception as exc:
        QtWidgets.QMessageBox.warning(self, "Resumo Custos", f"Falha ao guardar: {exc}")
        return

# ---------------------- PLANO DE CORTE (RESUMO) ----------------------

def generate_cut_plan_pdf(resumo_path: Path, output_pdf: Path, footer_label: str = '', kerf_mm: float = 3.0) -> None:
    # Plano de corte com resumo e layouts (usa rectpack se disponivel)
    try:
        from rectpack import newPacker, PackingMode, MaxRectsBssf
        rectpack_ok = True
    except Exception:
        rectpack_ok = False
    df = pd.read_excel(resumo_path, sheet_name='Resumo Geral', engine='openpyxl')
    col_tipo = 'tipo' if 'tipo' in df.columns else None
    col_familia = 'familia' if 'familia' in df.columns else None
    mask = None
    if col_tipo:
        mask = df[col_tipo].astype(str).str.upper().str.contains('PLACA', na=False)
    if (mask is None or not mask.any()) and col_familia:
        mask = df[col_familia].astype(str).str.upper().str.contains('PLACA', na=False)
    if mask is None or not mask.any():
        raise ValueError('Nenhum registo de placas encontrado (tipo/familia).')
    df = df[mask]
    if df.empty:
        raise ValueError('Nenhuma linha de PLACAS encontrada no Resumo Geral.')

    required = ['def_peca', 'qt_total', 'comp_res', 'larg_res', 'descricao_no_orcamento', 'comp_mp', 'larg_mp', 'esp_mp']
    for col in required:
        if col not in df.columns:
            raise ValueError(f'Coluna obrigatoria em falta: {col}')

    df['qt_total'] = pd.to_numeric(df['qt_total'], errors='coerce').fillna(0).astype(int)
    for col in ['comp_res', 'larg_res', 'comp_mp', 'larg_mp', 'esp_mp']:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    def has_grain(texto: str) -> bool:
        s = (texto or '').lower()
        if re.search(r'\b[h|m]\d{4}\b', s, re.IGNORECASE):
            return True
        woods = ['carvalho', 'nogueira', 'maple', 'roble', 'oak', 'walnut', 'freijo', 'pinus']
        return any(w in s for w in woods)

    summary_rows = []
    board_sets = []
    grouped = df.groupby(['descricao_no_orcamento', 'esp_mp', 'comp_mp', 'larg_mp'])
    for (ref, esp, board_w, board_h), group in grouped:
        if pd.isna(board_w) or pd.isna(board_h) or board_w <= 0 or board_h <= 0:
            continue
        pieces = []
        idx_counter = 0
        for _, row in group.iterrows():
            qty = int(row.get('qt_total', 0) or 0)
            w = float(row.get('comp_res', 0) or 0)
            h = float(row.get('larg_res', 0) or 0)
            if qty <= 0 or not math.isfinite(w) or not math.isfinite(h) or w <= 0 or h <= 0:
                continue
            desc = str(row.get('def_peca', ''))
            rot_allowed = not has_grain(str(ref) or desc)
            for _ in range(qty):
                pieces.append({'id': idx_counter, 'desc': desc, 'w': w, 'h': h, 'rot': rot_allowed})
                idx_counter += 1
        if not pieces:
            continue

        boards = []
        not_placed = []
        if rectpack_ok:
            packer = newPacker(pack_algo=MaxRectsBssf, mode=PackingMode.Offline, rotation=True)
            packer.add_bin(board_w, board_h, float('inf'))
            for pce in pieces:
                packer.add_rect(pce['w'] + kerf_mm, pce['h'] + kerf_mm, pce['id'])
            packer.pack()

            boards_map = {}
            for bid, rx, ry, rw, rh, rid in packer.rect_list():
                piece = next((pp for pp in pieces if pp['id'] == rid), None)
                desc = piece['desc'] if piece else ''
                w_use = float(max(rw - kerf_mm, 0))
                h_use = float(max(rh - kerf_mm, 0))
                boards_map.setdefault(bid, {'w': board_w, 'h': board_h, 'items': []})
                boards_map[bid]['items'].append({
                    'x': float(rx),
                    'y': float(ry),
                    'w': w_use if piece else float(rw),
                    'h': h_use if piece else float(rh),
                    'desc': desc,
                })
            boards = [boards_map[k] for k in sorted(boards_map.keys())]
            if not boards:
                not_placed = pieces
        else:
            for pce in pieces:
                boards.append({'w': board_w, 'h': board_h, 'items': [{'x': 0, 'y': 0, 'w': pce['w'], 'h': pce['h'], 'desc': pce['desc']}]})

        area_pieces = sum(pce['w'] * pce['h'] for pce in pieces)
        area_boards = len(boards) * board_w * board_h if boards else 0
        summary_rows.append({
            'referencia': ref,
            'esp_mm': esp,
            'dim_placa': f"{int(board_w)}x{int(board_h)}",
            'placas': len(boards),
            'aproveitamento_pct': round((area_pieces / area_boards * 100) if area_boards else 0, 2),
            'nao_alocadas': len(not_placed),
            'area_pecas_m2': round(area_pieces / 1_000_000, 3),
            'area_placas_m2': round(area_boards / 1_000_000, 3),
        })
        board_sets.append((ref, esp, board_w, board_h, boards, not_placed))

    from reportlab.lib.pagesizes import landscape
    c = NumberedFooterCanvas(str(output_pdf), pagesize=landscape(A4), footer_info={'numero': footer_label, 'data': ''})
    page_w, page_h = landscape(A4)
    c.setFont('Helvetica-Bold', 14)
    c.drawString(20, page_h - 30, 'Resumo Plano de Corte')
    table_data = [["Ref", "Esp (mm)", "Dimensao (mm)", "Placas", "Aproveitamento %", "Nao alocadas", "Area pecas (m2)", "Area placas (m2)"]]
    for row in summary_rows:
        table_data.append([
            row['referencia'], row['esp_mm'], row['dim_placa'], row['placas'], row['aproveitamento_pct'],
            row['nao_alocadas'], row.get('area_pecas_m2', ''), row.get('area_placas_m2', ''),
        ])
    try:
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ALIGN', (3,1), (7,-1), 'RIGHT'),
        ]))
        tw, th = table.wrapOn(c, page_w - 40, page_h - 80)
        table.drawOn(c, 20, page_h - 60 - th)
    except Exception:
        y = page_h - 50
        c.setFont('Helvetica', 9)
        for r in summary_rows:
            line = (f"Ref: {r['referencia']} | Esp: {r['esp_mm']} | Placas: {r['placas']} | "
                    f"Aproveitamento: {r['aproveitamento_pct']}% | Dim: {r['dim_placa']}")
            c.drawString(20, y, line)
            y -= 12
            if y < 40:
                c.showPage(); y = page_h - 40
    c.showPage()

    colors_palette = [colors.lightblue, colors.lightgreen, colors.lightpink, colors.lightgoldenrodyellow, colors.lavender, colors.peachpuff]
    for ref, esp, board_w, board_h, boards, not_placed in board_sets:
        if not boards:
            continue
        total_boards = len(boards)
        for idx, b in enumerate(boards, start=1):
            c.setPageSize(landscape(A4))
            page_w, page_h = landscape(A4)
            margin = 15 * mm
            scale = min((page_w - 2 * margin) / board_w, (page_h - 2 * margin) / board_h)
            ox = (page_w - board_w * scale) / 2
            oy = (page_h - board_h * scale) / 2
            c.setFont('Helvetica-Bold', 12)
            c.drawString(margin, page_h - margin + 5, f"{ref} | Esp: {esp} | Placa #{idx} de {total_boards}")
            c.rect(ox, oy, board_w * scale, board_h * scale, stroke=1, fill=0)

            # Legenda eixos da placa (XX/YY) baseada na dimensão da referência (ex.: 2440x2100)
            try:
                comp_mm = int(round(float(board_w)))
            except Exception:
                comp_mm = int(board_w) if board_w else 0
            try:
                larg_mm = int(round(float(board_h)))
            except Exception:
                larg_mm = int(board_h) if board_h else 0

            c.setFillColor(colors.grey)
            c.setFont('Helvetica', 8)
            # XX (horizontal) - abaixo da placa
            xx_y = max(oy - 6 * mm, 9 * mm)
            c.drawCentredString(ox + (board_w * scale) / 2, xx_y, f"XX (Comp): {comp_mm} mm")
            # YY (vertical) - à esquerda da placa, texto rodado
            yy_x = max(ox - 7 * mm, 9 * mm)
            c.saveState()
            c.translate(yy_x, oy + (board_h * scale) / 2)
            c.rotate(90)
            c.drawCentredString(0, 0, f"YY (Larg): {larg_mm} mm")
            c.restoreState()
            c.setFillColor(colors.black)

            for j, piece in enumerate(b['items']):
                col = colors_palette[j % len(colors_palette)]
                px = ox + piece['x'] * scale
                py = oy + piece['y'] * scale
                max_w = max(board_w - piece['x'], 0)
                max_h = max(board_h - piece['y'], 0)
                pw = min(piece['w'], max_w) * scale
                ph = min(piece['h'], max_h) * scale
                if pw <= 0 or ph <= 0:
                    continue
                c.setFillColor(col)
                c.rect(px, py, pw, ph, stroke=1, fill=1)
                c.setFillColor(colors.black)
                c.setFont('Helvetica', 7)
                c.drawString(px + 2, py + ph - 8, f"{piece['desc']} ({int(piece['w']-kerf_mm)}x{int(piece['h']-kerf_mm)})"[:70])
            c.showPage()
        if not_placed:
            c.setFont('Helvetica-Bold', 12)
            c.drawString(20, page_h - 30, f"Pecas nao alocadas - {ref}")
            c.setFont('Helvetica', 9)
            y = page_h - 50
            for pnt in not_placed:
                c.drawString(20, y, f"{pnt['desc']} ({pnt['w']}x{pnt['h']})")
                y -= 12
                if y < 40:
                    c.showPage(); y = page_h - 40
            c.showPage()
    c.save()


# Salvaguarda: adiciona métodos ausentes para evitar que a página quebre
def _ensure_relatorios_methods():
    cls = RelatoriosPage
    # Já existem implementações reais; apenas mantemos stub se ausente.
    if not hasattr(cls, "export_to_excel"):
        def export_to_excel(self) -> None:
            QtWidgets.QMessageBox.information(self, "Exportar Excel", "Funcionalidade indisponível.")
        cls.export_to_excel = export_to_excel  # type: ignore[attr-defined]

    if not hasattr(cls, "export_to_pdf"):
        def export_to_pdf(self) -> None:
            QtWidgets.QMessageBox.information(self, "Exportar PDF", "Funcionalidade indisponível.")
        cls.export_to_pdf = export_to_pdf  # type: ignore[attr-defined]


_ensure_relatorios_methods()

# -----------------------------------------------------------------------------
# Plano de corte: método externo e atribuição à classe (para evitar mexer no corpo)
# -----------------------------------------------------------------------------
def _export_cut_plan_pdf_impl(self) -> None:
    """Exporta plano de corte (resumo) em PDF a partir do Resumo_Custos."""
    if not REPORTLAB_AVAILABLE:
        QtWidgets.QMessageBox.warning(self, "Plano de Corte", "Biblioteca reportlab não encontrada.")
        return
    if not self.current_orcamento_id:
        try:
            self._show_toast(getattr(self, "btn_cut_plan", None), "Selecione um orçamento antes de exportar.", timeout_ms=3000)
        except Exception:
            QtWidgets.QMessageBox.information(self, "Plano de Corte", "Selecione um orçamento antes de exportar.")
        return
    if not self._current_orcamento or not self._current_items:
        self.refresh_preview()
        if not self._current_orcamento or not self._current_items:
            return
    orc = self._current_orcamento
    client = self._current_client
    export_dir = self._require_export_folder(orc, client, title="Plano de Corte")
    if export_dir is None:
        return

    resumo_path = export_dir / f"Resumo_Custos_{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}.xlsx"
    if not resumo_path.exists():
        try:
            self._export_resumo_custos(export_dir, orc, client)
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "Plano de Corte", f"Falha ao gerar Resumo_Custos: {exc}")
    if not resumo_path.exists():
        QtWidgets.QMessageBox.warning(self, "Plano de Corte", f"Resumo_Custos não encontrado:\n{resumo_path}")
        return

    output_pdf = export_dir / f"Plano_Corte_{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}.pdf"
    try:
        generate_cut_plan_pdf(
            resumo_path,
            output_pdf,
            footer_label=f"{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}",
        )
    except Exception as exc:
        logging.getLogger(__name__).exception("Erro ao gerar plano de corte")
        details = str(exc).strip() or type(exc).__name__
        QtWidgets.QMessageBox.critical(
            self,
            "Plano de Corte",
            f"Erro ao gerar plano de corte:\n{details}\n\nConsulte o ficheiro martelo_debug.log para mais detalhes.",
        )
        return
    try:
        self._show_toast(getattr(self, "btn_cut_plan", None), f"Plano de corte guardado em:\n{output_pdf}", timeout_ms=3000)
    except Exception:
        QtWidgets.QMessageBox.information(self, "Plano de Corte", f"Plano de corte guardado em:\n{output_pdf}")

# Atribuir método à classe (se não existir)
setattr(RelatoriosPage, "export_cut_plan_pdf", _export_cut_plan_pdf_impl)

def _export_excel_phc_full(self, export_dir: Path, orc: Orcamento) -> Path:
    """
    Gera um ficheiro Excel no formato esperado pelo PHC.
    """
    filename = f"{orc.num_orcamento or 'orcamento'}_{self._format_versao(orc.versao)}_PHC.xlsx"
    dest = export_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "PHC"
    headers = ["RefCliente", "Referencia", "Designacao", "XAltura", "YLargura", "ZEspessura", "Qtd", "Und", "Venda"]
    ws.append(headers)

    def _to_num(value):
        conv = getattr(self, "_decimal_value", None)
        if callable(conv):
            return conv(value)
        try:
            return float(value)
        except Exception:
            return None

    for item in self._current_items:
        desc_entries = self._parse_description(item.descricao)
        lines: list[str] = []
        if desc_entries:
            lines.append(desc_entries[0][1].upper())
            for kind, text in desc_entries[1:]:
                if kind == "dash":
                    lines.append(f"- {text}")
                elif kind == "star":
                    lines.append(f"* {text}")
                elif kind == "header2":
                    lines.append(text.upper())
                else:
                    lines.append(text)
        else:
            lines.append("")

        prefix = "COMP. MOB. - "
        first_line = f"{prefix}{lines[0]}" if lines[0] else prefix.rstrip()
        extra_lines = lines[1:]

        und_val = (getattr(item, "unidade", "") or "").strip() or "un"
        if und_val.lower() == "und":
            und_val = "un"

        ws.append(
            [
                item.codigo,
                "MOB",
                first_line,
                _to_num(item.altura),
                _to_num(item.largura),
                _to_num(item.profundidade),
                _to_num(item.qt),
                und_val,
                _to_num(item.preco_unitario),
            ]
        )
        for extra in extra_lines:
            ws.append(["", "", extra, None, None, None, None, None, None])

    # ajuste simples de largura das colunas para facilitar leitura
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 60)

    wb.save(dest)
    return dest

def _build_pdf_full(self, output_path: Path) -> None:
    """
    Gera o PDF do orçamento com layout ajustado:
    - margens em mm coerentes com as colWidths
    - colunas redesenhadas (descrição maior)
    - fontes e paddings ajustados para caber melhor na tabela
    - totais alinhados mais à direita
    """
    client = self._current_client
    cliente_source = self._cliente_info_source()
    orc = self._current_orcamento
    rows = self._current_items
    if not orc:
        raise ValueError("Nenhum orçamento disponível.")

    # --------------------------
    # Parâmetros de layout (fáceis de ajustar)
    # --------------------------
    left_margin = 5 * mm        # margem esquerda (mm -> pontos)
    right_margin = 5 * mm       # margem direita
    top_margin = 1 * mm
    bottom_margin = 1 * mm # margem inferior pequena para rodapé mais baixo

    # Tamanhos de fontes e leading
    HEADER_FONT_SIZE = 16
    HEADER_TABLE_FONTSIZE = 9   # <--- reduzido (cabeçalho da tabela)
    BODY_FONTSIZE = 9
    DESC_LEADING = 11

    # col_widths em mm que somam a largura útil da página:
    # A4 = 210 mm, largura útil = 210 - left - right
    # 210 - 5 - 5 = 200 mm
    # ajuste: descrição maior (80 mm) para caber melhor
    col_widths_mm = [9, 20, 80, 10, 11, 11, 11, 11, 18, 19]  # soma = 200 mm
    # converter para pontos (reportlab usa pontos internamente)
    col_widths = [w * mm for w in col_widths_mm]

    # --------------------------
    # Styles
    # --------------------------
    styles = getSampleStyleSheet()
    # info_style mais compacto para poupar espaço vertical
    info_style = ParagraphStyle("info", parent=styles["Normal"], fontSize=9, leading=11, spaceAfter=2)
    header_style = ParagraphStyle("header", parent=styles["Heading1"], alignment=2, fontSize=HEADER_FONT_SIZE, textColor=colors.HexColor("#133a63"))
    desc_style = ParagraphStyle("desc", parent=styles["Normal"], fontSize=BODY_FONTSIZE, leading=DESC_LEADING, spaceAfter=2)
    price_style = ParagraphStyle("price", parent=styles["Normal"], alignment=2, fontSize=BODY_FONTSIZE, leading=DESC_LEADING)

    # -------------------------------------------------------------------------------------
    # Criar o doc (precisamos do doc para obter doc.width/doc.height antes de construir story)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
        leftMargin=left_margin,
        rightMargin=right_margin,
    )

    # ------------------------------------------------------------------
    # Header: left_table (cliente) e right_table (titulo + nº/data)
    # ------------------------------------------------------------------
    left_elements: List = []
    logo_path = self._resolve_logo_path()
    if logo_path:
        try:
            img = Image(str(logo_path))
            img.drawHeight = 10 * mm
            img.drawWidth = 32 * mm
            left_elements.append([img])
        except Exception:
            pass

    client_name = self._cliente_display_nome() or ""
    contact_lines = [
        getattr(cliente_source, "morada", "") or "",
        getattr(cliente_source, "email", "") or "",
        (
            f"Telefone: {getattr(cliente_source, 'telefone', '') or ''} | "
            f"Telemóvel: {getattr(cliente_source, 'telemovel', '') or ''} | "
            f"N.º cliente PHC: {getattr(cliente_source, 'num_cliente_phc', '') or ''}"
        ),
    ]
    left_elements.append([Paragraph(f"<font size=12><b>{client_name}</b></font>", info_style)])
    left_elements.append([Paragraph("<br/>".join(filter(None, contact_lines)), info_style)])
    left_table = Table(left_elements, colWidths=[80 * mm])
    left_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), -8 * mm),
    ]))

    right_title = Paragraph("<font color='#103864' size=16><b>Relatório de Orçamento</b></font>", header_style)
    right_info = Paragraph(
        f"<font size=12>Nº Orçamento: <b>{orc.num_orcamento or ''}_{self._format_versao(orc.versao)}</b></font><br/>"
        f"<font color='#5f6368'>Data: {orc.data or ''}</font>",
        info_style,
    )
    right_table = Table([[right_title], [Spacer(1, 3 * mm)], [right_info]], colWidths=[80 * mm])
    right_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
    ]))

    header_table = Table([[left_table, right_table]], colWidths=[100 * mm, 80 * mm])
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))

    # ref / obra
    ref_para = Paragraph(f"<font color='#0a4ea1' size=12><b>Ref.: {orc.ref_cliente or '-'} </b></font>", info_style)
    obra_para = Paragraph(f"<font color='#d4111f' size=12><b>Obra: {orc.obra or '-'}</b></font>", ParagraphStyle("obra", parent=info_style, alignment=2))
    ref_table = Table([[ref_para, obra_para]], colWidths=[98 * mm, 76 * mm])
    ref_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, -1), -10 * mm),
        ("LEFTPADDING", (1, 0), (1, 0), -30),
        ("ALIGN", (1, 0), (1, 0), "LEFT"),
    ]))

    # ------------------------------------------------------------------
    # Construir a tabela de items (mas ainda NÃO a adicionamos ao story)
    # ------------------------------------------------------------------
    headers = ["Item", "Código", "Descrição", "Alt", "Larg", "Prof", "Und", "Qt", "Preço Unit", "Preço Total"]
    data = [headers]
    for item in rows:
        desc_para = Paragraph(self._format_description_pdf(item.descricao), desc_style)
        data.append([
            item.item, item.codigo, desc_para,
            _fmt_decimal(item.altura, 0), _fmt_decimal(item.largura, 0), _fmt_decimal(item.profundidade, 0),
            item.unidade, _fmt_decimal(item.qt, 2), _fmt_currency(item.preco_unitario),
            Paragraph(f"<b>{_fmt_currency(item.preco_total)}</b>", price_style),
        ])

    # paddings iniciais (vamos poder reduzir dinamicamente)
    pad_left = 3
    pad_right = 4
    pad_top = 2
    pad_bottom = 2

    def make_table_style(pl=pad_left, pr=pad_right, pt=pad_top, pb=pad_bottom):
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d7dce2")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), HEADER_TABLE_FONTSIZE),
            ("FONTSIZE", (0, 1), (-1, -1), BODY_FONTSIZE),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 1), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), pl),
            ("RIGHTPADDING", (0, 0), (-1, -1), pr),
            ("TOPPADDING", (0, 0), (-1, -1), pt),
            ("BOTTOMPADDING", (0, 0), (-1, -1), pb),
            ("ALIGN", (3, 1), (5, -1), "CENTER"),
            ("ALIGN", (7, 1), (8, -1), "RIGHT"),
            ("ALIGN", (9, 1), (9, -1), "RIGHT"),
        ])
        return style

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(make_table_style())

    # ------------------------------------------------------------------
    # Resumo / totais (SubTotal: com colon conforme pediste)
    # ------------------------------------------------------------------
    total_qt = sum((item.qt or Decimal("0")) for item in rows if not _is_separator_item(item))
    subtotal = sum((item.preco_total or Decimal("0")) for item in rows)
    iva = subtotal * IVA_RATE
    total = subtotal + iva

    # estilos para labels (right) e valores (right)
    label_right = ParagraphStyle("label_right", parent=styles["Normal"], alignment=2, fontSize=BODY_FONTSIZE, leading=DESC_LEADING)
    value_right = ParagraphStyle("value_right", parent=styles["Normal"], alignment=2, fontSize=BODY_FONTSIZE, leading=DESC_LEADING)

    # construir summary com Paragraphs — evita que HTML apareça em texto cru
    summary_col1 = 18 * mm
    summary_col2 = 30 * mm

    summary_table = Table(
        [
            [Paragraph("Total Qt.:", label_right), Paragraph(f"<b>{_fmt_decimal(total_qt, 2)}</b>", value_right)],
            [Paragraph("<font color='#0a2b6d'><b>SubTotal:</b></font>", label_right), Paragraph(f"<b>{_fmt_currency(subtotal)}</b>", value_right)],
            [Paragraph("IVA (23%):", label_right), Paragraph(f"{_fmt_currency(iva)}", value_right)],
            [Paragraph("Total Geral:", label_right), Paragraph(f"<b>{_fmt_currency(total)}</b>", value_right)],
        ],
        colWidths=[summary_col1, summary_col2],
    )

    # estilo para summary: alinhar tudo à direita, paddings reduzidos, box só no SUBTOTAL VALOR
    summary_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),  # garantir valores bem encostados à direita
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ("BOX", (1, 1), (1, 1), 1, colors.HexColor("#1d2f6f")),
                ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#dfe5f2")),
            ]
        )
    )
    summary_table.hAlign = "RIGHT"

    # Posicionar o summary exactamente à direita (push_right controla o "encaixe")
    page_width, _ = A4
    usable_width = page_width - left_margin - right_margin
    summary_width = summary_col1 + summary_col2
    left_space = usable_width - summary_width
    push_right = 12 * mm   # podes ajustar (aumenta este valor para encostar mais)
    left_space = max(0, left_space - push_right)

    container = Table([[ "", summary_table ]], colWidths=[left_space, summary_width])
    container.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 0),
                ("RIGHTPADDING", (1, 0), (1, 0), 0),
            ]
        )
    )

    # -------------------------------------------------------------------------------------
    #   Checagem de alturas: se tudo couber numa página, adicionamos; se faltar pouco,
    #   reduzimos paddings da tabela para forçar a caber numa página (evita página em branco).
    # -------------------------------------------------------------------------------------
    # obter alturas com wrap
    page_w = doc.width
    page_h = doc.height

    # calcular alturas dos blocos
    h_header = header_table.wrap(page_w, page_h)[1]
    h_ref = ref_table.wrap(page_w, page_h)[1]
    # pequeno spacer antes da tabela (6 pts)
    spacer1 = 6
    # obter altura da tabela com paddings actuais
    table_h = table.wrap(page_w, page_h)[1]
    # summary height
    summary_h = container.wrap(page_w, page_h)[1]
    total_needed = h_header + h_ref + spacer1 + table_h + spacer1 + summary_h

    # Se excede a página por poucos pontos, vamos reduzir os paddings gradualmente
    max_iterations = 6
    iter_count = 0
    # versão mais agressiva para reduzir paddings e forçar caber numa página
    max_iterations = 12
    iter_count = 0
    # reduzir em 1–2 pontos por iteração até um limite
    while total_needed > page_h and iter_count < max_iterations:
        pad_left = max(0, pad_left - 2)
        pad_right = max(0, pad_right - 2)
        pad_top = max(0, pad_top - 1)
        pad_bottom = max(0, pad_bottom - 1)
        table.setStyle(make_table_style(pl=pad_left, pr=pad_right, pt=pad_top, pb=pad_bottom))
        table_h = table.wrap(page_w, page_h)[1]
        total_needed = h_header + h_ref + spacer1 + table_h + spacer1 + summary_h
        iter_count += 1

    # Se continuar a ser maior que a página, deixamos partir em 2 páginas (cenário com muitos itens)
    # Agora construímos o story definitivo.
    story: List = []
    story.append(header_table)
    story.append(ref_table)
    story.append(Spacer(1, 6))
    story.append(table)
    story.append(Spacer(1, 6))
    story.append(container)

    # -------------------------------------------------------------------------------------
    # Construir documento com NumberedFooterCanvas (mantém rodapé com página X/Y)
    # -------------------------------------------------------------------------------------
    footer_info = {"data": orc.data or "", "numero": f"{orc.num_orcamento or ''}_{self._format_versao(orc.versao)}"}
    doc.build(story, canvasmaker=lambda *args, **kwargs: NumberedFooterCanvas(*args, footer_info=footer_info, **kwargs))
